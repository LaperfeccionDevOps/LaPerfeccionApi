from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from sqlalchemy import text


# ============================================================
# MAESTRO DE DOTACIÓN SOLICITADO POR OPERACIONES
#
# IMPORTANTE:
# - El segundo valor de cada tupla es el encabezado REAL del Excel.
# - El primer valor es una clave interna única.
# - Algunas columnas se repiten en el archivo solicitado por
#   Operaciones. Las claves internas permiten conservarlas sin
#   perder información al trabajar con diccionarios.
# ============================================================

COLUMNAS_MAESTRO_DOTACION = [
    ("compania", "compania"),
    ("Cuenta_de_Gasto", "Cuenta_de_Gasto"),
    ("desc_compania", "desc_compania"),
    ("empleado", "empleado"),
    ("pnombre", "pnombre"),
    ("snombre", "snombre"),
    ("papellido", "papellido"),
    ("sapellido", "sapellido"),
    ("fnacimiento", "fnacimiento"),
    ("tipo_empleado", "tipo_empleado"),
    ("pais", "pais"),
    ("desc_pais", "desc_pais"),
    ("departamento", "departamento"),
    ("desc_departamento", "desc_departamento"),
    ("lugar_nacimiento", "lugar_nacimiento"),
    ("direccion", "direccion"),
    ("barrio", "barrio"),
    ("telefono", "telefono"),
    ("tipo_doc_identidad", "tipo_doc_identidad"),
    ("ndoc_identidad", "ndoc_identidad"),
    ("ciudad_doc_ident", "ciudad_doc_ident"),
    ("pasaporte", "pasaporte"),
    ("otro_doc", "otro_doc"),
    ("dpto_resid", "dpto_resid"),
    ("desc_dpto_resid", "desc_dpto_resid"),
    ("ciudad_resid_empleado", "ciudad_resid_empleado"),
    ("desc_civil", "desc_civil"),
    ("nmilitar", "nmilitar"),
    ("distrito_militar", "distrito_militar"),
    ("nlic_conducir", "nlic_conducir"),
    ("njudicial", "njudicial"),
    ("nivel_estudios", "nivel_estudios"),
    ("desc_estudio", "desc_estudio"),
    ("sexo", "sexo"),
    ("sucursal", "sucursal"),
    ("desc_sucursal", "desc_sucursal"),
    ("cen1", "cen1"),
    ("desc_cen1", "desc_cen1"),
    ("cen2", "cen2"),
    ("desc_cen2", "desc_cen2"),
    ("cen3", "cen3"),
    ("desc_cen3", "desc_cen3"),
    ("cen4", "cen4"),
    ("desc_cen4", "desc_cen4"),
    ("cen5", "cen5"),
    ("desc_cen5", "desc_cen5"),
    ("escalafon", "escalafon"),
    ("desc_escalafon", "desc_escalafon"),
    ("tipo_contrato", "tipo_contrato"),
    ("desc_tipo_con", "desc_tipo_con"),
    ("regimen_contrato", "regimen_contrato"),
    ("desc_regimen_con", "desc_regimen_con"),
    ("ncontrato", "ncontrato"),
    ("fingreso", "fingreso"),
    ("fterminacion", "fterminacion"),
    ("fretiro", "fretiro"),
    ("motivo_retiro", "motivo_retiro"),
    ("descripcion", "descripcion"),
    ("activo_retirado", "activo_retirado"),
    ("entidad_salud", "entidad_salud"),
    ("desc_eps", "desc_eps"),
    ("suc_salud", "suc_salud"),
    ("entidad_pension", "entidad_pension"),
    ("desc_afp", "desc_afp"),
    ("suc_pension", "suc_pension"),
    ("entidad_riesgo", "entidad_riesgo"),
    ("desc_arp", "desc_arp"),
    ("suc_riesgo", "suc_riesgo"),
    ("caja_compensacion", "caja_compensacion"),
    ("desc_caja", "desc_caja"),
    ("fondo_cesantias", "fondo_cesantias"),
    ("desc_fondo", "desc_fondo"),
    ("centro_trabajo", "centro_trabajo"),
    ("tipo", "tipo"),
    ("desc_tipo_emp", "desc_tipo_emp"),
    ("indicador_reten", "indicador_reten"),
    ("porc_retencion", "porc_retencion"),
    ("aux_seguro", "aux_seguro"),
    ("aux_pension", "aux_pension"),
    ("aux_solidaridad", "aux_solidaridad"),
    ("fijo_variable", "fijo_variable"),
    ("ind_pension", "ind_pension"),
    ("ind_salud", "ind_salud"),
    ("ind_riesgo", "ind_riesgo"),
    ("tarifa_especial", "tarifa_especial"),
    ("porc_seguro", "porc_seguro"),
    ("turno", "turno"),
    ("numero_reg_eaab", "numero_reg_eaab"),
    ("tipo_cotizante", "tipo_cotizante"),
    ("desc_tipo_coti", "desc_tipo_coti"),
    ("subtipo_cotizante", "subtipo_cotizante"),
    ("desc_subt_coti", "desc_subt_coti"),
    ("extranjero_pension", "extranjero_pension"),
    ("reside_exterior", "reside_exterior"),
    ("activo_pensionado", "activo_pensionado"),
    ("posicion", "posicion"),
    ("empresa", "empresa"),
    ("tipo_sueldo", "tipo_sueldo"),
    ("nombre", "nombre"),
    ("forma_pago", "forma_pago"),
    ("lugar_deposito", "lugar_deposito"),
    ("desc_banco", "desc_banco"),
    ("cuenta", "cuenta"),
    ("tipo_cuenta", "tipo_cuenta"),

    # Encabezados repetidos del maestro solicitado
    ("posicion_2", "posicion"),
    ("nom_posicion", "nom_posicion"),
    ("empresa_2", "empresa"),
    ("nom_emp", "nom_emp"),
    ("nit_empresa", "nit_empresa"),
    ("cargo", "cargo"),
    ("nom_cargo", "nom_cargo"),
    ("unidad", "unidad"),
    ("nom_unidad", "nom_unidad"),
    ("escalafon_2", "escalafon"),
    ("tipo_contrato_2", "tipo_contrato"),
    ("descripcion_2", "descripcion"),
    ("email_empleado", "email_empleado"),
    ("porcentaje_riesgo_ant", "porcentaje_riesgo_ant"),
    ("porcentaje_riesgo", "porcentaje_riesgo"),
    ("Fijo_variable_2", "Fijo_variable"),
    ("indicador_sabado", "indicador_sabado"),
]


# ============================================================
# HELPERS
# ============================================================

def _texto(valor):
    if valor is None:
        return ""

    return str(valor).strip()


def _entero_si_posible(valor):
    if valor is None:
        return ""

    texto_valor = str(valor).strip()

    if not texto_valor:
        return ""

    try:
        return int(texto_valor)
    except Exception:
        return texto_valor


def _primer_valor(*valores):
    for valor in valores:
        if valor is not None and str(valor).strip() != "":
            return valor

    return ""


def _crear_mapa(rows, clave, descripcion):
    mapa = {}

    for row in rows:
        valor_clave = row.get(clave)
        valor_descripcion = row.get(descripcion)

        if valor_clave is None:
            continue

        clave_texto = str(valor_clave).strip()

        if not clave_texto:
            continue

        if (
            clave_texto not in mapa
            and valor_descripcion is not None
            and str(valor_descripcion).strip() != ""
        ):
            mapa[clave_texto] = valor_descripcion

    return mapa


def _buscar_mapa(mapa, valor, valor_defecto=""):
    if valor is None:
        return valor_defecto

    clave = str(valor).strip()

    if not clave:
        return valor_defecto

    return mapa.get(clave, valor_defecto)


# ============================================================
# CONSULTA ACTUAL A FN_REPORTESINERGY
# ============================================================

def consultar_datos_reporte_synergy(
    db,
    fecha_inicio: str,
    fecha_fin: str
):
    sql = text("""
        SELECT *
        FROM public.fn_ReporteSinergy(:fecha_inicio, :fecha_fin);
    """)

    rows = db.execute(
        sql,
        {
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
        }
    ).mappings().all()

    return rows


def normalizar_filas_reporte(rows):
    return [dict(row) for row in rows]


# ============================================================
# RETIROS FINALIZADOS
#
# fn_ReporteSinergy solo devuelve trabajadores en estado
# Contratado. Cuando una persona pasa a Retirado (35), deja de
# aparecer en esa función. Este complemento incorpora esos
# registros al maestro sin modificar la función SQL existente.
# ============================================================

def _separar_nombre_completo(valor, cantidad=2):
    partes = [
        parte
        for parte in _texto(valor).split()
        if parte
    ]

    if cantidad == 2:
        primero = partes[0] if partes else ""
        segundo = " ".join(partes[1:]) if len(partes) > 1 else ""
        return primero, segundo

    return "", ""


def _fila_maestro_vacia():
    return {
        clave_interna: ""
        for clave_interna, _
        in COLUMNAS_MAESTRO_DOTACION
    }


def _normalizar_fila_historica_a_maestro(fila_historica):
    fila_maestro = _fila_maestro_vacia()

    # Columnas sin duplicado: se copian directamente cuando existen.
    for clave_interna, _ in COLUMNAS_MAESTRO_DOTACION:
        if clave_interna in (
            "posicion_2",
            "empresa_2",
            "escalafon_2",
            "tipo_contrato_2",
            "descripcion_2",
            "Fijo_variable_2",
        ):
            continue

        if clave_interna in fila_historica:
            fila_maestro[clave_interna] = (
                fila_historica.get(clave_interna)
                if fila_historica.get(clave_interna) is not None
                else ""
            )

    # En PostgreSQL, los encabezados repetidos del Excel histórico
    # quedaron renombrados con sufijo _1.
    fila_maestro["posicion_2"] = (
        fila_historica.get("posicion_1")
        if fila_historica.get("posicion_1") is not None
        else ""
    )
    fila_maestro["empresa_2"] = (
        fila_historica.get("empresa_1")
        if fila_historica.get("empresa_1") is not None
        else ""
    )
    fila_maestro["escalafon_2"] = (
        fila_historica.get("escalafon_1")
        if fila_historica.get("escalafon_1") is not None
        else ""
    )
    fila_maestro["tipo_contrato_2"] = (
        fila_historica.get("tipo_contrato_1")
        if fila_historica.get("tipo_contrato_1") is not None
        else ""
    )
    fila_maestro["descripcion_2"] = (
        fila_historica.get("descripcion_1")
        if fila_historica.get("descripcion_1") is not None
        else ""
    )
    fila_maestro["Fijo_variable_2"] = (
        fila_historica.get("Fijo_variable_1")
        if fila_historica.get("Fijo_variable_1") is not None
        else ""
    )

    return fila_maestro


def _obtener_retiros_finalizados_maestro(db):
    sql_retiros = text("""
        SELECT
            rl."IdRetiroLaboral",
            rl."IdRegistroPersonal",
            rl."FechaRetiro",
            rl."IdMotivoRetiro",

            mr."Nombre" AS "NombreMotivoRetiro",

            rp."NumeroIdentificacion",
            rp."Nombres",
            rp."Apellidos",
            rp."Email",

            cb."Posicion",
            cb."Escalafon",
            cb."NumeroCuenta",
            cb."IdTipoContrato",

            banco."CodigoBanco",
            banco."DescripcionBanco",

            acc."IdCargo",
            cg."NombreCargo",

            (
                SELECT MAX(hl."FechaIngreso")
                FROM public."HistorialLaboral" hl
                WHERE hl."IdRegistroPersonal" =
                      rp."IdRegistroPersonal"
            ) AS "FechaIngreso"

        FROM public."RetiroLaboral" rl

        INNER JOIN public."RegistroPersonal" rp
            ON rp."IdRegistroPersonal" =
               rl."IdRegistroPersonal"

        LEFT JOIN public."MotivoRetiro" mr
            ON mr."IdMotivoRetiro" =
               rl."IdMotivoRetiro"

        LEFT JOIN public."ContratacionBasica" cb
            ON cb."IdRegistroPersonal" =
               rp."IdRegistroPersonal"

        LEFT JOIN public."Banco" banco
            ON banco."IdBanco" = cb."IdBanco"

        LEFT JOIN public."AsignacionCargoCliente" acc
            ON acc."IdRegistroPersonal" =
               rp."IdRegistroPersonal"

        LEFT JOIN public."Cargo" cg
            ON cg."IdCargo" = acc."IdCargo"

        WHERE rp."IdEstadoProceso" = 35

        ORDER BY
            rl."IdRegistroPersonal",
            rl."IdRetiroLaboral" DESC;
    """)

    try:
        rows_retiros = [
            dict(row)
            for row in db.execute(
                sql_retiros
            ).mappings().all()
        ]
    except Exception:
        return []

    # Un solo retiro por trabajador: se conserva el más reciente.
    retiros_por_documento = {}

    for row in rows_retiros:
        documento = row.get("NumeroIdentificacion")

        if documento is None:
            continue

        documento_key = str(documento).strip()

        if (
            documento_key
            and documento_key not in retiros_por_documento
        ):
            retiros_por_documento[documento_key] = row

    if not retiros_por_documento:
        return []

    documentos = list(retiros_por_documento.keys())

    # Si el trabajador existe en el maestro histórico, se usa esa
    # fila únicamente como base estructural para no perder datos que
    # no viven actualmente en tablas del aplicativo.
    sql_historico = text("""
        SELECT *
        FROM public.migracionactivossynergy
        WHERE CAST(ndoc_identidad AS TEXT) = ANY(:documentos);
    """)

    try:
        rows_historico = [
            dict(row)
            for row in db.execute(
                sql_historico,
                {"documentos": documentos}
            ).mappings().all()
        ]
    except Exception:
        rows_historico = []

    historico_por_documento = {}

    for row in rows_historico:
        documento = row.get("ndoc_identidad")

        if documento is None:
            continue

        documento_key = str(documento).strip()

        if (
            documento_key
            and documento_key not in historico_por_documento
        ):
            historico_por_documento[documento_key] = row

    resultado = []

    for documento_key, retiro in retiros_por_documento.items():
        historico = historico_por_documento.get(
            documento_key
        )

        if historico:
            fila_maestro = (
                _normalizar_fila_historica_a_maestro(
                    historico
                )
            )
        else:
            fila_maestro = _fila_maestro_vacia()

            nombres_1, nombres_2 = (
                _separar_nombre_completo(
                    retiro.get("Nombres")
                )
            )
            apellidos_1, apellidos_2 = (
                _separar_nombre_completo(
                    retiro.get("Apellidos")
                )
            )

            fila_maestro.update({
                "compania": 1,
                "desc_compania": (
                    "Aseos La Perfección SAS"
                ),
                "empleado": retiro.get(
                    "NumeroIdentificacion"
                ) or "",
                "pnombre": nombres_1,
                "snombre": nombres_2,
                "papellido": apellidos_1,
                "sapellido": apellidos_2,
                "ndoc_identidad": retiro.get(
                    "NumeroIdentificacion"
                ) or "",
                "sucursal": 1,
                "desc_sucursal": (
                    "ASEOS LA PERFECCIÓN SAS"
                ),
                "tipo": 1,
                "desc_tipo_emp": "Mensuales",
                "empresa": 1,
                "empresa_2": 1,
                "nom_emp": (
                    "Aseos La Perfección SAS"
                ),
                "nit_empresa": "800068462-4",
                "tipo_sueldo": 1,
                "nombre": "Salario interno pesos",
                "forma_pago": "CO",
                "extranjero_pension": "N",
                "reside_exterior": "N",
                "activo_pensionado": "A",
            })

        # Se priorizan los datos actuales del aplicativo cuando
        # existen. Esto evita reemplazar información vigente por
        # datos históricos del trabajador.
        posicion = retiro.get("Posicion")

        if posicion is not None and str(posicion).strip():
            fila_maestro["posicion"] = (
                _entero_si_posible(posicion)
            )
            fila_maestro["posicion_2"] = (
                _entero_si_posible(posicion)
            )

        escalafon = retiro.get("Escalafon")

        if escalafon is not None and str(escalafon).strip():
            fila_maestro["escalafon"] = (
                _entero_si_posible(escalafon)
            )
            fila_maestro["escalafon_2"] = (
                _entero_si_posible(escalafon)
            )

        tipo_contrato = retiro.get("IdTipoContrato")

        if (
            tipo_contrato is not None
            and str(tipo_contrato).strip()
        ):
            fila_maestro["tipo_contrato"] = (
                _entero_si_posible(tipo_contrato)
            )
            fila_maestro["tipo_contrato_2"] = (
                _entero_si_posible(tipo_contrato)
            )

        numero_cuenta = retiro.get("NumeroCuenta")

        if (
            numero_cuenta is not None
            and str(numero_cuenta).strip()
        ):
            fila_maestro["cuenta"] = numero_cuenta

        codigo_banco = retiro.get("CodigoBanco")

        if (
            codigo_banco is not None
            and str(codigo_banco).strip()
        ):
            fila_maestro["lugar_deposito"] = (
                _entero_si_posible(codigo_banco)
            )

        descripcion_banco = retiro.get(
            "DescripcionBanco"
        )

        if descripcion_banco:
            fila_maestro["desc_banco"] = (
                descripcion_banco
            )

        id_cargo = retiro.get("IdCargo")

        if id_cargo is not None and str(id_cargo).strip():
            fila_maestro["cargo"] = (
                _entero_si_posible(id_cargo)
            )

        nombre_cargo = retiro.get("NombreCargo")

        if nombre_cargo:
            fila_maestro["nom_cargo"] = nombre_cargo

        email = retiro.get("Email")

        if email:
            fila_maestro["email_empleado"] = email

        fecha_ingreso = retiro.get("FechaIngreso")

        if fecha_ingreso:
            fila_maestro["fingreso"] = fecha_ingreso

        # Campos de retiro: estos SIEMPRE provienen del flujo actual
        # de RetiroLaboral y prevalecen sobre cualquier histórico.
        fila_maestro["activo_retirado"] = "R"
        fila_maestro["fretiro"] = (
            retiro.get("FechaRetiro") or ""
        )
        fila_maestro["motivo_retiro"] = (
            _entero_si_posible(
                retiro.get("IdMotivoRetiro")
            )
        )
        fila_maestro["descripcion"] = (
            retiro.get("NombreMotivoRetiro")
            or ""
        )

        resultado.append(fila_maestro)

    return resultado


def _mezclar_activos_y_retirados(
    filas_activas,
    filas_retiradas
):
    mapa = {}
    orden = []

    for fila in filas_activas:
        documento = _primer_valor(
            fila.get("ndoc_identidad"),
            fila.get("empleado")
        )

        documento_key = (
            str(documento).strip()
            if documento is not None
            else ""
        )

        if documento_key:
            mapa[documento_key] = dict(fila)
            orden.append(documento_key)
        else:
            clave_temporal = (
                f"__sin_documento_activo_{len(orden)}"
            )
            mapa[clave_temporal] = dict(fila)
            orden.append(clave_temporal)

    for fila in filas_retiradas:
        documento = _primer_valor(
            fila.get("ndoc_identidad"),
            fila.get("empleado")
        )

        documento_key = (
            str(documento).strip()
            if documento is not None
            else ""
        )

        if not documento_key:
            continue

        if documento_key not in mapa:
            orden.append(documento_key)

        # Si el mismo documento llegara por ambos caminos,
        # el estado retirado prevalece.
        mapa[documento_key] = dict(fila)

    return [
        mapa[clave]
        for clave in orden
    ]


# ============================================================
# ENRIQUECIMIENTO ACTUAL DE CARGO
#
# Se conserva para no dañar el flujo existente.
# ============================================================

def enriquecer_filas_para_sheet_con_cargo(db, filas):
    if not filas:
        return filas

    ids_registro = []
    documentos = []

    for fila in filas:
        id_reg = (
            fila.get("idregistropersonal")
            or fila.get("IdRegistroPersonal")
        )

        if id_reg is not None:
            try:
                ids_registro.append(int(id_reg))
            except Exception:
                pass

        num_doc = (
            fila.get("num_doc_id")
            or fila.get("empleado")
            or fila.get("NumeroIdentificacion")
            or fila.get("ndoc_identidad")
        )

        if num_doc is not None and str(num_doc).strip() != "":
            documentos.append(str(num_doc).strip())

    ids_registro = list(set(ids_registro))
    documentos = list(set(documentos))

    mapa_cargo_por_id = {}
    mapa_cargo_por_doc = {}

    if ids_registro:
        sql_ids = text("""
            SELECT
                acc."IdRegistroPersonal",
                cg."NombreCargo"
            FROM public."AsignacionCargoCliente" acc
            LEFT JOIN public."Cargo" cg
                ON cg."IdCargo" = acc."IdCargo"
            WHERE acc."IdRegistroPersonal" = ANY(:ids_registro)
        """)

        rows_ids = db.execute(
            sql_ids,
            {"ids_registro": ids_registro}
        ).mappings().all()

        for row in rows_ids:
            try:
                id_registro = int(row["IdRegistroPersonal"])
                nombre_cargo = row.get("NombreCargo")

                if (
                    id_registro not in mapa_cargo_por_id
                    and nombre_cargo
                ):
                    mapa_cargo_por_id[id_registro] = nombre_cargo

            except Exception:
                pass

    if documentos:
        sql_docs = text("""
            SELECT
                rp."NumeroIdentificacion",
                acc."IdRegistroPersonal",
                cg."NombreCargo"
            FROM public."RegistroPersonal" rp
            LEFT JOIN public."AsignacionCargoCliente" acc
                ON acc."IdRegistroPersonal" = rp."IdRegistroPersonal"
            LEFT JOIN public."Cargo" cg
                ON cg."IdCargo" = acc."IdCargo"
            WHERE CAST(rp."NumeroIdentificacion" AS TEXT)
                = ANY(:documentos)
        """)

        rows_docs = db.execute(
            sql_docs,
            {"documentos": documentos}
        ).mappings().all()

        for row in rows_docs:
            doc = row.get("NumeroIdentificacion")
            cargo = row.get("NombreCargo")

            if (
                doc is not None
                and cargo
                and str(doc).strip() not in mapa_cargo_por_doc
            ):
                mapa_cargo_por_doc[str(doc).strip()] = cargo

    filas_enriquecidas = []

    for fila in filas:
        nueva = dict(fila)

        cargo_texto = None

        id_reg = (
            nueva.get("idregistropersonal")
            or nueva.get("IdRegistroPersonal")
        )

        if id_reg is not None:
            try:
                cargo_texto = mapa_cargo_por_id.get(int(id_reg))
            except Exception:
                pass

        if not cargo_texto:
            num_doc = (
                nueva.get("num_doc_id")
                or nueva.get("empleado")
                or nueva.get("NumeroIdentificacion")
                or nueva.get("ndoc_identidad")
            )

            if num_doc is not None:
                cargo_texto = mapa_cargo_por_doc.get(
                    str(num_doc).strip()
                )

        # Si todavía estamos trabajando con el formato antiguo,
        # se conserva exactamente el comportamiento anterior.
        if cargo_texto and "nom_cargo" not in nueva:
            nueva["cargo"] = cargo_texto

        # Si ya es una fila del maestro nuevo, no se pisa el código
        # numérico del cargo. Se actualiza únicamente el nombre.
        elif cargo_texto:
            nueva["nom_cargo"] = cargo_texto

        filas_enriquecidas.append(nueva)

    return filas_enriquecidas


# ============================================================
# CATÁLOGOS HISTÓRICOS SYNERGY
#
# La tabla histórica NO se modifica.
# Se usa únicamente como catálogo de referencia para completar
# las descripciones que el maestro nuevo solicita.
# ============================================================

def _obtener_catalogos_historicos(db):
    sql = text("""
        SELECT DISTINCT
            pais,
            desc_pais,

            departamento,
            desc_departamento,

            dpto_resid,
            desc_dpto_resid,

            nivel_estudios,
            desc_estudio,

            sucursal,
            desc_sucursal,

            cen1,
            desc_cen1,

            cen2,
            desc_cen2,

            cen3,
            desc_cen3,

            cen4,
            desc_cen4,

            cen5,
            desc_cen5,

            escalafon,
            desc_escalafon,

            tipo_contrato,
            desc_tipo_con,

            regimen_contrato,
            desc_regimen_con,

            entidad_salud,
            desc_eps,

            entidad_pension,
            desc_afp,

            entidad_riesgo,
            desc_arp,

            caja_compensacion,
            desc_caja,

            fondo_cesantias,
            desc_fondo,

            tipo,
            desc_tipo_emp,

            tipo_cotizante,
            desc_tipo_coti,

            subtipo_cotizante,
            desc_subt_coti,

            tipo_sueldo,
            nombre,

            lugar_deposito,
            desc_banco,

            posicion,
            posicion_1,
            nom_posicion,

            empresa,
            empresa_1,
            nom_emp,
            nit_empresa,

            cargo,
            nom_cargo,

            unidad,
            nom_unidad,

            centro_trabajo,
            porcentaje_riesgo_ant,
            porcentaje_riesgo,

            "Fijo_variable_1",
            indicador_sabado

        FROM public.migracionactivossynergy;
    """)

    try:
        rows = [
            dict(row)
            for row in db.execute(sql).mappings().all()
        ]
    except Exception:
        # La generación del archivo no debe caerse solamente porque
        # un catálogo histórico no esté disponible.
        rows = []

    catalogos = {
        "desc_pais": _crear_mapa(
            rows,
            "pais",
            "desc_pais"
        ),
        "desc_departamento": _crear_mapa(
            rows,
            "departamento",
            "desc_departamento"
        ),
        "desc_dpto_resid": _crear_mapa(
            rows,
            "dpto_resid",
            "desc_dpto_resid"
        ),
        "desc_estudio": _crear_mapa(
            rows,
            "nivel_estudios",
            "desc_estudio"
        ),
        "desc_sucursal": _crear_mapa(
            rows,
            "sucursal",
            "desc_sucursal"
        ),
        "desc_cen1": _crear_mapa(
            rows,
            "cen1",
            "desc_cen1"
        ),
        "desc_cen2": _crear_mapa(
            rows,
            "cen2",
            "desc_cen2"
        ),
        "desc_cen3": _crear_mapa(
            rows,
            "cen3",
            "desc_cen3"
        ),
        "desc_cen4": _crear_mapa(
            rows,
            "cen4",
            "desc_cen4"
        ),
        "desc_cen5": _crear_mapa(
            rows,
            "cen5",
            "desc_cen5"
        ),
        "desc_escalafon": _crear_mapa(
            rows,
            "escalafon",
            "desc_escalafon"
        ),
        "desc_tipo_con": _crear_mapa(
            rows,
            "tipo_contrato",
            "desc_tipo_con"
        ),
        "desc_regimen_con": _crear_mapa(
            rows,
            "regimen_contrato",
            "desc_regimen_con"
        ),
        "desc_eps": _crear_mapa(
            rows,
            "entidad_salud",
            "desc_eps"
        ),
        "desc_afp": _crear_mapa(
            rows,
            "entidad_pension",
            "desc_afp"
        ),
        "desc_arp": _crear_mapa(
            rows,
            "entidad_riesgo",
            "desc_arp"
        ),
        "desc_caja": _crear_mapa(
            rows,
            "caja_compensacion",
            "desc_caja"
        ),
        "desc_fondo": _crear_mapa(
            rows,
            "fondo_cesantias",
            "desc_fondo"
        ),
        "desc_tipo_emp": _crear_mapa(
            rows,
            "tipo",
            "desc_tipo_emp"
        ),
        "desc_tipo_coti": _crear_mapa(
            rows,
            "tipo_cotizante",
            "desc_tipo_coti"
        ),
        "desc_subt_coti": _crear_mapa(
            rows,
            "subtipo_cotizante",
            "desc_subt_coti"
        ),
        "nombre_tipo_sueldo": _crear_mapa(
            rows,
            "tipo_sueldo",
            "nombre"
        ),
        "desc_banco": _crear_mapa(
            rows,
            "lugar_deposito",
            "desc_banco"
        ),
        "nom_posicion": {},
        "unidad_por_posicion": {},
        "nom_unidad_por_posicion": {},
        "nom_cargo": _crear_mapa(
            rows,
            "cargo",
            "nom_cargo"
        ),
        "empresa_nombre": {},
        "empresa_nit": {},
        "riesgo_anterior": {},
        "riesgo_actual": {},
        "fijo_variable_2": {},
        "indicador_sabado": {},
    }

    for row in rows:
        posicion = _primer_valor(
            row.get("posicion_1"),
            row.get("posicion")
        )

        if posicion is not None and str(posicion).strip():
            posicion_key = str(posicion).strip()

            if (
                posicion_key not in catalogos["nom_posicion"]
                and row.get("nom_posicion")
            ):
                catalogos["nom_posicion"][posicion_key] = (
                    row.get("nom_posicion")
                )

            if (
                posicion_key not in catalogos["unidad_por_posicion"]
                and row.get("unidad") is not None
            ):
                catalogos["unidad_por_posicion"][posicion_key] = (
                    row.get("unidad")
                )

            if (
                posicion_key
                not in catalogos["nom_unidad_por_posicion"]
                and row.get("nom_unidad")
            ):
                catalogos["nom_unidad_por_posicion"][
                    posicion_key
                ] = row.get("nom_unidad")

            if (
                posicion_key not in catalogos["fijo_variable_2"]
                and row.get("Fijo_variable_1")
            ):
                catalogos["fijo_variable_2"][posicion_key] = (
                    row.get("Fijo_variable_1")
                )

            if (
                posicion_key not in catalogos["indicador_sabado"]
                and row.get("indicador_sabado")
            ):
                catalogos["indicador_sabado"][posicion_key] = (
                    row.get("indicador_sabado")
                )

        empresa = _primer_valor(
            row.get("empresa_1"),
            row.get("empresa")
        )

        if empresa is not None and str(empresa).strip():
            empresa_key = str(empresa).strip()

            if (
                empresa_key not in catalogos["empresa_nombre"]
                and row.get("nom_emp")
            ):
                catalogos["empresa_nombre"][empresa_key] = (
                    row.get("nom_emp")
                )

            if (
                empresa_key not in catalogos["empresa_nit"]
                and row.get("nit_empresa")
            ):
                catalogos["empresa_nit"][empresa_key] = (
                    row.get("nit_empresa")
                )

        centro_trabajo = row.get("centro_trabajo")

        if (
            centro_trabajo is not None
            and str(centro_trabajo).strip()
        ):
            centro_key = str(centro_trabajo).strip()

            if (
                centro_key not in catalogos["riesgo_anterior"]
                and row.get("porcentaje_riesgo_ant")
            ):
                catalogos["riesgo_anterior"][centro_key] = (
                    row.get("porcentaje_riesgo_ant")
                )

            if (
                centro_key not in catalogos["riesgo_actual"]
                and row.get("porcentaje_riesgo")
            ):
                catalogos["riesgo_actual"][centro_key] = (
                    row.get("porcentaje_riesgo")
                )

    return catalogos


# ============================================================
# DATOS ACTUALES COMPLEMENTARIOS
#
# Se usan los datos actuales del aplicativo.
# No se reemplazan por datos históricos del trabajador.
# ============================================================

def _obtener_datos_actuales_por_documento(db, filas):
    documentos = []

    for fila in filas:
        documento = _primer_valor(
            fila.get("num_doc_id"),
            fila.get("empleado"),
            fila.get("NumeroIdentificacion"),
            fila.get("ndoc_identidad")
        )

        if documento is not None and str(documento).strip():
            documentos.append(str(documento).strip())

    documentos = list(set(documentos))

    if not documentos:
        return {}

    sql = text("""
        SELECT
            rp."NumeroIdentificacion",
            acc."IdCargo",
            cg."NombreCargo",
            cb."Posicion",
            cb."Escalafon",
            cb."NumeroCuenta",
            cb."IdTipoContrato",
            banco."CodigoBanco",
            banco."DescripcionBanco"

        FROM public."RegistroPersonal" rp

        LEFT JOIN public."ContratacionBasica" cb
            ON cb."IdRegistroPersonal" =
               rp."IdRegistroPersonal"

        LEFT JOIN public."Banco" banco
            ON banco."IdBanco" = cb."IdBanco"

        LEFT JOIN public."AsignacionCargoCliente" acc
            ON acc."IdRegistroPersonal" =
               rp."IdRegistroPersonal"

        LEFT JOIN public."Cargo" cg
            ON cg."IdCargo" = acc."IdCargo"

        WHERE CAST(rp."NumeroIdentificacion" AS TEXT)
            = ANY(:documentos);
    """)

    try:
        rows = db.execute(
            sql,
            {"documentos": documentos}
        ).mappings().all()
    except Exception:
        return {}

    mapa = {}

    for row in rows:
        documento = row.get("NumeroIdentificacion")

        if documento is None:
            continue

        documento_key = str(documento).strip()

        if documento_key not in mapa:
            mapa[documento_key] = dict(row)
            continue

        # Si ya existía una fila y esta nueva trae cargo,
        # se prefiere la que tiene información más completa.
        anterior = mapa[documento_key]

        if (
            not anterior.get("IdCargo")
            and row.get("IdCargo")
        ):
            mapa[documento_key] = dict(row)

    return mapa


# ============================================================
# ADAPTACIÓN AL MAESTRO NUEVO
# ============================================================

def adaptar_filas_maestro_nuevo(db, filas):
    if not filas:
        return []

    catalogos = _obtener_catalogos_historicos(db)
    datos_actuales = _obtener_datos_actuales_por_documento(
        db,
        filas
    )

    resultado = []

    for fila_original in filas:
        fila = dict(fila_original)

        documento = _primer_valor(
            fila.get("num_doc_id"),
            fila.get("empleado"),
            fila.get("NumeroIdentificacion"),
            fila.get("ndoc_identidad")
        )

        documento_key = (
            str(documento).strip()
            if documento is not None
            else ""
        )

        actual = datos_actuales.get(documento_key, {})

        pais = _primer_valor(
            fila.get("pais"),
            169
        )

        departamento = _primer_valor(
            fila.get("departamento"),
            11
        )

        dpto_resid = _primer_valor(
            fila.get("depto_residencia"),
            departamento
        )

        sucursal = _primer_valor(
            fila.get("sucursal"),
            1
        )

        cen1 = fila.get("centro_costos_1") or ""
        cen2 = fila.get("centro_costos_2") or ""
        cen3 = fila.get("centro_costos_3") or ""
        cen4 = fila.get("centro_costos_4") or ""
        cen5 = fila.get("centro_costos_5") or ""

        escalafon = _primer_valor(
            actual.get("Escalafon"),
            fila.get("escalafon")
        )

        tipo_contrato = _primer_valor(
            actual.get("IdTipoContrato"),
            fila.get("tipo_contrato")
        )

        regimen_contrato = _primer_valor(
            fila.get("regimen"),
            2
        )

        entidad_salud = fila.get("entidad_salud") or ""
        entidad_pension = fila.get("entidad_pension") or ""
        entidad_riesgo = fila.get("entidad_riesgo") or ""
        caja_compensacion = (
            fila.get("caja_compensacion")
            or "CCF24"
        )
        fondo_cesantias = (
            fila.get("fondo_cesantias")
            or ""
        )

        tipo_cotizante = _primer_valor(
            fila.get("tipo_cotizante"),
            1
        )

        subtipo_cotizante = _primer_valor(
            fila.get("subtipo_cotizante"),
            0
        )

        posicion = _primer_valor(
            actual.get("Posicion"),
            fila.get("posicion")
        )

        empresa = _primer_valor(
            fila.get("empresa"),
            1
        )

        tipo_sueldo = _primer_valor(
            fila.get("tipo_sueldo"),
            1
        )

        codigo_banco = _primer_valor(
            actual.get("CodigoBanco"),
            fila.get("corporacion")
        )

        descripcion_banco = _primer_valor(
            actual.get("DescripcionBanco"),
            _buscar_mapa(
                catalogos["desc_banco"],
                codigo_banco
            )
        )

        id_cargo = actual.get("IdCargo") or ""
        nombre_cargo = _primer_valor(
            actual.get("NombreCargo"),
            _buscar_mapa(
                catalogos["nom_cargo"],
                id_cargo
            ),
            fila.get("cargo")
        )

        centro_trabajo = fila.get("centro_trabajo") or ""

        unidad = _buscar_mapa(
            catalogos["unidad_por_posicion"],
            posicion
        )

        nom_unidad = _buscar_mapa(
            catalogos["nom_unidad_por_posicion"],
            posicion
        )

        fijo_variable = ""

        tipo_sueldo_empleado = _texto(
            fila.get("tipo_sueldo_empleado")
        ).upper()

        if tipo_sueldo_empleado == "F":
            fijo_variable = "fijo"
        elif tipo_sueldo_empleado == "V":
            fijo_variable = "variable"

        fijo_variable_2 = _buscar_mapa(
            catalogos["fijo_variable_2"],
            posicion
        )

        if not fijo_variable_2:
            if tipo_sueldo_empleado == "F":
                fijo_variable_2 = "sueldo fijo"
            elif tipo_sueldo_empleado == "V":
                fijo_variable_2 = "sueldo variable"

        indicador_sabado = _primer_valor(
            fila.get("trabaja_sabado"),
            _buscar_mapa(
                catalogos["indicador_sabado"],
                posicion
            )
        )

        activo_retirado = _primer_valor(
            fila.get("estado"),
            "A"
        )

        # Para trabajadores activos, los datos de retiro deben quedar vacíos.
        # fn_ReporteSinergy actualmente devuelve "0" como motivo de retiro,
        # pero ese valor no corresponde al maestro solicitado por Operaciones.
        if str(activo_retirado).strip().upper() == "A":
            fecha_retiro = ""
            motivo_retiro = ""
            descripcion_retiro = ""
        else:
            fecha_retiro = fila.get("fecha_retiro") or ""
            motivo_retiro = _entero_si_posible(
                fila.get("motivo_retiro")
            )
            descripcion_retiro = fila.get("descripcion") or ""

        fila_maestro = {
            "compania": 1,
            "Cuenta_de_Gasto": fila.get("cuenta_gasto") or "",
            "desc_compania": "Aseos La Perfección SAS",

            "empleado": fila.get("empleado") or documento,
            "pnombre": fila.get("pnombre") or "",
            "snombre": fila.get("snombre") or "",
            "papellido": fila.get("papellido") or "",
            "sapellido": fila.get("sapellido") or "",

            "fnacimiento": fila.get("fecha_nacimiento") or "",

            "tipo_empleado": _entero_si_posible(
                fila.get("tipo_empleado")
            ),

            "pais": _entero_si_posible(pais),
            "desc_pais": _buscar_mapa(
                catalogos["desc_pais"],
                pais,
                "169 - COLOMBIA"
                if str(pais).strip() == "169"
                else ""
            ),

            "departamento": _entero_si_posible(
                departamento
            ),

            "desc_departamento": _buscar_mapa(
                catalogos["desc_departamento"],
                departamento
            ),

            "lugar_nacimiento": (
                fila.get("lugar_nacimiento")
                or ""
            ),

            "direccion": fila.get("direccion") or "",
            "barrio": fila.get("barrio") or "",
            "telefono": fila.get("telefono") or "",

            "tipo_doc_identidad": (
                fila.get("tipo_doc_id")
                or ""
            ),

            "ndoc_identidad": (
                fila.get("num_doc_id")
                or documento
            ),

            "ciudad_doc_ident": (
                fila.get("ciudad_doc_id")
                or ""
            ),

            "pasaporte": fila.get("pasaporte") or "",

            "otro_doc": "",

            "dpto_resid": _entero_si_posible(
                dpto_resid
            ),

            "desc_dpto_resid": _buscar_mapa(
                catalogos["desc_dpto_resid"],
                dpto_resid
            ),

            "ciudad_resid_empleado": (
                fila.get("municipio_resid")
                or ""
            ),

            "desc_civil": (
                fila.get("estado_civil")
                or ""
            ),

            "nmilitar": (
                fila.get("libreta_militar")
                or ""
            ),

            "distrito_militar": "",

            "nlic_conducir": (
                fila.get("licencia_conducir")
                or ""
            ),

            "njudicial": (
                fila.get("certif_juducial")
                or ""
            ),

            "nivel_estudios": _entero_si_posible(
                fila.get("nivel_estudios")
            ),

            "desc_estudio": _buscar_mapa(
                catalogos["desc_estudio"],
                fila.get("nivel_estudios")
            ),

            "sexo": fila.get("sexo") or "",

            "sucursal": _entero_si_posible(
                sucursal
            ),

            "desc_sucursal": _buscar_mapa(
                catalogos["desc_sucursal"],
                sucursal,
                "ASEOS LA PERFECCIÓN SAS"
            ),

            "cen1": cen1,
            "desc_cen1": _buscar_mapa(
                catalogos["desc_cen1"],
                cen1
            ),

            "cen2": cen2,
            "desc_cen2": _buscar_mapa(
                catalogos["desc_cen2"],
                cen2
            ),

            "cen3": cen3,
            "desc_cen3": _buscar_mapa(
                catalogos["desc_cen3"],
                cen3
            ),

            "cen4": cen4,
            "desc_cen4": _buscar_mapa(
                catalogos["desc_cen4"],
                cen4
            ),

            "cen5": cen5,
            "desc_cen5": _buscar_mapa(
                catalogos["desc_cen5"],
                cen5
            ),

            "escalafon": _entero_si_posible(
                escalafon
            ),

            "desc_escalafon": _buscar_mapa(
                catalogos["desc_escalafon"],
                escalafon
            ),

            "tipo_contrato": _entero_si_posible(
                tipo_contrato
            ),

            "desc_tipo_con": _buscar_mapa(
                catalogos["desc_tipo_con"],
                tipo_contrato
            ),

            "regimen_contrato": _entero_si_posible(
                regimen_contrato
            ),

            "desc_regimen_con": _buscar_mapa(
                catalogos["desc_regimen_con"],
                regimen_contrato
            ),

            "ncontrato": _entero_si_posible(
                fila.get("ncontrato")
            ),

            "fingreso": (
                fila.get("fecha_ingreso")
                or ""
            ),

            "fterminacion": (
                fila.get("fecha_terminacion")
                or ""
            ),

            "fretiro": fecha_retiro,

            "motivo_retiro": motivo_retiro,

            "descripcion": descripcion_retiro,

            "activo_retirado": activo_retirado,

            "entidad_salud": entidad_salud,

            "desc_eps": _buscar_mapa(
                catalogos["desc_eps"],
                entidad_salud
            ),

            "suc_salud": _entero_si_posible(
                fila.get("sucursal_salud")
            ),

            "entidad_pension": entidad_pension,

            "desc_afp": _buscar_mapa(
                catalogos["desc_afp"],
                entidad_pension
            ),

            "suc_pension": _entero_si_posible(
                fila.get("sucursal_pension")
            ),

            "entidad_riesgo": entidad_riesgo,

            "desc_arp": _buscar_mapa(
                catalogos["desc_arp"],
                entidad_riesgo
            ),

            "suc_riesgo": _entero_si_posible(
                fila.get("sucur_Ent_riesgo")
            ),

            "caja_compensacion": (
                caja_compensacion
            ),

            "desc_caja": _buscar_mapa(
                catalogos["desc_caja"],
                caja_compensacion,
                "COMPENSAR"
                if caja_compensacion == "CCF24"
                else ""
            ),

            "fondo_cesantias": fondo_cesantias,

            "desc_fondo": _buscar_mapa(
                catalogos["desc_fondo"],
                fondo_cesantias
            ),

            "centro_trabajo": centro_trabajo,

            "tipo": 1,

            "desc_tipo_emp": _buscar_mapa(
                catalogos["desc_tipo_emp"],
                1,
                "Mensuales"
            ),

            "indicador_reten": (
                fila.get("indicador_retencion")
                or ""
            ),

            "porc_retencion": (
                fila.get("porcentaje_retencion")
                or ""
            ),

            "aux_seguro": (
                fila.get("auxilio_seguro")
                or ""
            ),

            "aux_pension": (
                fila.get("auxilio_pension")
                or ""
            ),

            "aux_solidaridad": (
                fila.get("auxilio_solidaridad")
                or ""
            ),

            "fijo_variable": fijo_variable,

            "ind_pension": (
                "aporta"
                if entidad_pension
                else ""
            ),

            "ind_salud": (
                "aporta"
                if entidad_salud
                else ""
            ),

            "ind_riesgo": (
                "aporta"
                if entidad_riesgo
                else ""
            ),

            "tarifa_especial": _entero_si_posible(
                fila.get("tarifa_especial")
            ),

            "porc_seguro": _entero_si_posible(
                fila.get("porcentaje_seguro")
            ),

            "turno": "",
            "numero_reg_eaab": "",

            "tipo_cotizante": _entero_si_posible(
                tipo_cotizante
            ),

            "desc_tipo_coti": _buscar_mapa(
                catalogos["desc_tipo_coti"],
                tipo_cotizante
            ),

            "subtipo_cotizante": _entero_si_posible(
                subtipo_cotizante
            ),

            "desc_subt_coti": _buscar_mapa(
                catalogos["desc_subt_coti"],
                subtipo_cotizante
            ),

            "extranjero_pension": (
                fila.get("extranjero_pension")
                or "N"
            ),

            "reside_exterior": (
                fila.get("reside_exterior")
                or "N"
            ),

            "activo_pensionado": (
                fila.get("activo_pensionado")
                or "A"
            ),

            "posicion": _entero_si_posible(
                posicion
            ),

            "empresa": _entero_si_posible(
                empresa
            ),

            "tipo_sueldo": _entero_si_posible(
                tipo_sueldo
            ),

            "nombre": _buscar_mapa(
                catalogos["nombre_tipo_sueldo"],
                tipo_sueldo,
                "Salario interno pesos"
            ),

            "forma_pago": (
                fila.get("tipo_pago")
                or "CO"
            ),

            "lugar_deposito": _entero_si_posible(
                codigo_banco
            ),

            "desc_banco": descripcion_banco,

            "cuenta": _primer_valor(
                actual.get("NumeroCuenta"),
                fila.get("cuenta")
            ),

            "tipo_cuenta": _entero_si_posible(
                fila.get("tipo_cuenta")
            ),

            "posicion_2": _entero_si_posible(
                posicion
            ),

            "nom_posicion": _buscar_mapa(
                catalogos["nom_posicion"],
                posicion
            ),

            "empresa_2": _entero_si_posible(
                empresa
            ),

            "nom_emp": _buscar_mapa(
                catalogos["empresa_nombre"],
                empresa,
                "Aseos La Perfección SAS"
            ),

            "nit_empresa": _buscar_mapa(
                catalogos["empresa_nit"],
                empresa,
                "800068462-4"
            ),

            "cargo": _entero_si_posible(
                id_cargo
            ),

            "nom_cargo": nombre_cargo,

            "unidad": _entero_si_posible(
                unidad
            ),

            "nom_unidad": nom_unidad,

            "escalafon_2": _entero_si_posible(
                escalafon
            ),

            "tipo_contrato_2": _entero_si_posible(
                tipo_contrato
            ),

            "descripcion_2": _buscar_mapa(
                catalogos["desc_tipo_con"],
                tipo_contrato
            ),

            "email_empleado": (
                fila.get("email")
                or ""
            ),

            "porcentaje_riesgo_ant": _buscar_mapa(
                catalogos["riesgo_anterior"],
                centro_trabajo
            ),

            "porcentaje_riesgo": _buscar_mapa(
                catalogos["riesgo_actual"],
                centro_trabajo
            ),

            "Fijo_variable_2": fijo_variable_2,

            "indicador_sabado": indicador_sabado,
        }

        resultado.append(fila_maestro)

    return resultado


# ============================================================
# ADAPTACIÓN BÁSICA
#
# Solo se utiliza como respaldo cuando generar_excel_reporte()
# es llamado sin una sesión de base de datos.
#
# Para obtener TODAS las descripciones del maestro nuevo debe
# llamarse generar_excel_reporte(filas, db).
# ============================================================

def _adaptar_filas_maestro_sin_db(filas):
    resultado = []

    for fila in filas:
        posicion = fila.get("posicion") or ""
        tipo_contrato = fila.get("tipo_contrato") or ""
        empresa = fila.get("empresa") or 1

        activo_retirado = fila.get("estado") or "A"

        if str(activo_retirado).strip().upper() == "A":
            fecha_retiro = ""
            motivo_retiro = ""
            descripcion_retiro = ""
        else:
            fecha_retiro = fila.get("fecha_retiro") or ""
            motivo_retiro = _entero_si_posible(
                fila.get("motivo_retiro")
            )
            descripcion_retiro = fila.get("descripcion") or ""

        fila_maestro = {
            "compania": 1,
            "Cuenta_de_Gasto": fila.get("cuenta_gasto") or "",
            "desc_compania": "Aseos La Perfección SAS",
            "empleado": fila.get("empleado") or "",
            "pnombre": fila.get("pnombre") or "",
            "snombre": fila.get("snombre") or "",
            "papellido": fila.get("papellido") or "",
            "sapellido": fila.get("sapellido") or "",
            "fnacimiento": fila.get("fecha_nacimiento") or "",
            "tipo_empleado": _entero_si_posible(
                fila.get("tipo_empleado")
            ),
            "pais": _entero_si_posible(
                fila.get("pais")
            ),
            "desc_pais": "",
            "departamento": _entero_si_posible(
                fila.get("departamento")
            ),
            "desc_departamento": "",
            "lugar_nacimiento": (
                fila.get("lugar_nacimiento")
                or ""
            ),
            "direccion": fila.get("direccion") or "",
            "barrio": fila.get("barrio") or "",
            "telefono": fila.get("telefono") or "",
            "tipo_doc_identidad": (
                fila.get("tipo_doc_id")
                or ""
            ),
            "ndoc_identidad": (
                fila.get("num_doc_id")
                or ""
            ),
            "ciudad_doc_ident": (
                fila.get("ciudad_doc_id")
                or ""
            ),
            "pasaporte": fila.get("pasaporte") or "",
            "otro_doc": "",
            "dpto_resid": _entero_si_posible(
                fila.get("depto_residencia")
            ),
            "desc_dpto_resid": "",
            "ciudad_resid_empleado": (
                fila.get("municipio_resid")
                or ""
            ),
            "desc_civil": (
                fila.get("estado_civil")
                or ""
            ),
            "nmilitar": (
                fila.get("libreta_militar")
                or ""
            ),
            "distrito_militar": "",
            "nlic_conducir": (
                fila.get("licencia_conducir")
                or ""
            ),
            "njudicial": (
                fila.get("certif_juducial")
                or ""
            ),
            "nivel_estudios": _entero_si_posible(
                fila.get("nivel_estudios")
            ),
            "desc_estudio": "",
            "sexo": fila.get("sexo") or "",
            "sucursal": _entero_si_posible(
                fila.get("sucursal")
            ),
            "desc_sucursal": "ASEOS LA PERFECCIÓN SAS",
            "cen1": fila.get("centro_costos_1") or "",
            "desc_cen1": "",
            "cen2": fila.get("centro_costos_2") or "",
            "desc_cen2": "",
            "cen3": fila.get("centro_costos_3") or "",
            "desc_cen3": "",
            "cen4": fila.get("centro_costos_4") or "",
            "desc_cen4": "",
            "cen5": fila.get("centro_costos_5") or "",
            "desc_cen5": "",
            "escalafon": _entero_si_posible(
                fila.get("escalafon")
            ),
            "desc_escalafon": "",
            "tipo_contrato": _entero_si_posible(
                tipo_contrato
            ),
            "desc_tipo_con": "",
            "regimen_contrato": _entero_si_posible(
                fila.get("regimen")
            ),
            "desc_regimen_con": "",
            "ncontrato": _entero_si_posible(
                fila.get("ncontrato")
            ),
            "fingreso": fila.get("fecha_ingreso") or "",
            "fterminacion": (
                fila.get("fecha_terminacion")
                or ""
            ),
            "fretiro": fecha_retiro,
            "motivo_retiro": motivo_retiro,
            "descripcion": descripcion_retiro,
            "activo_retirado": activo_retirado,
            "entidad_salud": (
                fila.get("entidad_salud")
                or ""
            ),
            "desc_eps": "",
            "suc_salud": _entero_si_posible(
                fila.get("sucursal_salud")
            ),
            "entidad_pension": (
                fila.get("entidad_pension")
                or ""
            ),
            "desc_afp": "",
            "suc_pension": _entero_si_posible(
                fila.get("sucursal_pension")
            ),
            "entidad_riesgo": (
                fila.get("entidad_riesgo")
                or ""
            ),
            "desc_arp": "",
            "suc_riesgo": _entero_si_posible(
                fila.get("sucur_Ent_riesgo")
            ),
            "caja_compensacion": (
                fila.get("caja_compensacion")
                or ""
            ),
            "desc_caja": "",
            "fondo_cesantias": (
                fila.get("fondo_cesantias")
                or ""
            ),
            "desc_fondo": "",
            "centro_trabajo": (
                fila.get("centro_trabajo")
                or ""
            ),
            "tipo": 1,
            "desc_tipo_emp": "",
            "indicador_reten": (
                fila.get("indicador_retencion")
                or ""
            ),
            "porc_retencion": (
                fila.get("porcentaje_retencion")
                or ""
            ),
            "aux_seguro": (
                fila.get("auxilio_seguro")
                or ""
            ),
            "aux_pension": (
                fila.get("auxilio_pension")
                or ""
            ),
            "aux_solidaridad": (
                fila.get("auxilio_solidaridad")
                or ""
            ),
            "fijo_variable": "",
            "ind_pension": "",
            "ind_salud": "",
            "ind_riesgo": "",
            "tarifa_especial": _entero_si_posible(
                fila.get("tarifa_especial")
            ),
            "porc_seguro": _entero_si_posible(
                fila.get("porcentaje_seguro")
            ),
            "turno": "",
            "numero_reg_eaab": "",
            "tipo_cotizante": _entero_si_posible(
                fila.get("tipo_cotizante")
            ),
            "desc_tipo_coti": "",
            "subtipo_cotizante": _entero_si_posible(
                fila.get("subtipo_cotizante")
            ),
            "desc_subt_coti": "",
            "extranjero_pension": (
                fila.get("extranjero_pension")
                or "N"
            ),
            "reside_exterior": (
                fila.get("reside_exterior")
                or "N"
            ),
            "activo_pensionado": (
                fila.get("activo_pensionado")
                or "A"
            ),
            "posicion": _entero_si_posible(
                posicion
            ),
            "empresa": _entero_si_posible(
                empresa
            ),
            "tipo_sueldo": _entero_si_posible(
                fila.get("tipo_sueldo")
            ),
            "nombre": "Salario interno pesos",
            "forma_pago": (
                fila.get("tipo_pago")
                or "CO"
            ),
            "lugar_deposito": _entero_si_posible(
                fila.get("corporacion")
            ),
            "desc_banco": "",
            "cuenta": fila.get("cuenta") or "",
            "tipo_cuenta": _entero_si_posible(
                fila.get("tipo_cuenta")
            ),
            "posicion_2": _entero_si_posible(
                posicion
            ),
            "nom_posicion": "",
            "empresa_2": _entero_si_posible(
                empresa
            ),
            "nom_emp": "Aseos La Perfección SAS",
            "nit_empresa": "800068462-4",
            "cargo": "",
            "nom_cargo": fila.get("cargo") or "",
            "unidad": "",
            "nom_unidad": "",
            "escalafon_2": _entero_si_posible(
                fila.get("escalafon")
            ),
            "tipo_contrato_2": _entero_si_posible(
                tipo_contrato
            ),
            "descripcion_2": "",
            "email_empleado": fila.get("email") or "",
            "porcentaje_riesgo_ant": "",
            "porcentaje_riesgo": "",
            "Fijo_variable_2": "",
            "indicador_sabado": (
                fila.get("trabaja_sabado")
                or ""
            ),
        }

        resultado.append(fila_maestro)

    return resultado


# ============================================================
# PREPARACIÓN FINAL DEL MAESTRO DE DOTACIÓN
#
# Esta función concentra en un solo punto la transformación que
# necesitan tanto el Excel como Google Sheet. De esta manera,
# ambos destinos consumen exactamente las mismas filas, con las
# mismas 121 columnas y la misma lógica de activos/retiros.
# ============================================================

def preparar_filas_maestro_dotacion(filas, db=None):
    if not filas:
        return []

    # Si todavía llegan las filas antiguas generadas por
    # fn_ReporteSinergy, se convierten al nuevo maestro.
    ya_es_maestro_nuevo = (
        isinstance(filas[0], dict)
        and "fnacimiento" in filas[0]
        and "activo_retirado" in filas[0]
        and "email_empleado" in filas[0]
    )

    if ya_es_maestro_nuevo:
        filas_maestro = [
            dict(fila)
            for fila in filas
        ]

    elif db is not None:
        filas_maestro = adaptar_filas_maestro_nuevo(
            db,
            filas
        )

    else:
        filas_maestro = _adaptar_filas_maestro_sin_db(
            filas
        )

    # Cuando existe una sesión de BD, se incorporan también los
    # retiros finalizados. fn_ReporteSinergy deja de devolverlos
    # después de pasar a estado 35, por eso se agregan aquí.
    if db is not None:
        filas_retiradas = (
            _obtener_retiros_finalizados_maestro(db)
        )

        filas_maestro = _mezclar_activos_y_retirados(
            filas_maestro,
            filas_retiradas
        )

    return filas_maestro


# ============================================================
# GENERACIÓN EXCEL
# ============================================================

def generar_excel_reporte(filas, db=None):
    if not filas:
        return None

    filas_maestro = preparar_filas_maestro_dotacion(
        filas,
        db
    )

    wb = Workbook()

    ws = wb.active
    ws.title = "Reporte"

    # ========================================================
    # ENCABEZADOS EXACTOS DEL ARCHIVO SOLICITADO
    # ========================================================

    headers = [
        encabezado
        for _, encabezado
        in COLUMNAS_MAESTRO_DOTACION
    ]

    for col_num, header in enumerate(
        headers,
        start=1
    ):
        cell = ws.cell(
            row=1,
            column=col_num,
            value=header
        )

        cell.font = Font(bold=True)

    # ========================================================
    # DATOS
    #
    # Se escriben por orden explícito.
    # NO se utiliza list(fila.values()) porque existen
    # encabezados repetidos en el maestro nuevo.
    # ========================================================

    for row_num, fila in enumerate(
        filas_maestro,
        start=2
    ):
        for col_num, (
            clave_interna,
            _
        ) in enumerate(
            COLUMNAS_MAESTRO_DOTACION,
            start=1
        ):
            valor = fila.get(
                clave_interna,
                ""
            )

            ws.cell(
                row=row_num,
                column=col_num,
                value=valor
            )

    # Filtros en encabezados
    ws.auto_filter.ref = (
        f"A1:"
        f"{get_column_letter(len(headers))}"
        f"{max(1, len(filas_maestro) + 1)}"
    )

    # Congelar primera fila
    ws.freeze_panes = "A2"

    # ========================================================
    # AJUSTAR ANCHO DE COLUMNAS
    # ========================================================

    for col in ws.columns:
        max_length = 0

        col_letter = get_column_letter(
            col[0].column
        )

        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )
            except Exception:
                pass

        # Se limita el ancho para evitar columnas gigantes
        # por direcciones, correos u observaciones.
        ancho = min(
            max(max_length + 2, 10),
            45
        )

        ws.column_dimensions[
            col_letter
        ].width = ancho

    # ========================================================
    # NOMBRE Y RUTA
    # ========================================================

    nombre_archivo = (
        "Registro_Dotacion_Actual.xlsx"
    )

    base_dir = (
        Path(__file__)
        .resolve()
        .parent
        .parent
    )

    carpeta_dotacion = (
        base_dir
        / "dotacion"
    )

    carpeta_dotacion.mkdir(
        parents=True,
        exist_ok=True
    )

    ruta = (
        carpeta_dotacion
        / nombre_archivo
    )

    wb.save(
        str(ruta)
    )

    return str(ruta)