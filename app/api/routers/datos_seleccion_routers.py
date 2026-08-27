# api/routers/datos_seleccion_routers.py
from datetime import datetime, timezone
from typing import Annotated, Any
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Body
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, Reference, LineChart, BarChart
from openpyxl.chart.label import DataLabelList

from domain.schemas.registro_personal_update_schema import RegistroPersonalUpdateRequest
from repositories.registro_personal_repo import RegistroPersonalRepository
from infrastructure.db.deps import get_db
from services.datos_seleccion_service import DatosSeleccionService
from domain.schemas.datos_seleccion_schema import (
    DatosSeleccionUpsertRequest,
    DatosSeleccionResponse,
)


router = APIRouter(prefix="/api/datos-seleccion", tags=["datos-seleccion"])
service = DatosSeleccionService()
registro_personal_repo = RegistroPersonalRepository()


@router.put("/registro-personal/{id_registro_personal}")
def actualizar_registro_personal(
    id_registro_personal: int,
    body: Annotated[RegistroPersonalUpdateRequest, Body()],
    db: Annotated[Session, Depends(get_db)],
):
    data = body.model_dump(exclude_none=True)
    if not data:
        raise HTTPException(status_code=400, detail="No hay datos para actualizar")

    data_registro = {
        k: v for k, v in data.items()
        if k not in ("DireccionDatosAdicionales", "IdGrupoSanguineo")
    }

    if data_registro:
        updated = registro_personal_repo.update_by_id(db, id_registro_personal, data_registro)
        if updated == 0:
            raise HTTPException(status_code=404, detail="No se encontró el registro personal")

    if "DireccionDatosAdicionales" in data:
        updated_datos = registro_personal_repo.update_direccion_datos_adicionales(
            db,
            id_registro_personal,
            data["DireccionDatosAdicionales"],
            data.get("IdGrupoSanguineo", 0)
        )
        if updated_datos == 0:
            raise HTTPException(status_code=404, detail="No se encontró DatosAdicionales para este registro personal")

    return {"ok": True, "message": "Registro actualizado", "IdRegistroPersonal": id_registro_personal}


def _parse_bool(value: Any) -> bool | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        s = value.strip().lower()
        if s in ("si", "sí", "true", "1"):
            return True
        if s in ("no", "false", "0"):
            return False
    return None


def _normalizar_estado_dashboard(estado: Any) -> str:
    estado = str(estado or "").strip().upper()
    estado = estado.replace("CONTRATACION", "CONTRATACIÓN")
    estado = estado.replace("EXAMENES", "EXÁMENES")
    estado = estado.replace("REFERENCIACION", "REFERENCIACIÓN")

    if estado.startswith("AVANZA"):
        return "AVANZA A CONTRATACIÓN"
    if estado.startswith("ENTREVISTA"):
        return "ENTREVISTA JEFE INMEDIATO" if "JEFE" in estado else "ENTREVISTA"
    if estado.startswith("PENDIENTE"):
        return "PENDIENTE DE CONTRATACIÓN"
    if estado.startswith("CONTRATADO"):
        return "CONTRATADO"
    if estado.startswith("RECHAZADO"):
        return "RECHAZADO"
    if estado.startswith("DESISTE"):
        return "DESISTE DEL PROCESO"
    if estado.startswith("SEGURIDAD"):
        return "SEGURIDAD"
    if estado.startswith("NUEVO"):
        return "NUEVO"
    if estado.startswith("REFERENCIACIÓN"):
        return "REFERENCIACIÓN"
    if estado.startswith("EXÁMENES"):
        return "EXÁMENES"
    if estado.startswith("ABIERTO"):
        return "ABIERTO"

    return estado or "SIN ESTADO"


def _normalizar_motivo_dashboard(motivo: Any) -> str:
    motivo = str(motivo or "").strip().upper()
    motivo = motivo.replace("CONTRATACION", "CONTRATACIÓN")
    motivo = motivo.replace("EXAMENES", "EXÁMENES")
    motivo = motivo.replace("DOCUMENTACION", "DOCUMENTACIÓN")
    return motivo


def _es_rechazo_contratacion(motivo: Any) -> bool:
    motivo_normalizado = _normalizar_motivo_dashboard(motivo)
    if not motivo_normalizado or motivo_normalizado == "SIN_MOTIVO":
        return False
    return "CONTRATACIÓN" in motivo_normalizado


@router.get("/dashboard-indicadores")
def obtener_dashboard_indicadores_seleccion(
    db: Annotated[Session, Depends(get_db)],
    anio: int | None = None,
    mes: int | None = None,
):
    """
    Dashboard gerencial exclusivo de Selección.

    REGLAS GENERALES
    ============================================================
    - Universo:
      Personas creadas realmente mediante el aplicativo desde
      el 1 de marzo de 2026.

    - No incluye:
      * Activos migrados.
      * Registros identificados como migración.
      * Ajustes administrativos de migración.
      * Registros técnicos de prueba conocidos.

    - Fecha base:
      RegistroPersonal.FechaCreacion.

    - Tarjetas principales:
      1. Total personal registrado.
      2. Personal que avanzó a Contratación.
      3. Personal rechazado en Selección.

    - Porcentajes:
      * Total registrados = 100 %.
      * Avanzan a Contratación = avanzan / registrados.
      * Rechazados Selección = rechazados / registrados.
      * Estados = cantidad estado / registrados.
      * Motivos rechazo = cantidad motivo / rechazados Selección.

    - Serie mensual:
      Agrupa por mes de RegistroPersonal.FechaCreacion y devuelve:
      registrados, avanzan, rechazados y sus porcentajes.

    - Detalle mensual:
      Cada mes incluye:
      tarjetas, estados y motivos de rechazo.

    - Rechazos de Contratación:
      No alimentan los indicadores ni los motivos de Selección.

    SOLO CONSULTA.
    No inserta, no actualiza y no elimina información.
    """

    if mes is not None and (mes < 1 or mes > 12):
        raise HTTPException(
            status_code=400,
            detail="El mes debe estar entre 1 y 12.",
        )

    if anio is not None and anio < 2000:
        raise HTTPException(
            status_code=400,
            detail="El año consultado no es válido.",
        )

    # ============================================================
    # 1. UNIVERSO GENERAL DE SELECCIÓN
    # ============================================================
    #
    # IMPORTANTE:
    # El filtro de año se aplica aquí para construir la serie
    # histórica del año consultado.
    #
    # El filtro de mes NO se aplica en SQL.
    # Se aplica después en Python para conservar la serie completa
    # del año y al mismo tiempo entregar el detalle del mes elegido.
    # ============================================================

    rows = db.execute(
        text("""
            WITH universo_seleccion AS (
                SELECT
                    rp."IdRegistroPersonal",
                    rp."IdEstadoProceso",
                    rp."FechaCreacion",
                    rp."UsuarioActualizacion",

                    EXISTS (
                        SELECT 1
                        FROM public."HistorialEstadoContratacion" hec24
                        WHERE
                            hec24."IdRegistroPersonal"
                                = rp."IdRegistroPersonal"
                            AND hec24."EstadoNuevo" = 24
                    ) AS tiene_estado_24,

                    EXISTS (
                        SELECT 1
                        FROM public."HistorialEstadoContratacion" hec25
                        WHERE
                            hec25."IdRegistroPersonal"
                                = rp."IdRegistroPersonal"
                            AND hec25."EstadoNuevo" = 25
                    ) AS tiene_estado_25,

                    EXISTS (
                        SELECT 1
                        FROM public."ContratacionBasica" cb
                        WHERE
                            cb."IdRegistroPersonal"
                                = rp."IdRegistroPersonal"
                            AND cb."FechaIngreso" IS NOT NULL
                    ) AS tiene_contratacion_basica,

                    EXISTS (
                        SELECT 1
                        FROM public."HistorialEstadoContratacion" hec_rechazo
                        WHERE
                            hec_rechazo."IdRegistroPersonal"
                                = rp."IdRegistroPersonal"
                            AND hec_rechazo."EstadoNuevo" = 28
                            AND UPPER(
                                TRIM(
                                    COALESCE(
                                        hec_rechazo."Modulo",
                                        ''
                                    )
                                )
                            ) = 'CONTRATACION'
                    ) AS tiene_rechazo_contratacion_historial

                FROM public."RegistroPersonal" rp

                WHERE
                    rp."FechaCreacion"
                        >= TIMESTAMPTZ '2026-03-01 00:00:00-05'

                    AND (
                        :anio IS NULL
                        OR EXTRACT(
                            YEAR FROM rp."FechaCreacion"
                        ) = :anio
                    )

                    -- Excluir activos históricos migrados
                    AND NOT EXISTS (
                        SELECT 1
                        FROM public."HistorialLaboral" hl
                        WHERE
                            hl."IdRegistroPersonal"
                                = rp."IdRegistroPersonal"
                            AND UPPER(
                                TRIM(
                                    COALESCE(
                                        hl."TipoVinculacion",
                                        ''
                                    )
                                )
                            ) = 'ACTIVO MIGRADO'
                    )

                    -- Excluir etiquetas de migración
                    AND LOWER(
                        COALESCE(
                            rp."UsuarioActualizacion",
                            ''
                        )
                    ) NOT LIKE '%migracion%'

                    AND LOWER(
                        COALESCE(
                            rp."UsuarioActualizacion",
                            ''
                        )
                    ) NOT LIKE '%migrado%'

                    -- Excluir ajuste histórico administrativo
                    AND COALESCE(
                        rp."UsuarioActualizacion",
                        ''
                    ) <> 'ajuste_no_activos_maestro_2026_06_22'

                    -- Registros técnicos conocidos de prueba
                    AND rp."NumeroIdentificacion"::text NOT IN (
                        '91011506',
                        '0987654',
                        '951357'
                    )
            )

            SELECT
                us."IdRegistroPersonal",
                us."IdEstadoProceso",
                us."FechaCreacion" AS fecha_registro,

                us.tiene_estado_24,
                us.tiene_estado_25,
                us.tiene_contratacion_basica,
                us.tiene_rechazo_contratacion_historial,

                CASE
                    -- En los indicadores de Selección, una persona que ya
                    -- avanzó a Contratación conserva ese resultado aunque
                    -- posteriormente cambie de estado o sea rechazada allá.
                    WHEN (
                        us.tiene_estado_24
                        OR us.tiene_estado_25
                        OR us.tiene_contratacion_basica
                        OR us.tiene_rechazo_contratacion_historial
                        OR UPPER(
                            TRIM(
                                COALESCE(
                                    mcp."MotivoCierre",
                                    ''
                                )
                            )
                        ) IN (
                            'NO ASISTE A CONTRATACION',
                            'NO ASISTE A CONTRATACIÓN'
                        )
                        OR us."IdEstadoProceso" IN (
                            24,
                            25,
                            30,
                            31,
                            32,
                            33,
                            34,
                            35
                        )
                    )
                    THEN 'AVANZA A CONTRATACIÓN'

                    -- Solo queda como rechazo de Selección cuando nunca
                    -- existe evidencia de haber avanzado a Contratación.
                    WHEN us."IdEstadoProceso" = 28
                    THEN 'RECHAZADO'

                    WHEN us."IdEstadoProceso" = 27
                    THEN 'DESISTE DEL PROCESO'

                    WHEN us."IdEstadoProceso" = 26
                    THEN 'REFERENCIACIÓN'

                    WHEN us."IdEstadoProceso" = 22
                    THEN 'SEGURIDAD'

                    WHEN us."IdEstadoProceso" = 21
                    THEN 'EXÁMENES'

                    WHEN us."IdEstadoProceso" = 20
                    THEN 'ENTREVISTA JEFE INMEDIATO'

                    WHEN us."IdEstadoProceso" = 19
                    THEN 'ENTREVISTA'

                    WHEN us."IdEstadoProceso" = 18
                    THEN 'NUEVO'

                    ELSE CONCAT(
                        'ESTADO ',
                        COALESCE(
                            us."IdEstadoProceso"::text,
                            'SIN DEFINIR'
                        )
                    )
                END AS estado_seleccion,

                (
                    us.tiene_estado_24
                    OR us.tiene_estado_25
                    OR us.tiene_contratacion_basica
                    OR us.tiene_rechazo_contratacion_historial
                    OR UPPER(
                        TRIM(
                            COALESCE(
                                mcp."MotivoCierre",
                                ''
                            )
                        )
                    ) IN (
                        'NO ASISTE A CONTRATACION',
                        'NO ASISTE A CONTRATACIÓN'
                    )
                    OR us."IdEstadoProceso" IN (
                        24,
                        25,
                        30,
                        31,
                        32,
                        33,
                        34,
                        35
                    )
                ) AS avanzo_contratacion,

                mcp."MotivoCierre" AS motivo_rechazo,

                COALESCE(
                    mcp."FechaActualizacion",
                    mcp."FechaCreacion"
                ) AS fecha_motivo_rechazo,

                CASE
                    WHEN us."IdEstadoProceso" <> 28
                    THEN NULL

                    WHEN (
                        us.tiene_estado_24
                        OR us.tiene_estado_25
                        OR us.tiene_contratacion_basica
                        OR us.tiene_rechazo_contratacion_historial
                        OR UPPER(
                            TRIM(
                                COALESCE(
                                    mcp."MotivoCierre",
                                    ''
                                )
                            )
                        ) IN (
                            'NO ASISTE A CONTRATACION',
                            'NO ASISTE A CONTRATACIÓN'
                        )
                    )
                    THEN 'CONTRATACION'

                    ELSE 'SELECCION'
                END AS origen_rechazo

            FROM universo_seleccion us

            LEFT JOIN LATERAL (
                SELECT
                    mcp_detalle."MotivoCierre",
                    mcp_detalle."FechaCreacion",
                    mcp_detalle."FechaActualizacion"

                FROM public."MotivoCierreProceso" mcp_detalle

                WHERE
                    mcp_detalle."IdRegistroPersonal"
                        = us."IdRegistroPersonal"

                ORDER BY
                    COALESCE(
                        mcp_detalle."FechaActualizacion",
                        mcp_detalle."FechaCreacion"
                    ) DESC,
                    mcp_detalle."IdMotivoCierre" DESC

                LIMIT 1
            ) mcp ON TRUE

            ORDER BY
                us."FechaCreacion",
                us."IdRegistroPersonal";
        """),
        {
            "anio": anio,
        },
    ).mappings().all()

    filas_universo = [
        dict(row)
        for row in rows
    ]

    # ============================================================
    # 2. CATÁLOGOS DEL DASHBOARD
    # ============================================================

    estados_base = [
        "NUEVO",
        "ENTREVISTA",
        "ENTREVISTA JEFE INMEDIATO",
        "EXÁMENES",
        "SEGURIDAD",
        "REFERENCIACIÓN",
        "AVANZA A CONTRATACIÓN",
        "PENDIENTE DE CONTRATACIÓN",
        "CONTRATADO",
        "DESISTE DEL PROCESO",
        "RECHAZADO",
    ]

    equivalencias_motivos = {
        "DESISTE DEL PROCESO":
            "Desiste del proceso",

        "NO CUMPLE PERFIL":
            "No cumple perfil",

        "NO ASISTE A EXAMENES MEDICOS":
            "No asiste a exámenes médicos",

        "NO ASISTE A EXÁMENES MEDICOS":
            "No asiste a exámenes médicos",

        "NO ASISTE A EXÁMENES MÉDICOS":
            "No asiste a exámenes médicos",

        "EXAMENES NO APTOS":
            "Exámenes no aptos",

        "EXÁMENES NO APTOS":
            "Exámenes no aptos",

        "DOCUMENTACION INCOMPLETA":
            "Documentación incompleta",

        "DOCUMENTACIÓN INCOMPLETA":
            "Documentación incompleta",

        "ESTUDIO DE SEGURIDAD":
            "Estudio de seguridad",

        "REINTEGRO NO APROBADO":
            "Reintegro no aprobado",

        "NO SUPERA PRUEBA FISICA":
            "No supera prueba física",

        "NO SUPERA PRUEBA FÍSICA":
            "No supera prueba física",

        "OTRO":
            "Otro",
    }

    motivos_base = [
        "Desiste del proceso",
        "No cumple perfil",
        "No asiste a exámenes médicos",
        "Exámenes no aptos",
        "Documentación incompleta",
        "Estudio de seguridad",
        "Reintegro no aprobado",
        "No supera prueba física",
        "Otro",
        "Sin motivo registrado",
    ]

    meses_nombre = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }

    # ============================================================
    # 3. FUNCIONES INTERNAS DE CÁLCULO
    # ============================================================

    def es_rechazo_seleccion(fila: dict) -> bool:
        estado = _normalizar_estado_dashboard(
            fila.get("estado_seleccion")
        )

        if estado != "RECHAZADO":
            return False

        motivo_normalizado = _normalizar_motivo_dashboard(
            fila.get("motivo_rechazo")
        )

        es_rechazo_contratacion = (
            bool(fila.get("tiene_estado_24"))
            or bool(fila.get("tiene_estado_25"))
            or bool(fila.get("tiene_contratacion_basica"))
            or bool(
                fila.get(
                    "tiene_rechazo_contratacion_historial"
                )
            )
            or motivo_normalizado
                == "NO ASISTE A CONTRATACIÓN"
        )

        return not es_rechazo_contratacion

    def calcular_estados(
        filas_periodo: list[dict],
    ) -> list[dict]:
        total_periodo = len(filas_periodo)

        conteo = {
            estado: 0
            for estado in estados_base
        }

        for fila in filas_periodo:
            estado = _normalizar_estado_dashboard(
                fila.get("estado_seleccion")
            )

            if estado not in conteo:
                conteo[estado] = 0

            conteo[estado] += 1

        orden = estados_base + [
            estado
            for estado in conteo
            if estado not in estados_base
        ]

        return [
            {
                "estado": estado,
                "cantidad": conteo.get(
                    estado,
                    0,
                ),
                "porcentaje": (
                    round(
                        (
                            conteo.get(
                                estado,
                                0,
                            )
                            / total_periodo
                        )
                        * 100,
                        2,
                    )
                    if total_periodo
                    else 0
                ),
            }
            for estado in orden
        ]

    def calcular_motivos_rechazo(
        filas_periodo: list[dict],
    ) -> tuple[list[dict], int, int]:
        conteo = {
            motivo: 0
            for motivo in motivos_base
        }

        total_rechazados = 0
        rechazados_excluidos = 0

        for fila in filas_periodo:
            estado = _normalizar_estado_dashboard(
                fila.get("estado_seleccion")
            )

            if estado != "RECHAZADO":
                continue

            motivo_original = fila.get(
                "motivo_rechazo"
            )

            motivo_normalizado = (
                _normalizar_motivo_dashboard(
                    motivo_original
                )
            )

            if not es_rechazo_seleccion(fila):
                rechazados_excluidos += 1
                continue

            total_rechazados += 1

            if (
                not motivo_normalizado
                or motivo_normalizado == "SIN_MOTIVO"
            ):
                motivo_final = (
                    "Sin motivo registrado"
                )
            else:
                motivo_final = (
                    equivalencias_motivos.get(
                        motivo_normalizado,
                        str(
                            motivo_original
                        ).strip(),
                    )
                )

            if motivo_final not in conteo:
                conteo[motivo_final] = 0

            conteo[motivo_final] += 1

        motivos = [
            {
                "motivo": motivo,
                "cantidad": cantidad,
                "porcentaje": (
                    round(
                        (
                            cantidad
                            / total_rechazados
                        )
                        * 100,
                        2,
                    )
                    if total_rechazados
                    else 0
                ),
            }
            for motivo, cantidad in conteo.items()
        ]

        return (
            motivos,
            total_rechazados,
            rechazados_excluidos,
        )

    def construir_resumen(
        filas_periodo: list[dict],
    ) -> dict:
        total_periodo = len(
            filas_periodo
        )

        total_avanzan = sum(
            1
            for fila in filas_periodo
            if bool(
                fila.get(
                    "avanzo_contratacion"
                )
            )
        )

        (
            motivos,
            total_rechazados,
            rechazados_excluidos,
        ) = calcular_motivos_rechazo(
            filas_periodo
        )

        estados = calcular_estados(
            filas_periodo
        )

        conteo_estados = {
            item["estado"]:
                item["cantidad"]
            for item in estados
        }

        contratados = (
            conteo_estados.get(
                "CONTRATADO",
                0,
            )
        )

        desistidos = (
            conteo_estados.get(
                "DESISTE DEL PROCESO",
                0,
            )
        )

        rechazados_estado_actual = (
            conteo_estados.get(
                "RECHAZADO",
                0,
            )
        )

        pendientes_contratacion = (
            conteo_estados.get(
                "PENDIENTE DE CONTRATACIÓN",
                0,
            )
        )

        en_proceso = max(
            total_periodo
            - contratados
            - desistidos
            - rechazados_estado_actual,
            0,
        )

        porcentaje_avanzan = (
            round(
                (
                    total_avanzan
                    / total_periodo
                )
                * 100,
                2,
            )
            if total_periodo
            else 0
        )

        porcentaje_rechazados = (
            round(
                (
                    total_rechazados
                    / total_periodo
                )
                * 100,
                2,
            )
            if total_periodo
            else 0
        )

        return {
            "tarjetas": {
                "registrados": {
                    "cantidad": total_periodo,
                    "porcentaje": (
                        100.0
                        if total_periodo
                        else 0
                    ),
                },
                "avanzan_contratacion": {
                    "cantidad":
                        total_avanzan,
                    "porcentaje":
                        porcentaje_avanzan,
                },
                "rechazados_seleccion": {
                    "cantidad":
                        total_rechazados,
                    "porcentaje":
                        porcentaje_rechazados,
                },
            },

            "total":
                total_periodo,

            "registrados_seleccion":
                total_periodo,

            "avanza_contratacion":
                total_avanzan,

            "total_personas_avanzadas_contratacion":
                total_avanzan,

            "porcentaje_avanza_contratacion":
                porcentaje_avanzan,

            "rechazados_generales":
                total_rechazados,

            "rechazados_seleccion":
                total_rechazados,

            "porcentaje_rechazados_seleccion":
                porcentaje_rechazados,

            "rechazados_estado_actual":
                rechazados_estado_actual,

            "rechazados_excluidos_contratacion":
                rechazados_excluidos,

            "desistidos":
                desistidos,

            "contratados":
                contratados,

            "pendiente_contratacion":
                pendientes_contratacion,

            "en_proceso":
                en_proceso,

            "estados":
                estados,

            "estados_con_datos": [
                item
                for item in estados
                if item["cantidad"] > 0
            ],

            "motivos_rechazo_generales":
                motivos,

            "motivos_rechazo_generales_con_datos": [
                item
                for item in motivos
                if item["cantidad"] > 0
            ],
        }

    # ============================================================
    # 4. SERIE MENSUAL
    # ============================================================

    filas_por_mes: dict[str, list[dict]] = {}

    for fila in filas_universo:
        fecha = fila.get(
            "fecha_registro"
        )

        if not isinstance(
            fecha,
            datetime,
        ):
            continue

        clave = fecha.strftime(
            "%Y-%m"
        )

        if clave not in filas_por_mes:
            filas_por_mes[clave] = []

        filas_por_mes[clave].append(
            fila
        )

    serie_mensual = []
    detalle_mensual = []

    for clave in sorted(
        filas_por_mes.keys()
    ):
        filas_mes = filas_por_mes[
            clave
        ]

        fecha_referencia = (
            filas_mes[0].get(
                "fecha_registro"
            )
        )

        if not isinstance(
            fecha_referencia,
            datetime,
        ):
            continue

        resumen_mes = construir_resumen(
            filas_mes
        )

        tarjetas_mes = resumen_mes[
            "tarjetas"
        ]

        item_serie = {
            "clave": clave,
            "anio":
                fecha_referencia.year,
            "numero_mes":
                fecha_referencia.month,
            "mes":
                meses_nombre[
                    fecha_referencia.month
                ],
            "etiqueta": (
                f"{meses_nombre[fecha_referencia.month].capitalize()} "
                f"{fecha_referencia.year}"
            ),

            "registrados":
                tarjetas_mes[
                    "registrados"
                ]["cantidad"],

            "porcentaje_registrados":
                tarjetas_mes[
                    "registrados"
                ]["porcentaje"],

            "avanzan_contratacion":
                tarjetas_mes[
                    "avanzan_contratacion"
                ]["cantidad"],

            "porcentaje_avanzan_contratacion":
                tarjetas_mes[
                    "avanzan_contratacion"
                ]["porcentaje"],

            "rechazados_seleccion":
                tarjetas_mes[
                    "rechazados_seleccion"
                ]["cantidad"],

            "porcentaje_rechazados_seleccion":
                tarjetas_mes[
                    "rechazados_seleccion"
                ]["porcentaje"],
        }

        serie_mensual.append(
            item_serie
        )

        detalle_mensual.append(
            {
                **item_serie,
                "tarjetas":
                    resumen_mes[
                        "tarjetas"
                    ],
                "estados":
                    resumen_mes[
                        "estados"
                    ],
                "estados_con_datos":
                    resumen_mes[
                        "estados_con_datos"
                    ],
                "motivos_rechazo":
                    resumen_mes[
                        "motivos_rechazo_generales"
                    ],
                "motivos_rechazo_con_datos":
                    resumen_mes[
                        "motivos_rechazo_generales_con_datos"
                    ],
            }
        )

    # ============================================================
    # 5. PERÍODO ACTUAL SOLICITADO
    # ============================================================
    #
    # Sin mes:
    #   devuelve el universo completo del año consultado o global.
    #
    # Con mes:
    #   devuelve solamente la cohorte creada en ese mes.
    #
    # La gráfica conserva la serie completa del año.
    # ============================================================

    if mes is not None:
        filas_periodo = [
            fila
            for fila in filas_universo
            if isinstance(
                fila.get(
                    "fecha_registro"
                ),
                datetime,
            )
            and fila[
                "fecha_registro"
            ].month == mes
        ]
    else:
        filas_periodo = list(
            filas_universo
        )

    resumen_periodo = construir_resumen(
        filas_periodo
    )

    # ============================================================
    # 6. RESPUESTA FINAL
    # ============================================================

    return {
        "filtros": {
            "anio": anio,
            "mes": mes,
            "fecha_inicio_aplicativo":
                "2026-03-01",
            "modo": (
                "mensual"
                if mes is not None
                else (
                    "anual"
                    if anio is not None
                    else "global"
                )
            ),
        },

        # --------------------------------------------------------
        # Compatibilidad con el frontend actual
        # --------------------------------------------------------

        "total":
            resumen_periodo[
                "total"
            ],

        "registrados_seleccion":
            resumen_periodo[
                "registrados_seleccion"
            ],

        "avanza_contratacion":
            resumen_periodo[
                "avanza_contratacion"
            ],

        "total_personas_avanzadas_contratacion":
            resumen_periodo[
                "total_personas_avanzadas_contratacion"
            ],

        "rechazados_generales":
            resumen_periodo[
                "rechazados_generales"
            ],

        "rechazados_seleccion":
            resumen_periodo[
                "rechazados_seleccion"
            ],

        "rechazados_estado_actual":
            resumen_periodo[
                "rechazados_estado_actual"
            ],

        "rechazados_excluidos_contratacion":
            resumen_periodo[
                "rechazados_excluidos_contratacion"
            ],

        "desistidos":
            resumen_periodo[
                "desistidos"
            ],

        "contratados":
            resumen_periodo[
                "contratados"
            ],

        "pendiente_contratacion":
            resumen_periodo[
                "pendiente_contratacion"
            ],

        "en_proceso":
            resumen_periodo[
                "en_proceso"
            ],

        "estados":
            resumen_periodo[
                "estados"
            ],

        "estados_con_datos":
            resumen_periodo[
                "estados_con_datos"
            ],

        "motivos_rechazo_generales":
            resumen_periodo[
                "motivos_rechazo_generales"
            ],

        "motivos_rechazo_generales_con_datos":
            resumen_periodo[
                "motivos_rechazo_generales_con_datos"
            ],

        # Se conservan para no romper el frontend viejo
        "rechazados": 0,
        "motivos_rechazo": [],
        "motivos_rechazo_con_datos": [],

        # --------------------------------------------------------
        # Nueva estructura del Panel Gerencial de Selección
        # --------------------------------------------------------

        "tarjetas":
            resumen_periodo[
                "tarjetas"
            ],

        "porcentajes": {
            "registrados":
                (
                    100.0
                    if resumen_periodo[
                        "total"
                    ]
                    else 0
                ),

            "avanzan_contratacion":
                resumen_periodo[
                    "porcentaje_avanza_contratacion"
                ],

            "rechazados_seleccion":
                resumen_periodo[
                    "porcentaje_rechazados_seleccion"
                ],
        },

        "serie_mensual":
            serie_mensual,

        # Alias temporal para compatibilidad
        "registros_por_mes": [
            {
                "mes":
                    item["etiqueta"],
                "registros":
                    item["registrados"],
                "avanzan_contratacion":
                    item["avanzan_contratacion"],
                "rechazados_seleccion":
                    item["rechazados_seleccion"],
            }
            for item in serie_mensual
        ],

        "detalle_mensual":
            detalle_mensual,

        "periodo_actual":
            {
                "anio": anio,
                "mes": mes,
                **resumen_periodo,
            },

        "auditoria": {
            "universo": (
                "RegistroPersonal.FechaCreacion "
                "desde 2026-03-01"
            ),

            "filtro_periodo":
                "RegistroPersonal.FechaCreacion",

            "estado_actual": (
                "Resultado vigente del proceso "
                "de Selección"
            ),

            "avance_contratacion": (
                "Estado 24 o evidencia posterior "
                "de contratación"
            ),

            "rechazo_seleccion": (
                "Estado 28 sin evidencia de "
                "rechazo de Contratación"
            ),

            "rechazo_contratacion_excluido": (
                "Estado 28 con estado 24, "
                "estado 25, ContratacionBasica, "
                "movimiento de rechazo en "
                "Contratación o motivo histórico "
                "No asiste a Contratación"
            ),

            "porcentaje_tarjetas": (
                "Avanzan y rechazados se calculan "
                "sobre el total registrado del período"
            ),

            "porcentaje_estados": (
                "Cada estado se calcula sobre "
                "el total registrado del período"
            ),

            "porcentaje_motivos": (
                "Cada motivo se calcula sobre "
                "el total de rechazados en Selección "
                "del período"
            ),

            "serie_mensual": (
                "Agrupada por "
                "RegistroPersonal.FechaCreacion"
            ),

            "excluye_migrados":
                True,

            "solo_consulta":
                True,
        },
    }


# ============================================================
# KPI 2 SELECCIÓN - TASA DE ROTACIÓN DEL NUEVO PERSONAL
# ============================================================
# Este endpoint es independiente del dashboard principal de Selección.
#
# Reglas:
# - Universo: personal con ContratacionBasica.FechaIngreso desde 2026-03-01.
# - Excluye personal migrado y registros técnicos conocidos de prueba.
# - Fecha real de retiro: PazYSalvoOperaciones.FechaUltimoDiaLaborado.
# - Cuando existe un ciclo laboral ACTIVO o EN_PROCESO, solo se toma un
#   Paz y Salvo cuya FechaUltimoDiaLaborado pertenezca temporalmente a ese
#   ciclo. Así un retiro histórico no se cruza con un reintegro vigente.
# - Cortes acumulados oficiales: 7, 15, 30 y 60 días.
# - Una persona es evaluable para un corte cuando:
#     1. ya transcurrió ese número de días desde su FechaIngreso, o
#     2. ya se retiró dentro de ese corte.
# - La tasa solo se publica cuando al menos el 80 % de la cohorte válida
#   ya alcanzó naturalmente el corte.
# - Los retiros tempranos se contabilizan, pero por sí solos no hacen
#   madurar el indicador.
# - Si el corte aún no madura, la tasa se devuelve como None para que
#   el frontend muestre "Pendiente de maduración".
# - El mes se interpreta como cohorte de ingreso, no como mes de retiro.
#
# SOLO CONSULTA.
# No inserta, no actualiza y no elimina información.
# ============================================================


@router.get("/dashboard-indicadores-rotacion-nuevo-personal")
def obtener_dashboard_rotacion_nuevo_personal(
    db: Annotated[Session, Depends(get_db)],
    anio: int | None = None,
    mes: int | None = None,
):
    """
    KPI 2 de Selección: tasa de rotación del nuevo personal.

    La cohorte se define por ContratacionBasica.FechaIngreso.

    El retiro efectivo se toma de
    PazYSalvoOperaciones.FechaUltimoDiaLaborado.

    Si el trabajador tiene un ciclo laboral ACTIVO o EN_PROCESO,
    el retiro usado por el KPI debe corresponder temporalmente a ese
    ciclo. Esto evita cruzar un retiro histórico con el ingreso de un
    reintegro vigente.
    """

    if mes is not None and (mes < 1 or mes > 12):
        raise HTTPException(
            status_code=400,
            detail="El mes debe estar entre 1 y 12.",
        )

    if anio is not None and anio < 2000:
        raise HTTPException(
            status_code=400,
            detail="El año consultado no es válido.",
        )

    rows = db.execute(
        text("""
            SELECT
                rp."IdRegistroPersonal" AS id_registro_personal,
                cb."FechaIngreso" AS fecha_ingreso,
                pso."FechaUltimoDiaLaborado" AS ultimo_dia_laborado,

                CASE
                    WHEN
                        pso."FechaUltimoDiaLaborado" IS NOT NULL
                        AND pso."FechaUltimoDiaLaborado" >= cb."FechaIngreso"
                    THEN
                        pso."FechaUltimoDiaLaborado" - cb."FechaIngreso"
                    ELSE NULL
                END AS dias_permanencia

            FROM public."RegistroPersonal" rp

            INNER JOIN public."ContratacionBasica" cb
                ON cb."IdRegistroPersonal" = rp."IdRegistroPersonal"

            -- Ciclo laboral vigente.
            --
            -- Se usa únicamente para evitar que un retiro de un ciclo
            -- histórico se mezcle con la FechaIngreso del reintegro actual.
            -- No modifica información y no altera trabajadores que todavía
            -- no tengan ciclos en VinculacionLaboral.
            LEFT JOIN LATERAL (
                SELECT
                    vl."IdVinculacionLaboral",
                    vl."NumeroCiclo",
                    vl."TipoVinculacion",
                    vl."EstadoVinculacion",
                    vl."FechaIngreso"
                FROM public."VinculacionLaboral" vl
                WHERE
                    vl."IdRegistroPersonal" = rp."IdRegistroPersonal"
                    AND vl."EstadoVinculacion" IN (
                        'EN_PROCESO',
                        'ACTIVO'
                    )
                ORDER BY
                    vl."NumeroCiclo" DESC,
                    vl."IdVinculacionLaboral" DESC
                LIMIT 1
            ) ciclo_vigente ON TRUE

            LEFT JOIN LATERAL (
                SELECT
                    p."FechaUltimoDiaLaborado"
                FROM public."PazYSalvoOperaciones" p
                WHERE
                    p."IdRegistroPersonal" = rp."IdRegistroPersonal"

                    -- Si no existe ciclo vigente, se conserva exactamente
                    -- el comportamiento anterior del KPI y se toma el último
                    -- Paz y Salvo del trabajador.
                    --
                    -- Si sí existe ciclo vigente con FechaIngreso, solo se
                    -- acepta un retiro del mismo periodo laboral. Un retiro
                    -- anterior pertenece a un ciclo histórico y no debe
                    -- marcar el reintegro actual como fecha inconsistente.
                    AND (
                        ciclo_vigente."IdVinculacionLaboral" IS NULL
                        OR ciclo_vigente."FechaIngreso" IS NULL
                        OR p."FechaUltimoDiaLaborado"
                            >= ciclo_vigente."FechaIngreso"
                    )

                ORDER BY
                    p."FechaCarga" DESC NULLS LAST,
                    p."IdPazYSalvo" DESC
                LIMIT 1
            ) pso ON TRUE

            WHERE
                cb."FechaIngreso" >= DATE '2026-03-01'

                AND (
                    :anio IS NULL
                    OR EXTRACT(YEAR FROM cb."FechaIngreso") = :anio
                )

                AND NOT EXISTS (
                    SELECT 1
                    FROM public."HistorialLaboral" hl
                    WHERE
                        hl."IdRegistroPersonal" = rp."IdRegistroPersonal"
                        AND UPPER(
                            TRIM(
                                COALESCE(
                                    hl."TipoVinculacion",
                                    ''
                                )
                            )
                        ) = 'ACTIVO MIGRADO'
                )

                AND LOWER(
                    COALESCE(
                        rp."UsuarioActualizacion",
                        ''
                    )
                ) NOT LIKE '%migracion%'

                AND LOWER(
                    COALESCE(
                        rp."UsuarioActualizacion",
                        ''
                    )
                ) NOT LIKE '%migrado%'

                AND COALESCE(
                    rp."UsuarioActualizacion",
                    ''
                ) <> 'ajuste_no_activos_maestro_2026_06_22'

                AND rp."NumeroIdentificacion"::text NOT IN (
                    '91011506',
                    '0987654',
                    '951357'
                )

            ORDER BY
                cb."FechaIngreso",
                rp."IdRegistroPersonal";
        """),
        {
            "anio": anio,
        },
    ).mappings().all()

    filas_universo = [
        dict(row)
        for row in rows
    ]

    cortes = (7, 15, 30, 60)
    umbral_maduracion = 0.80

    # Se consulta una sola vez la fecha actual de la base de datos.
    # Esto evita ejecutar consultas adicionales por cada trabajador
    # y mantiene todos los cálculos con la misma fecha de referencia.
    fecha_actual_bd = db.execute(
        text("SELECT CURRENT_DATE")
    ).scalar_one()

    def construir_resumen_rotacion(
        filas_periodo: list[dict],
    ) -> dict:
        total_contratados = len(filas_periodo)

        sin_retiro_registrado = sum(
            1
            for fila in filas_periodo
            if fila.get("ultimo_dia_laborado") is None
        )

        fechas_inconsistentes = sum(
            1
            for fila in filas_periodo
            if fila.get("ultimo_dia_laborado") is not None
            and fila.get("fecha_ingreso") is not None
            and fila.get("ultimo_dia_laborado") < fila.get("fecha_ingreso")
        )

        retiros_mas_60 = sum(
            1
            for fila in filas_periodo
            if fila.get("dias_permanencia") is not None
            and fila.get("dias_permanencia") > 60
        )

        rangos_exclusivos = {
            "0_7": 0,
            "8_15": 0,
            "16_30": 0,
            "31_60": 0,
        }

        for fila in filas_periodo:
            dias = fila.get("dias_permanencia")

            if dias is None:
                continue

            if 0 <= dias <= 7:
                rangos_exclusivos["0_7"] += 1
            elif 8 <= dias <= 15:
                rangos_exclusivos["8_15"] += 1
            elif 16 <= dias <= 30:
                rangos_exclusivos["16_30"] += 1
            elif 31 <= dias <= 60:
                rangos_exclusivos["31_60"] += 1

        tarjetas_cortes = {}

        for corte in cortes:
            evaluables = 0
            evaluables_naturales = 0
            retiros = 0

            for fila in filas_periodo:
                fecha_ingreso = fila.get("fecha_ingreso")
                dias = fila.get("dias_permanencia")
                ultimo_dia_laborado = fila.get("ultimo_dia_laborado")

                if fecha_ingreso is None:
                    continue

                fecha_inconsistente = (
                    ultimo_dia_laborado is not None
                    and ultimo_dia_laborado < fecha_ingreso
                )

                if fecha_inconsistente:
                    continue

                ya_maduro = (
                    (fecha_actual_bd - fecha_ingreso).days >= corte
                )

                retiro_dentro_corte = (
                    dias is not None
                    and 0 <= dias <= corte
                )

                if ya_maduro:
                    evaluables_naturales += 1

                if ya_maduro or retiro_dentro_corte:
                    evaluables += 1

                if retiro_dentro_corte:
                    retiros += 1

            base_maduracion = max(
                total_contratados - fechas_inconsistentes,
                0,
            )

            porcentaje_maduracion = (
                round(
                    evaluables_naturales * 100.0 / base_maduracion,
                    2,
                )
                if base_maduracion > 0
                else 0
            )

            corte_maduro = (
                base_maduracion > 0
                and (
                    evaluables_naturales / base_maduracion
                ) >= umbral_maduracion
            )

            tasa = (
                round(
                    retiros * 100.0 / evaluables,
                    2,
                )
                if corte_maduro and evaluables > 0
                else None
            )

            tarjetas_cortes[str(corte)] = {
                "corte_dias": corte,
                "evaluables": evaluables,
                "evaluables_naturales": evaluables_naturales,
                "retiros": retiros,
                "tasa": tasa,
                "maduro": corte_maduro,
                "porcentaje_maduracion": porcentaje_maduracion,
                "umbral_maduracion": 80.0,
                "estado": (
                    "EVALUABLE"
                    if corte_maduro
                    else "PENDIENTE_MADURACION"
                ),
            }

        return {
            "total_contratados": total_contratados,

            "tarjetas": {
                "total_contratados": {
                    "cantidad": total_contratados,
                    "porcentaje": 100.0 if total_contratados else 0,
                },
                "hasta_7_dias": tarjetas_cortes["7"],
                "hasta_15_dias": tarjetas_cortes["15"],
                "hasta_30_dias": tarjetas_cortes["30"],
                "hasta_60_dias": tarjetas_cortes["60"],
            },

            "rangos_exclusivos": {
                "retiro_0_7": rangos_exclusivos["0_7"],
                "retiro_8_15": rangos_exclusivos["8_15"],
                "retiro_16_30": rangos_exclusivos["16_30"],
                "retiro_31_60": rangos_exclusivos["31_60"],
                "retiro_mas_60": retiros_mas_60,
                "sin_retiro_registrado": sin_retiro_registrado,
                "fechas_inconsistentes": fechas_inconsistentes,
            },

            "cortes": tarjetas_cortes,
        }

    filas_por_mes: dict[str, list[dict]] = {}

    for fila in filas_universo:
        fecha_ingreso = fila.get("fecha_ingreso")

        if fecha_ingreso is None:
            continue

        clave = (
            f"{fecha_ingreso.year}-"
            f"{str(fecha_ingreso.month).zfill(2)}"
        )

        if clave not in filas_por_mes:
            filas_por_mes[clave] = []

        filas_por_mes[clave].append(fila)

    meses_nombre = {
        1: "enero",
        2: "febrero",
        3: "marzo",
        4: "abril",
        5: "mayo",
        6: "junio",
        7: "julio",
        8: "agosto",
        9: "septiembre",
        10: "octubre",
        11: "noviembre",
        12: "diciembre",
    }

    serie_mensual = []
    detalle_mensual = []

    for clave in sorted(filas_por_mes.keys()):
        filas_mes = filas_por_mes[clave]

        fecha_referencia = filas_mes[0].get("fecha_ingreso")

        if fecha_referencia is None:
            continue

        resumen_mes = construir_resumen_rotacion(
            filas_mes
        )

        item_serie = {
            "clave": clave,
            "anio": fecha_referencia.year,
            "numero_mes": fecha_referencia.month,
            "mes": meses_nombre[
                fecha_referencia.month
            ],
            "etiqueta": (
                f"{meses_nombre[fecha_referencia.month].capitalize()} "
                f"{fecha_referencia.year}"
            ),

            "total_contratados":
                resumen_mes["total_contratados"],

            "tasa_hasta_7":
                resumen_mes["cortes"]["7"]["tasa"],

            "tasa_hasta_15":
                resumen_mes["cortes"]["15"]["tasa"],

            "tasa_hasta_30":
                resumen_mes["cortes"]["30"]["tasa"],

            "tasa_hasta_60":
                resumen_mes["cortes"]["60"]["tasa"],

            "evaluables_7":
                resumen_mes["cortes"]["7"]["evaluables"],

            "evaluables_15":
                resumen_mes["cortes"]["15"]["evaluables"],

            "evaluables_30":
                resumen_mes["cortes"]["30"]["evaluables"],

            "evaluables_60":
                resumen_mes["cortes"]["60"]["evaluables"],

            "retiros_hasta_7":
                resumen_mes["cortes"]["7"]["retiros"],

            "retiros_hasta_15":
                resumen_mes["cortes"]["15"]["retiros"],

            "retiros_hasta_30":
                resumen_mes["cortes"]["30"]["retiros"],

            "retiros_hasta_60":
                resumen_mes["cortes"]["60"]["retiros"],
        }

        serie_mensual.append(item_serie)

        detalle_mensual.append(
            {
                **item_serie,
                "tarjetas": resumen_mes["tarjetas"],
                "rangos_exclusivos":
                    resumen_mes["rangos_exclusivos"],
                "cortes":
                    resumen_mes["cortes"],
            }
        )

    if mes is not None:
        filas_periodo = [
            fila
            for fila in filas_universo
            if fila.get("fecha_ingreso") is not None
            and fila["fecha_ingreso"].month == mes
        ]
    else:
        filas_periodo = list(filas_universo)

    resumen_periodo = construir_resumen_rotacion(
        filas_periodo
    )

    return {
        "filtros": {
            "anio": anio,
            "mes": mes,
            "fecha_inicio_kpi": "2026-03-01",
            "fecha_base": "ContratacionBasica.FechaIngreso",
            "modo": (
                "mensual"
                if mes is not None
                else (
                    "anual"
                    if anio is not None
                    else "global"
                )
            ),
        },

        "kpi": {
            "codigo": "KPI_2_SELECCION",
            "nombre":
                "Tasa de rotación del nuevo personal",
            "descripcion": (
                "Mide la permanencia inicial del personal nuevo "
                "contratado y los retiros ocurridos dentro de los "
                "primeros 7, 15, 30 y 60 días."
            ),
        },

        "total_contratados":
            resumen_periodo["total_contratados"],

        "tarjetas":
            resumen_periodo["tarjetas"],

        "rangos_exclusivos":
            resumen_periodo["rangos_exclusivos"],

        "cortes":
            resumen_periodo["cortes"],

        "serie_mensual":
            serie_mensual,

        "detalle_mensual":
            detalle_mensual,

        "periodo_actual": {
            "anio": anio,
            "mes": mes,
            **resumen_periodo,
        },

        "auditoria": {
            "universo": (
                "ContratacionBasica.FechaIngreso desde 2026-03-01"
            ),
            "fecha_ingreso":
                "ContratacionBasica.FechaIngreso",
            "fecha_retiro_real":
                "PazYSalvoOperaciones.FechaUltimoDiaLaborado",
            "regla_ciclo_laboral": (
                "Si existe un ciclo ACTIVO o EN_PROCESO, el KPI solo "
                "considera retiros con FechaUltimoDiaLaborado igual o "
                "posterior a la FechaIngreso de ese ciclo. Los retiros "
                "anteriores se consideran históricos y no se cruzan con "
                "el reintegro vigente."
            ),
            "cohorte": (
                "El mes corresponde al mes de ingreso, "
                "no al mes de retiro."
            ),
            "cortes": [
                7,
                15,
                30,
                60,
            ],
            "regla_evaluable": (
                "Una persona entra al denominador de un corte "
                "si ya alcanzó ese número de días desde su ingreso "
                "o si ya se retiró dentro de ese mismo corte."
            ),
            "regla_maduracion": (
                "La tasa solo se publica cuando al menos el 80 % "
                "de la cohorte válida ya alcanzó naturalmente el corte. "
                "Los retiros tempranos se contabilizan, pero por sí solos "
                "no hacen madurar el indicador."
            ),
            "umbral_maduracion_porcentaje": 80.0,
            "sin_evaluables": (
                "La tasa se devuelve como null para indicar "
                "pendiente de maduración."
            ),
            "fechas_inconsistentes": (
                "Cuenta registros con FechaUltimoDiaLaborado anterior "
                "a ContratacionBasica.FechaIngreso. Estos registros no "
                "alimentan los rangos de retiro ni las tasas."
            ),
            "excluye_migrados": True,
            "solo_consulta": True,
        },
    }


# ============================================================
# DASHBOARD EXCLUSIVO DE CONTRATACIÓN
# ============================================================
# Este bloque es independiente del dashboard de Selección.
#
# No modifica:
# - Flujo de contratación
# - Endpoint /api/contratado
# - Endpoint /api/rechazo-contratacion
# - Reporte Synergy
# - Excel
# - Drive
# - Google Sheet
#
# Criterios temporales de la consulta general:
# - Registrados: RegistroPersonal.FechaCreacion.
# - Avanzan: fecha real del movimiento al estado 24.
# - Contratados: fecha real del movimiento al estado 25 por botón C.
# - Tiempo promedio: contrataciones finalizadas en el periodo consultado.
#
# Consulta individual:
# - El usuario busca visualmente por nombre o identificación.
# - El frontend conserva internamente IdRegistroPersonal.
# - El detalle individual consulta el historial completo del trabajador.
# ============================================================


def _formatear_duracion_segundos(total_segundos: int | None) -> str | None:
    """Convierte segundos a una duración legible en español."""
    if total_segundos is None:
        return None

    total_segundos = max(total_segundos, 0)
    dias, resto = divmod(total_segundos, 86400)
    horas, resto = divmod(resto, 3600)
    minutos, segundos = divmod(resto, 60)

    partes = []

    if dias > 0:
        partes.append(f"{dias} día" if dias == 1 else f"{dias} días")

    if horas > 0:
        partes.append(f"{horas} hora" if horas == 1 else f"{horas} horas")

    if minutos > 0:
        partes.append(
            f"{minutos} minuto" if minutos == 1 else f"{minutos} minutos"
        )

    # En tiempos de varios días se prioriza una lectura ejecutiva.
    # Para tiempos menores a un día se conservan también los segundos.
    if dias == 0 and (segundos > 0 or not partes):
        partes.append(
            f"{segundos} segundo" if segundos == 1 else f"{segundos} segundos"
        )

    return ", ".join(partes)


def _nombre_estado_proceso(id_estado: int | None) -> str:
    estados = {
        18: "Nuevo",
        19: "Entrevista",
        20: "Entrevista jefe inmediato",
        21: "Exámenes",
        22: "Seguridad",
        24: "Avanza a contratación",
        25: "Contratado",
        26: "Referenciación",
        27: "Desiste del proceso",
        28: "Rechazado",
        30: "Abierto",
        34: "Pendiente de contratación",
    }
    return estados.get(id_estado, f"Estado {id_estado}" if id_estado else "Sin estado")


@router.get("/buscar-trabajadores-contratacion")
def buscar_trabajadores_dashboard_contratacion(
    db: Annotated[Session, Depends(get_db)],
    texto_busqueda: str,
    limite: int = 20,
):
    """
    Busca cualquier trabajador por nombre, apellido o identificación.

    Esta consulta individual permite encontrar trabajadores creados:
    - Por el flujo normal del aplicativo.
    - Por migración histórica.
    - Por ajustes administrativos posteriores.

    Los filtros de exclusión de migración se conservan únicamente en
    los indicadores generales del dashboard de Contratación.
    """
    texto_limpio = (texto_busqueda or "").strip()

    if len(texto_limpio) < 2:
        raise HTTPException(
            status_code=400,
            detail="Escriba al menos 2 caracteres para buscar.",
        )

    limite_seguro = min(max(limite, 1), 50)
    patron = f"%{texto_limpio}%"

    rows = db.execute(
        text("""
            SELECT
                rp."IdRegistroPersonal" AS id_registro_personal,
                TRIM(
                    CONCAT_WS(
                        ' ',
                        NULLIF(TRIM(rp."Nombres"), ''),
                        NULLIF(TRIM(rp."Apellidos"), '')
                    )
                ) AS nombre_completo,
                rp."NumeroIdentificacion"::text AS numero_identificacion,
                rp."IdEstadoProceso" AS id_estado_proceso,
                (
                    EXISTS (
                        SELECT 1
                        FROM public."HistorialLaboral" hl
                        WHERE
                            hl."IdRegistroPersonal" = rp."IdRegistroPersonal"
                            AND UPPER(
                                TRIM(COALESCE(hl."TipoVinculacion", ''))
                            ) = 'ACTIVO MIGRADO'
                    )
                    OR EXISTS (
                        SELECT 1
                        FROM public."ContratacionBasica" cb
                        WHERE
                            cb."IdRegistroPersonal" = rp."IdRegistroPersonal"
                            AND cb."FechaIngreso" IS NOT NULL
                            AND cb."FechaIngreso"::date
                                < rp."FechaCreacion"::date
                    )
                ) AS es_activo_migrado
            FROM public."RegistroPersonal" rp
            WHERE
                (
                    rp."NumeroIdentificacion"::text ILIKE :patron
                    OR CONCAT_WS(
                        ' ',
                        COALESCE(rp."Nombres", ''),
                        COALESCE(rp."Apellidos", '')
                    ) ILIKE :patron
                    OR COALESCE(rp."Nombres", '') ILIKE :patron
                    OR COALESCE(rp."Apellidos", '') ILIKE :patron
                )
            ORDER BY
                CASE
                    WHEN rp."NumeroIdentificacion"::text = :texto_exacto
                    THEN 0
                    ELSE 1
                END,
                nombre_completo ASC,
                rp."IdRegistroPersonal" DESC
            LIMIT :limite;
        """),
        {
            "patron": patron,
            "texto_exacto": texto_limpio,
            "limite": limite_seguro,
        },
    ).mappings().all()

    resultados = []

    for row in rows:
        item = dict(row)

        item["estado_actual"] = _nombre_estado_proceso(
            item.get("id_estado_proceso")
        )

        item["origen_registro"] = (
            "Migración histórica"
            if item.get("es_activo_migrado")
            else "Aplicativo"
        )

        resultados.append(item)

    return {
        "texto_busqueda": texto_limpio,
        "total_resultados": len(resultados),
        "resultados": resultados,
    }


def _obtener_dashboard_contratacion_individual(
    id_registro_personal: int,
    db: Session,
):
    """
    Consulta individual de cualquier trabajador.

    A diferencia de los indicadores generales, esta consulta no excluye:
    - Trabajadores migrados.
    - Usuarios con etiquetas de migración.
    - Registros de ajustes administrativos.

    La fecha de contratación se obtiene en este orden:
    1. HistorialEstadoContratacion, estado 25.
    2. ContratacionBasica.FechaIngreso como respaldo.

    Esta función solo consulta información y no modifica la base de datos.
    """

    trabajador = db.execute(
        text("""
            SELECT
                rp."IdRegistroPersonal" AS id_registro_personal,
                TRIM(
                    CONCAT_WS(
                        ' ',
                        NULLIF(TRIM(rp."Nombres"), ''),
                        NULLIF(TRIM(rp."Apellidos"), '')
                    )
                ) AS nombre_completo,
                rp."NumeroIdentificacion"::text AS numero_identificacion,
                rp."IdEstadoProceso" AS id_estado_proceso,
                rp."FechaCreacion" AS fecha_registro_seleccion,
                rp."UsuarioActualizacion" AS usuario_actualizacion,
                (
                    SELECT MIN(cb."FechaIngreso")::timestamptz
                    FROM public."ContratacionBasica" cb
                    WHERE
                        cb."IdRegistroPersonal" = rp."IdRegistroPersonal"
                        AND cb."FechaIngreso" IS NOT NULL
                ) AS fecha_ingreso_contratacion_basica,
                (
                    EXISTS (
                        SELECT 1
                        FROM public."HistorialLaboral" hl
                        WHERE
                            hl."IdRegistroPersonal" = rp."IdRegistroPersonal"
                            AND UPPER(
                                TRIM(COALESCE(hl."TipoVinculacion", ''))
                            ) = 'ACTIVO MIGRADO'
                    )
                    OR EXISTS (
                        SELECT 1
                        FROM public."ContratacionBasica" cb
                        WHERE
                            cb."IdRegistroPersonal" = rp."IdRegistroPersonal"
                            AND cb."FechaIngreso" IS NOT NULL
                            AND cb."FechaIngreso"::date
                                < rp."FechaCreacion"::date
                    )
                ) AS es_activo_migrado
            FROM public."RegistroPersonal" rp
            WHERE
                rp."IdRegistroPersonal" = :id_registro_personal;
        """),
        {
            "id_registro_personal": id_registro_personal,
        },
    ).mappings().first()

    if not trabajador:
        raise HTTPException(
            status_code=404,
            detail="No se encontró el trabajador consultado.",
        )

    trazabilidad = db.execute(
        text("""
            WITH fecha_25_historial AS (
                SELECT
                    MIN(hec."FechaMovimiento") AS fecha_estado_25
                FROM public."HistorialEstadoContratacion" hec
                WHERE
                    hec."IdRegistroPersonal" = :id_registro_personal
                    AND hec."EstadoNuevo" = 25
            ),
            fecha_contratacion_basica AS (
                SELECT
                    MIN(cb."FechaIngreso")::timestamptz
                        AS fecha_ingreso_contratacion
                FROM public."ContratacionBasica" cb
                WHERE
                    cb."IdRegistroPersonal" = :id_registro_personal
                    AND cb."FechaIngreso" IS NOT NULL
            ),
            fecha_contratacion_final AS (
                SELECT
                    COALESCE(
                        f25.fecha_estado_25,
                        fcb.fecha_ingreso_contratacion
                    ) AS fecha_estado_25,
                    CASE
                        WHEN f25.fecha_estado_25 IS NOT NULL
                        THEN 'HISTORIAL_ESTADO_CONTRATACION'

                        WHEN fcb.fecha_ingreso_contratacion IS NOT NULL
                        THEN 'CONTRATACION_BASICA'

                        ELSE NULL
                    END AS fuente_fecha_contratacion
                FROM fecha_25_historial f25
                CROSS JOIN fecha_contratacion_basica fcb
            ),
            fecha_24 AS (
                SELECT
                    MAX(hec."FechaMovimiento") AS fecha_estado_24
                FROM public."HistorialEstadoContratacion" hec
                CROSS JOIN fecha_contratacion_final fcf
                WHERE
                    hec."IdRegistroPersonal" = :id_registro_personal
                    AND hec."EstadoNuevo" = 24
                    AND (
                        fcf.fecha_estado_25 IS NULL
                        OR hec."FechaMovimiento" <= fcf.fecha_estado_25
                    )
            )
            SELECT
                f24.fecha_estado_24,
                fcf.fecha_estado_25,
                fcf.fuente_fecha_contratacion,
                CASE
                    WHEN f24.fecha_estado_24 IS NOT NULL
                         AND fcf.fecha_estado_25 IS NOT NULL
                         AND fcf.fecha_estado_25 >= f24.fecha_estado_24
                    THEN EXTRACT(
                        EPOCH FROM (
                            fcf.fecha_estado_25 - f24.fecha_estado_24
                        )
                    )
                    ELSE NULL
                END AS tiempo_segundos
            FROM fecha_24 f24
            CROSS JOIN fecha_contratacion_final fcf;
        """),
        {
            "id_registro_personal": id_registro_personal,
        },
    ).mappings().first()

    rechazo = db.execute(
        text("""
            SELECT
                hec."FechaMovimiento" AS fecha_rechazo,
                hec."UsuarioMovimiento" AS usuario_rechazo,
                hec."OrigenMovimiento" AS origen_movimiento,
                CASE
                    WHEN hec."OrigenMovimiento" = 'BOTON_NC'
                    THEN COALESCE(
                        NULLIF(TRIM(orc."ObservacionesRechazo"), ''),
                        'Sin observación'
                    )

                    WHEN hec."OrigenMovimiento"
                        = 'HISTORICO_MOTIVO_CIERRE'
                    THEN COALESCE(
                        NULLIF(TRIM(mcp."MotivoCierre"), ''),
                        'Sin motivo de cierre'
                    )

                    ELSE 'Sin motivo'
                END AS motivo_rechazo
            FROM public."HistorialEstadoContratacion" hec

            LEFT JOIN LATERAL (
                SELECT
                    orc_detalle."ObservacionesRechazo"
                FROM public."ObsRechazoContratacion" orc_detalle
                WHERE
                    orc_detalle."IdRegistroPersonal"
                        = hec."IdRegistroPersonal"
                ORDER BY
                    orc_detalle."IdObsRechazoContratacion" DESC
                LIMIT 1
            ) orc
                ON hec."OrigenMovimiento" = 'BOTON_NC'

            LEFT JOIN LATERAL (
                SELECT
                    mcp_detalle."MotivoCierre"
                FROM public."MotivoCierreProceso" mcp_detalle
                WHERE
                    mcp_detalle."IdRegistroPersonal"
                        = hec."IdRegistroPersonal"
                ORDER BY
                    COALESCE(
                        mcp_detalle."FechaActualizacion",
                        mcp_detalle."FechaCreacion"
                    ) DESC,
                    mcp_detalle."IdMotivoCierre" DESC
                LIMIT 1
            ) mcp
                ON hec."OrigenMovimiento"
                    = 'HISTORICO_MOTIVO_CIERRE'

            WHERE
                hec."IdRegistroPersonal" = :id_registro_personal
                AND hec."EstadoNuevo" = 28
                AND hec."OrigenMovimiento" IN (
                    'BOTON_NC',
                    'HISTORICO_MOTIVO_CIERRE'
                )
                AND hec."Modulo" = 'CONTRATACION'

            ORDER BY
                hec."FechaMovimiento" DESC,
                hec."IdHistorialEstadoContratacion" DESC

            LIMIT 1;
        """),
        {
            "id_registro_personal": id_registro_personal,
        },
    ).mappings().first()

    datos_trabajador = dict(trabajador)

    fecha_estado_24 = (
        trazabilidad.get("fecha_estado_24")
        if trazabilidad
        else None
    )

    fecha_estado_25 = (
        trazabilidad.get("fecha_estado_25")
        if trazabilidad
        else None
    )

    fuente_fecha_contratacion = (
        trazabilidad.get("fuente_fecha_contratacion")
        if trazabilidad
        else None
    )

    tiempo_segundos_raw = (
        trazabilidad.get("tiempo_segundos")
        if trazabilidad
        else None
    )

    tiempo_segundos = (
        round(float(tiempo_segundos_raw))
        if tiempo_segundos_raw is not None
        else None
    )

    id_estado_actual = datos_trabajador.get("id_estado_proceso")
    es_activo_migrado = bool(
        datos_trabajador.get("es_activo_migrado")
    )

    fecha_ingreso_historica = datos_trabajador.get(
        "fecha_ingreso_contratacion_basica"
    )

    # En un trabajador histórico, los movimientos 24 y 25 reconstruidos
    # posteriormente no representan su proceso real dentro del aplicativo.
    # Por eso se conserva la fecha real de ContratacionBasica, se elimina
    # el avance ficticio y no se calcula tiempo de contratación.
    if es_activo_migrado:
        fecha_estado_24 = None
        tiempo_segundos = None

        if fecha_ingreso_historica is not None:
            fecha_estado_25 = fecha_ingreso_historica
            fuente_fecha_contratacion = "CONTRATACION_BASICA"

    fecha_rechazo = (
        rechazo.get("fecha_rechazo")
        if rechazo
        else None
    )

    usuario_rechazo = (
        rechazo.get("usuario_rechazo")
        if rechazo
        else None
    )

    origen_movimiento_rechazo = (
        rechazo.get("origen_movimiento")
        if rechazo
        else None
    )

    motivo_rechazo = (
        rechazo.get("motivo_rechazo")
        if rechazo
        else None
    )

    fue_rechazado = rechazo is not None
    tiene_contratacion = fecha_estado_25 is not None

    confirmada_por_historial = (
        fuente_fecha_contratacion
        == "HISTORIAL_ESTADO_CONTRATACION"
    )

    fecha_registro_seleccion = datos_trabajador.get(
        "fecha_registro_seleccion"
    )

    origen_registro = (
        "Migración histórica"
        if es_activo_migrado
        else "Aplicativo"
    )

    if fuente_fecha_contratacion == "CONTRATACION_BASICA":
        mensaje_fecha_contratacion = (
            "La fecha de contratación fue obtenida de "
            "ContratacionBasica.FechaIngreso porque el trabajador "
            "no tiene un movimiento disponible al estado 25."
        )
    elif fuente_fecha_contratacion == "HISTORIAL_ESTADO_CONTRATACION":
        mensaje_fecha_contratacion = (
            "La fecha de contratación fue obtenida del historial "
            "de estados del trabajador."
        )
    else:
        mensaje_fecha_contratacion = (
            "No se encontró una fecha de contratación en el historial "
            "ni en ContratacionBasica."
        )

    if es_activo_migrado:
        mensaje_tiempo = (
            "No se calcula porque el trabajador corresponde a información "
            "histórica y no recorrió dentro del aplicativo el flujo 24 a 25."
        )
    elif tiempo_segundos is not None:
        mensaje_tiempo = (
            "Tiempo real del trabajador entre el avance a contratación "
            "y su fecha de contratación."
        )
    elif tiene_contratacion and fecha_estado_24 is None:
        mensaje_tiempo = (
            "Existe fecha de contratación, pero no existe una fecha "
            "histórica verificable de avance al estado 24; por eso no "
            "es posible calcular el tiempo de contratación."
        )
    else:
        mensaje_tiempo = (
            "El trabajador todavía no tiene una trazabilidad completa "
            "entre el avance a contratación y la contratación."
        )

    return {
        "modo_consulta": "individual",

        "trabajador": {
            "id_registro_personal": datos_trabajador.get(
                "id_registro_personal"
            ),
            "nombre_completo": datos_trabajador.get(
                "nombre_completo"
            ),
            "numero_identificacion": datos_trabajador.get(
                "numero_identificacion"
            ),
            "id_estado_actual": id_estado_actual,
            "estado_actual": _nombre_estado_proceso(
                id_estado_actual
            ),
            "origen_registro": origen_registro,
            "es_activo_migrado": es_activo_migrado,
            "usuario_actualizacion": datos_trabajador.get(
                "usuario_actualizacion"
            ),
        },

        "registro_seleccion": {
            "existe": fecha_registro_seleccion is not None,
            "fecha": fecha_registro_seleccion,
            "nota": (
                "En trabajadores migrados esta fecha puede corresponder "
                "al momento de incorporación histórica al aplicativo."
                if es_activo_migrado
                else
                "Fecha de creación del registro en Selección."
            ),
        },

        "avance_contratacion": {
            "existe": (
                fecha_estado_24 is not None
                and not es_activo_migrado
            ),
            "fecha": (
                fecha_estado_24
                if not es_activo_migrado
                else None
            ),
            "fuente": (
                "HistorialEstadoContratacion"
                if fecha_estado_24 is not None
                else None
            ),
        },

        "contratacion": {
            "existe": tiene_contratacion,
            "fecha": fecha_estado_25,
            "fuente": fuente_fecha_contratacion,
            "confirmada_por_historial": confirmada_por_historial,
            "obtenida_de_contratacion_basica": (
                fuente_fecha_contratacion == "CONTRATACION_BASICA"
            ),
            "mensaje": mensaje_fecha_contratacion,
        },

        "rechazo": {
            "existe": fue_rechazado,
            "fecha": fecha_rechazo,
            "usuario": usuario_rechazo,
            "origen_movimiento": origen_movimiento_rechazo,
            "motivo": motivo_rechazo,
            "fuente": "HistorialEstadoContratacion",
        },

        "tiempo_contratacion": {
            "total_segundos": (
                tiempo_segundos
                if not es_activo_migrado
                else None
            ),
            "total_minutos": (
                round(tiempo_segundos / 60, 2)
                if tiempo_segundos is not None
                and not es_activo_migrado
                else None
            ),
            "formateado": (
                _formatear_duracion_segundos(tiempo_segundos)
                if not es_activo_migrado
                else None
            ),
            "disponible": (
                tiempo_segundos is not None
                and not es_activo_migrado
            ),
            "fecha_inicio": fecha_estado_24,
            "fecha_fin": fecha_estado_25,
            "fuente_inicio": (
                "Estado 24 - HistorialEstadoContratacion"
                if fecha_estado_24 is not None
                else None
            ),
            "fuente_fin": fuente_fecha_contratacion,
            "mensaje": mensaje_tiempo,
        },

        "linea_tiempo": [
            {
                "evento": "Registro en Selección",
                "completado": fecha_registro_seleccion is not None,
                "fecha": fecha_registro_seleccion,
                "fuente": "RegistroPersonal.FechaCreacion",
            },
            {
                "evento": "Avanza a Contratación",
                "completado": fecha_estado_24 is not None,
                "fecha": fecha_estado_24,
                "fuente": (
                    "HistorialEstadoContratacion"
                    if fecha_estado_24 is not None
                    else None
                ),
            },
            {
                "evento": "Contratado",
                "completado": tiene_contratacion,
                "fecha": fecha_estado_25,
                "fuente": fuente_fecha_contratacion,
            },
            {
                "evento": "Rechazado en Contratación",
                "completado": fue_rechazado,
                "fecha": fecha_rechazo,
                "usuario": usuario_rechazo,
                "origen_movimiento": origen_movimiento_rechazo,
                "motivo": motivo_rechazo,
            },
        ],

        "alcance_consulta": {
            "incluye_personal_migrado": True,
            "incluye_personal_aplicativo": True,
            "afecta_indicadores_generales": False,
            "modifica_base_datos": False,
        },

        "exclusiones": {
            "activo_migrado_historial_laboral": False,
            "usuarios_migracion": False,
            "ajuste_no_activos_maestro": False,
            "nota": (
                "Las exclusiones de migración se aplican únicamente "
                "a los indicadores generales, no a la consulta individual."
            ),
        },
    }

@router.get("/dashboard-contratacion")
def obtener_dashboard_contratacion(
    db: Annotated[Session, Depends(get_db)],
    anio: int | None = None,
    mes: int | None = None,
    id_registro_personal: int | None = None,
):
    """
    Dashboard independiente para el módulo de Contratación.

    Reglas de la consulta general:
    - Registrados por Selección: RegistroPersonal.FechaCreacion.
    - Avanzan a Contratación:
      1. movimiento real al estado 24;
      2. rechazo real registrado por Contratación, cuya transición confirma
         que la persona se encontraba en el estado 24; o
      3. contratación real confirmada para uno de los seis casos operativos
         que quedaron sin historial por una falla anterior.
    - Contratados:
      1. movimiento real al estado 25 mediante BOTON_C; o
      2. uno de los seis casos operativos sin historial, usando
         ContratacionBasica.FechaIngreso.
    - Rechazados en Contratación:
      únicamente personas con movimiento real desde el estado 24
      al estado 28. Quienes permanecen en estado 24 no se cuentan
      como rechazados.
    - Cohorte del periodo:
      personas que avanzaron al estado 24 dentro del periodo consultado.
      El mes de pertenencia queda definido por la fecha del avance.
    - Contratados de la cohorte:
      integrantes que, después del avance, presentan evidencia de contratación
      mediante estado 25 o ContratacionBasica.FechaIngreso. Permanecen como
      contratados aunque posteriormente pasen a Retiros u otro estado.
    - Rechazados de la cohorte:
      integrantes que presentan rechazo después del avance y nunca alcanzaron
      contratación.
    - Pendientes de la cohorte:
      integrantes que todavía no presentan evidencia de contratación ni rechazo.
    - Contrataciones finalizadas en el periodo:
      movimientos reales al estado 25 ocurridos dentro del periodo,
      aunque el avance al estado 24 haya sucedido en un periodo anterior.
    - Tiempo promedio:
      contrataciones finalizadas en el periodo con trazabilidad real 24 a 25.

    No se reconstruyen movimientos históricos en la base de datos.
    Los casos legacy provenientes de Achill y los registros de prueba
    no se agregan como contratados sin historial.

    Consulta individual:
    - Se recibe IdRegistroPersonal internamente después de que el usuario
      selecciona un resultado buscado por nombre o identificación.
    - Se consulta el historial completo del trabajador, sin limitar sus
      eventos por año o mes.
    """
    if mes is not None and (mes < 1 or mes > 12):
        raise HTTPException(
            status_code=400,
            detail="El mes debe estar entre 1 y 12.",
        )

    if anio is not None and anio < 2000:
        raise HTTPException(
            status_code=400,
            detail="El año consultado no es válido.",
        )

    if id_registro_personal is not None:
        if id_registro_personal <= 0:
            raise HTTPException(
                status_code=400,
                detail="El IdRegistroPersonal consultado no es válido.",
            )
        return _obtener_dashboard_contratacion_individual(
            id_registro_personal=id_registro_personal,
            db=db,
        )

    # Casos reales contratados mediante el aplicativo que quedaron sin
    # movimientos 24 y 25 en HistorialEstadoContratacion.
    #
    # Se identifican por documento para no depender de IdRegistroPersonal
    # diferentes entre QA y producción.
    documentos_contratados_sin_historial = (
        "1043634001",  # Luis Fernando Martínez Babilonia
        "1108834963",  # Cristian Giovanny Tocarema Lugo
        "1104435573",  # Eliana Toledo Ruz
        "45549624",    # María Marcela Romaña Mosquera
        "1003213604",  # Lorena Paola Puente García
        "1016105764",  # José de Jesús Conde Ospino
    )

    resultado = db.execute(
        text("""
            WITH parametros_periodo AS (
                SELECT
                    CASE
                        WHEN :anio IS NOT NULL AND :mes IS NOT NULL
                        THEN (
                            MAKE_TIMESTAMPTZ(
                                :anio,
                                :mes,
                                1,
                                0,
                                0,
                                0,
                                'America/Bogota'
                            ) + INTERVAL '1 month'
                        )
                        WHEN :anio IS NOT NULL
                        THEN (
                            MAKE_TIMESTAMPTZ(
                                :anio,
                                1,
                                1,
                                0,
                                0,
                                0,
                                'America/Bogota'
                            ) + INTERVAL '1 year'
                        )
                        ELSE CURRENT_TIMESTAMP
                    END AS fecha_corte
            ),
            universo_base AS (
                SELECT
                    rp."IdRegistroPersonal",
                    rp."IdEstadoProceso",
                    rp."NumeroIdentificacion"::text
                        AS numero_identificacion,
                    rp."FechaCreacion",
                    rp."UsuarioActualizacion"
                FROM public."RegistroPersonal" rp
                WHERE
                    NOT EXISTS (
                        SELECT 1
                        FROM public."HistorialLaboral" hl
                        WHERE
                            hl."IdRegistroPersonal" = rp."IdRegistroPersonal"
                            AND UPPER(
                                TRIM(COALESCE(hl."TipoVinculacion", ''))
                            ) = 'ACTIVO MIGRADO'
                    )
                    AND LOWER(
                        COALESCE(rp."UsuarioActualizacion", '')
                    ) NOT LIKE '%migracion%'
                    AND LOWER(
                        COALESCE(rp."UsuarioActualizacion", '')
                    ) NOT LIKE '%migrado%'
                    AND COALESCE(rp."UsuarioActualizacion", '')
                        <> 'ajuste_no_activos_maestro_2026_06_22'
            ),
            registrados_periodo AS (
                SELECT ub."IdRegistroPersonal"
                FROM universo_base ub
                WHERE
                    ub."FechaCreacion"
                        >= TIMESTAMPTZ '2026-03-01 00:00:00-05'
                    AND (
                        :anio IS NULL
                        OR EXTRACT(YEAR FROM ub."FechaCreacion") = :anio
                    )
                    AND (
                        :mes IS NULL
                        OR EXTRACT(MONTH FROM ub."FechaCreacion") = :mes
                    )
            ),
            avances_historial_periodo AS (
                SELECT DISTINCT
                    hec."IdRegistroPersonal",
                    hec."FechaMovimiento" AS fecha_avance,
                    'HISTORIAL_ESTADO_24'::text AS fuente_avance
                FROM public."HistorialEstadoContratacion" hec
                INNER JOIN universo_base ub
                    ON ub."IdRegistroPersonal" = hec."IdRegistroPersonal"
                WHERE
                    hec."EstadoNuevo" = 24
                    AND hec."FechaMovimiento"
                        >= TIMESTAMPTZ '2026-03-01 00:00:00-05'
                    AND (
                        :anio IS NULL
                        OR EXTRACT(YEAR FROM hec."FechaMovimiento") = :anio
                    )
                    AND (
                        :mes IS NULL
                        OR EXTRACT(MONTH FROM hec."FechaMovimiento") = :mes
                    )
            ),
            rechazos_contratacion_periodo AS (
                SELECT DISTINCT ON (hec."IdRegistroPersonal")
                    hec."IdRegistroPersonal",
                    hec."FechaMovimiento" AS fecha_avance,
                    'RECHAZO_CONTRATACION'::text AS fuente_avance
                FROM public."HistorialEstadoContratacion" hec
                INNER JOIN universo_base ub
                    ON ub."IdRegistroPersonal" = hec."IdRegistroPersonal"
                WHERE
                    hec."EstadoNuevo" = 28
                    AND UPPER(TRIM(COALESCE(hec."Modulo", '')))
                        = 'CONTRATACION'
                    AND hec."FechaMovimiento"
                        >= TIMESTAMPTZ '2026-03-01 00:00:00-05'
                    AND (
                        :anio IS NULL
                        OR EXTRACT(YEAR FROM hec."FechaMovimiento") = :anio
                    )
                    AND (
                        :mes IS NULL
                        OR EXTRACT(MONTH FROM hec."FechaMovimiento") = :mes
                    )
                ORDER BY
                    hec."IdRegistroPersonal",
                    hec."FechaMovimiento" ASC,
                    hec."IdHistorialEstadoContratacion" ASC
            ),
            contrataciones_historial_periodo AS (
                SELECT DISTINCT ON (hec25."IdRegistroPersonal")
                    hec25."IdRegistroPersonal",
                    inicio.fecha_estado_24,
                    hec25."FechaMovimiento" AS fecha_estado_25,
                    'HISTORIAL_ESTADO_25'::text AS fuente_contratacion
                FROM public."HistorialEstadoContratacion" hec25
                INNER JOIN universo_base ub
                    ON ub."IdRegistroPersonal" = hec25."IdRegistroPersonal"
                LEFT JOIN LATERAL (
                    SELECT MAX(hec24."FechaMovimiento") AS fecha_estado_24
                    FROM public."HistorialEstadoContratacion" hec24
                    WHERE
                        hec24."IdRegistroPersonal"
                            = hec25."IdRegistroPersonal"
                        AND hec24."EstadoNuevo" = 24
                        AND hec24."FechaMovimiento"
                            <= hec25."FechaMovimiento"
                ) inicio ON TRUE
                WHERE
                    hec25."EstadoNuevo" = 25
                    AND hec25."FechaMovimiento"
                        >= TIMESTAMPTZ '2026-03-01 00:00:00-05'
                    AND (
                        :anio IS NULL
                        OR EXTRACT(YEAR FROM hec25."FechaMovimiento") = :anio
                    )
                    AND (
                        :mes IS NULL
                        OR EXTRACT(MONTH FROM hec25."FechaMovimiento") = :mes
                    )
                ORDER BY
                    hec25."IdRegistroPersonal",
                    hec25."FechaMovimiento" ASC,
                    hec25."IdHistorialEstadoContratacion" ASC
            ),
            contrataciones_sin_historial_periodo AS (
                SELECT DISTINCT ON (ub."IdRegistroPersonal")
                    ub."IdRegistroPersonal",
                    NULL::timestamptz AS fecha_estado_24,
                    cb."FechaIngreso"::timestamptz AS fecha_estado_25,
                    'CONTRATACION_BASICA_SIN_HISTORIAL'::text
                        AS fuente_contratacion
                FROM universo_base ub
                INNER JOIN public."ContratacionBasica" cb
                    ON cb."IdRegistroPersonal" = ub."IdRegistroPersonal"
                WHERE
                    ub.numero_identificacion IN (
                        :doc_sin_historial_1,
                        :doc_sin_historial_2,
                        :doc_sin_historial_3,
                        :doc_sin_historial_4,
                        :doc_sin_historial_5,
                        :doc_sin_historial_6
                    )
                    AND cb."FechaIngreso" IS NOT NULL
                    AND cb."FechaIngreso"::date >= DATE '2026-03-01'
                    AND NOT EXISTS (
                        SELECT 1
                        FROM public."HistorialEstadoContratacion" hec25
                        WHERE
                            hec25."IdRegistroPersonal"
                                = ub."IdRegistroPersonal"
                            AND hec25."EstadoNuevo" = 25
                    )
                    AND (
                        :anio IS NULL
                        OR EXTRACT(YEAR FROM cb."FechaIngreso") = :anio
                    )
                    AND (
                        :mes IS NULL
                        OR EXTRACT(MONTH FROM cb."FechaIngreso") = :mes
                    )
                ORDER BY
                    ub."IdRegistroPersonal",
                    cb."FechaIngreso" ASC
            ),
            contrataciones_periodo AS (
                SELECT
                    "IdRegistroPersonal",
                    fecha_estado_24,
                    fecha_estado_25,
                    fuente_contratacion
                FROM contrataciones_historial_periodo

                UNION ALL

                SELECT
                    "IdRegistroPersonal",
                    fecha_estado_24,
                    fecha_estado_25,
                    fuente_contratacion
                FROM contrataciones_sin_historial_periodo
            ),
            avances_sin_historial_periodo AS (
                SELECT
                    c."IdRegistroPersonal",
                    c.fecha_estado_25 AS fecha_avance,
                    'CONTRATACION_REAL_SIN_HISTORIAL'::text
                        AS fuente_avance
                FROM contrataciones_sin_historial_periodo c
            ),
            avances_periodo AS (
                SELECT DISTINCT ON ("IdRegistroPersonal")
                    "IdRegistroPersonal",
                    fecha_avance,
                    fuente_avance
                FROM (
                    SELECT
                        "IdRegistroPersonal",
                        fecha_avance,
                        fuente_avance
                    FROM avances_historial_periodo

                    UNION ALL

                    SELECT
                        "IdRegistroPersonal",
                        fecha_avance,
                        fuente_avance
                    FROM rechazos_contratacion_periodo

                    UNION ALL

                    SELECT
                        "IdRegistroPersonal",
                        fecha_avance,
                        fuente_avance
                    FROM avances_sin_historial_periodo
                ) avances_unificados
                ORDER BY
                    "IdRegistroPersonal",
                    fecha_avance ASC
            ),
            resultado_final_cohorte AS (
                SELECT
                    ap."IdRegistroPersonal",
                    ap.fecha_avance,
                    ap.fuente_avance,
                    (
                        SELECT MIN(cb."FechaIngreso")::timestamptz
                        FROM public."ContratacionBasica" cb
                        WHERE
                            cb."IdRegistroPersonal" = ap."IdRegistroPersonal"
                            AND cb."FechaIngreso" IS NOT NULL
                    ) AS fecha_ingreso,
                    (
                        SELECT MIN(hec25."FechaMovimiento")
                        FROM public."HistorialEstadoContratacion" hec25
                        WHERE
                            hec25."IdRegistroPersonal" = ap."IdRegistroPersonal"
                            AND hec25."EstadoNuevo" = 25
                            AND hec25."FechaMovimiento" >= ap.fecha_avance
                    ) AS fecha_estado_25,
                    (
                        SELECT MIN(hec28."FechaMovimiento")
                        FROM public."HistorialEstadoContratacion" hec28
                        WHERE
                            hec28."IdRegistroPersonal" = ap."IdRegistroPersonal"
                            AND hec28."EstadoNuevo" = 28
                            AND hec28."FechaMovimiento" >= ap.fecha_avance
                    ) AS fecha_estado_28
                FROM avances_periodo ap
            ),
            decisiones_cohorte AS (
                SELECT
                    rfc."IdRegistroPersonal",
                    rfc.fecha_avance,
                    rfc.fuente_avance,
                    CASE
                        WHEN (
                            rfc.fuente_avance
                                = 'CONTRATACION_REAL_SIN_HISTORIAL'
                            OR rfc.fecha_ingreso IS NOT NULL
                            OR rfc.fecha_estado_25 IS NOT NULL
                        )
                        THEN 25

                        WHEN (
                            rfc.fuente_avance = 'RECHAZO_CONTRATACION'
                            OR rfc.fecha_estado_28 IS NOT NULL
                        )
                        THEN 28

                        ELSE NULL
                    END AS estado_decision,
                    CASE
                        WHEN (
                            rfc.fuente_avance
                                = 'CONTRATACION_REAL_SIN_HISTORIAL'
                            OR rfc.fecha_ingreso IS NOT NULL
                            OR rfc.fecha_estado_25 IS NOT NULL
                        )
                        THEN COALESCE(
                            rfc.fecha_estado_25,
                            rfc.fecha_ingreso,
                            rfc.fecha_avance
                        )

                        WHEN (
                            rfc.fuente_avance = 'RECHAZO_CONTRATACION'
                            OR rfc.fecha_estado_28 IS NOT NULL
                        )
                        THEN COALESCE(
                            rfc.fecha_estado_28,
                            rfc.fecha_avance
                        )

                        ELSE NULL
                    END AS fecha_decision
                FROM resultado_final_cohorte rfc
            ),
            contratados_cohorte AS (
                SELECT "IdRegistroPersonal"
                FROM decisiones_cohorte
                WHERE estado_decision = 25
            ),
            rechazados_cohorte AS (
                SELECT "IdRegistroPersonal", fecha_decision AS fecha_rechazo
                FROM decisiones_cohorte
                WHERE estado_decision = 28
            ),
            pendientes_contratacion_periodo AS (
                SELECT "IdRegistroPersonal"
                FROM decisiones_cohorte
                WHERE estado_decision IS NULL
            ),
            casos_tiempo_medible AS (
                SELECT
                    fecha_estado_24,
                    fecha_estado_25
                FROM contrataciones_periodo
                WHERE
                    fecha_estado_24 IS NOT NULL
                    AND fecha_estado_25 IS NOT NULL
                    AND fecha_estado_25 >= fecha_estado_24
            )
            SELECT
                (SELECT COUNT(*) FROM registrados_periodo)
                    AS registrados_seleccion,
                (SELECT COUNT(*) FROM avances_periodo)
                    AS avanzan_contratacion,
                (SELECT COUNT(*) FROM contrataciones_periodo)
                    AS contrataciones_finalizadas_periodo,
                (SELECT COUNT(*) FROM contratados_cohorte)
                    AS contratados_cohorte,
                (SELECT COUNT(*) FROM pendientes_contratacion_periodo)
                    AS pendientes_contratacion,
                (
                    SELECT AVG(
                        EXTRACT(
                            EPOCH FROM (
                                fecha_estado_25 - fecha_estado_24
                            )
                        )
                    )
                    FROM casos_tiempo_medible
                ) AS promedio_segundos,
                (SELECT COUNT(*) FROM casos_tiempo_medible)
                    AS casos_tiempo_medidos,
                (SELECT COUNT(*) FROM rechazados_cohorte)
                    AS rechazados_contratacion,
                (
                    SELECT COUNT(*)
                    FROM contrataciones_sin_historial_periodo
                ) AS contratados_sin_historial;
        """),
        {
            "anio": anio,
            "mes": mes,
            "doc_sin_historial_1": documentos_contratados_sin_historial[0],
            "doc_sin_historial_2": documentos_contratados_sin_historial[1],
            "doc_sin_historial_3": documentos_contratados_sin_historial[2],
            "doc_sin_historial_4": documentos_contratados_sin_historial[3],
            "doc_sin_historial_5": documentos_contratados_sin_historial[4],
            "doc_sin_historial_6": documentos_contratados_sin_historial[5],
        },
    ).mappings().first()

    registrados_seleccion = int(
        resultado.get("registrados_seleccion") or 0
    )
    avanzan_contratacion = int(
        resultado.get("avanzan_contratacion") or 0
    )
    contrataciones_finalizadas_periodo = int(
        resultado.get("contrataciones_finalizadas_periodo") or 0
    )
    contratados_cohorte = int(
        resultado.get("contratados_cohorte") or 0
    )
    pendientes_contratacion = int(
        resultado.get("pendientes_contratacion") or 0
    )
    contratados_sin_historial = int(
        resultado.get("contratados_sin_historial") or 0
    )
    rechazados_contratacion = int(
        resultado.get("rechazados_contratacion") or 0
    )
    casos_tiempo_medidos = int(
        resultado.get("casos_tiempo_medidos") or 0
    )

    promedio_segundos_raw = resultado.get("promedio_segundos")
    promedio_segundos = (
        round(float(promedio_segundos_raw))
        if promedio_segundos_raw is not None
        else None
    )
    promedio_minutos = (
        round(promedio_segundos / 60, 2)
        if promedio_segundos is not None
        else None
    )
    promedio_formateado = _formatear_duracion_segundos(promedio_segundos)

    motivos_rows = db.execute(
        text("""
            WITH parametros_periodo AS (
                SELECT
                    CASE
                        WHEN :anio IS NOT NULL AND :mes IS NOT NULL
                        THEN (
                            MAKE_TIMESTAMPTZ(
                                :anio,
                                :mes,
                                1,
                                0,
                                0,
                                0,
                                'America/Bogota'
                            ) + INTERVAL '1 month'
                        )
                        WHEN :anio IS NOT NULL
                        THEN (
                            MAKE_TIMESTAMPTZ(
                                :anio,
                                1,
                                1,
                                0,
                                0,
                                0,
                                'America/Bogota'
                            ) + INTERVAL '1 year'
                        )
                        ELSE CURRENT_TIMESTAMP
                    END AS fecha_corte
            ),
            universo_base AS (
                SELECT
                    rp."IdRegistroPersonal",
                    rp."NumeroIdentificacion"::text
                        AS numero_identificacion,
                    rp."UsuarioActualizacion"
                FROM public."RegistroPersonal" rp
                WHERE
                    NOT EXISTS (
                        SELECT 1
                        FROM public."HistorialLaboral" hl
                        WHERE
                            hl."IdRegistroPersonal" = rp."IdRegistroPersonal"
                            AND UPPER(
                                TRIM(COALESCE(hl."TipoVinculacion", ''))
                            ) = 'ACTIVO MIGRADO'
                    )
                    AND LOWER(
                        COALESCE(rp."UsuarioActualizacion", '')
                    ) NOT LIKE '%migracion%'
                    AND LOWER(
                        COALESCE(rp."UsuarioActualizacion", '')
                    ) NOT LIKE '%migrado%'
                    AND COALESCE(rp."UsuarioActualizacion", '')
                        <> 'ajuste_no_activos_maestro_2026_06_22'
            ),
            avances_historial_periodo AS (
                SELECT DISTINCT ON (hec."IdRegistroPersonal")
                    hec."IdRegistroPersonal",
                    hec."FechaMovimiento" AS fecha_avance,
                    'HISTORIAL_ESTADO_24'::text AS fuente_avance
                FROM public."HistorialEstadoContratacion" hec
                INNER JOIN universo_base ub
                    ON ub."IdRegistroPersonal" = hec."IdRegistroPersonal"
                WHERE
                    hec."EstadoNuevo" = 24
                    AND hec."FechaMovimiento"
                        >= TIMESTAMPTZ '2026-03-01 00:00:00-05'
                    AND (
                        :anio IS NULL
                        OR EXTRACT(YEAR FROM hec."FechaMovimiento") = :anio
                    )
                    AND (
                        :mes IS NULL
                        OR EXTRACT(MONTH FROM hec."FechaMovimiento") = :mes
                    )
                ORDER BY
                    hec."IdRegistroPersonal",
                    hec."FechaMovimiento" ASC,
                    hec."IdHistorialEstadoContratacion" ASC
            ),
            rechazos_contratacion_periodo AS (
                SELECT DISTINCT ON (hec."IdRegistroPersonal")
                    hec."IdRegistroPersonal",
                    hec."FechaMovimiento" AS fecha_avance,
                    'RECHAZO_CONTRATACION'::text AS fuente_avance
                FROM public."HistorialEstadoContratacion" hec
                INNER JOIN universo_base ub
                    ON ub."IdRegistroPersonal" = hec."IdRegistroPersonal"
                WHERE
                    hec."EstadoNuevo" = 28
                    AND UPPER(TRIM(COALESCE(hec."Modulo", '')))
                        = 'CONTRATACION'
                    AND hec."FechaMovimiento"
                        >= TIMESTAMPTZ '2026-03-01 00:00:00-05'
                    AND (
                        :anio IS NULL
                        OR EXTRACT(YEAR FROM hec."FechaMovimiento") = :anio
                    )
                    AND (
                        :mes IS NULL
                        OR EXTRACT(MONTH FROM hec."FechaMovimiento") = :mes
                    )
                ORDER BY
                    hec."IdRegistroPersonal",
                    hec."FechaMovimiento" ASC,
                    hec."IdHistorialEstadoContratacion" ASC
            ),
            avances_sin_historial_periodo AS (
                SELECT DISTINCT ON (ub."IdRegistroPersonal")
                    ub."IdRegistroPersonal",
                    cb."FechaIngreso"::timestamptz AS fecha_avance,
                    'CONTRATACION_REAL_SIN_HISTORIAL'::text
                        AS fuente_avance
                FROM universo_base ub
                INNER JOIN public."ContratacionBasica" cb
                    ON cb."IdRegistroPersonal" = ub."IdRegistroPersonal"
                WHERE
                    ub.numero_identificacion IN (
                        :doc_sin_historial_1,
                        :doc_sin_historial_2,
                        :doc_sin_historial_3,
                        :doc_sin_historial_4,
                        :doc_sin_historial_5,
                        :doc_sin_historial_6
                    )
                    AND cb."FechaIngreso" IS NOT NULL
                    AND cb."FechaIngreso"::date >= DATE '2026-03-01'
                    AND NOT EXISTS (
                        SELECT 1
                        FROM public."HistorialEstadoContratacion" hec25
                        WHERE
                            hec25."IdRegistroPersonal"
                                = ub."IdRegistroPersonal"
                            AND hec25."EstadoNuevo" = 25
                    )
                    AND (
                        :anio IS NULL
                        OR EXTRACT(YEAR FROM cb."FechaIngreso") = :anio
                    )
                    AND (
                        :mes IS NULL
                        OR EXTRACT(MONTH FROM cb."FechaIngreso") = :mes
                    )
                ORDER BY
                    ub."IdRegistroPersonal",
                    cb."FechaIngreso" ASC
            ),
            avances_periodo AS (
                SELECT DISTINCT ON ("IdRegistroPersonal")
                    "IdRegistroPersonal",
                    fecha_avance,
                    fuente_avance
                FROM (
                    SELECT * FROM avances_historial_periodo
                    UNION ALL
                    SELECT * FROM rechazos_contratacion_periodo
                    UNION ALL
                    SELECT * FROM avances_sin_historial_periodo
                ) avances_unificados
                ORDER BY
                    "IdRegistroPersonal",
                    fecha_avance ASC
            ),
            resultado_final_cohorte AS (
                SELECT
                    ap."IdRegistroPersonal",
                    ap.fecha_avance,
                    ap.fuente_avance,
                    (
                        SELECT MIN(cb."FechaIngreso")::timestamptz
                        FROM public."ContratacionBasica" cb
                        WHERE
                            cb."IdRegistroPersonal" = ap."IdRegistroPersonal"
                            AND cb."FechaIngreso" IS NOT NULL
                    ) AS fecha_ingreso,
                    (
                        SELECT MIN(hec25."FechaMovimiento")
                        FROM public."HistorialEstadoContratacion" hec25
                        WHERE
                            hec25."IdRegistroPersonal" = ap."IdRegistroPersonal"
                            AND hec25."EstadoNuevo" = 25
                            AND hec25."FechaMovimiento" >= ap.fecha_avance
                    ) AS fecha_estado_25,
                    (
                        SELECT MIN(hec28."FechaMovimiento")
                        FROM public."HistorialEstadoContratacion" hec28
                        WHERE
                            hec28."IdRegistroPersonal" = ap."IdRegistroPersonal"
                            AND hec28."EstadoNuevo" = 28
                            AND hec28."FechaMovimiento" >= ap.fecha_avance
                    ) AS fecha_estado_28
                FROM avances_periodo ap
            ),
            decisiones_cohorte AS (
                SELECT
                    rfc."IdRegistroPersonal",
                    CASE
                        WHEN (
                            rfc.fuente_avance
                                = 'CONTRATACION_REAL_SIN_HISTORIAL'
                            OR rfc.fecha_ingreso IS NOT NULL
                            OR rfc.fecha_estado_25 IS NOT NULL
                        )
                        THEN 25

                        WHEN (
                            rfc.fuente_avance = 'RECHAZO_CONTRATACION'
                            OR rfc.fecha_estado_28 IS NOT NULL
                        )
                        THEN 28

                        ELSE NULL
                    END AS estado_decision
                FROM resultado_final_cohorte rfc
            ),
            rechazados_periodo AS (
                SELECT "IdRegistroPersonal"
                FROM decisiones_cohorte
                WHERE estado_decision = 28
            ),
            rechazos_con_motivo AS (
                SELECT
                    rp."IdRegistroPersonal",
                    COALESCE(
                        NULLIF(TRIM(orc."ObservacionesRechazo"), ''),
                        NULLIF(TRIM(mcp."MotivoCierre"), ''),
                        'Sin motivo registrado'
                    ) AS motivo
                FROM rechazados_periodo rp
                LEFT JOIN LATERAL (
                    SELECT orc_detalle."ObservacionesRechazo"
                    FROM public."ObsRechazoContratacion" orc_detalle
                    WHERE
                        orc_detalle."IdRegistroPersonal"
                            = rp."IdRegistroPersonal"
                    ORDER BY
                        orc_detalle."IdObsRechazoContratacion" DESC
                    LIMIT 1
                ) orc ON TRUE
                LEFT JOIN LATERAL (
                    SELECT mcp_detalle."MotivoCierre"
                    FROM public."MotivoCierreProceso" mcp_detalle
                    WHERE
                        mcp_detalle."IdRegistroPersonal"
                            = rp."IdRegistroPersonal"
                    ORDER BY
                        COALESCE(
                            mcp_detalle."FechaActualizacion",
                            mcp_detalle."FechaCreacion"
                        ) DESC,
                        mcp_detalle."IdMotivoCierre" DESC
                    LIMIT 1
                ) mcp ON TRUE
            )
            SELECT
                motivo,
                COUNT(*) AS cantidad
            FROM rechazos_con_motivo
            GROUP BY motivo
            ORDER BY cantidad DESC, motivo ASC;
        """),
        {
            "anio": anio,
            "mes": mes,
            "doc_sin_historial_1": documentos_contratados_sin_historial[0],
            "doc_sin_historial_2": documentos_contratados_sin_historial[1],
            "doc_sin_historial_3": documentos_contratados_sin_historial[2],
            "doc_sin_historial_4": documentos_contratados_sin_historial[3],
            "doc_sin_historial_5": documentos_contratados_sin_historial[4],
            "doc_sin_historial_6": documentos_contratados_sin_historial[5],
        },
    ).mappings().all()

    motivos_rechazo_contratacion = []
    for motivo_row in motivos_rows:
        cantidad = int(motivo_row.get("cantidad") or 0)
        porcentaje = (
            round((cantidad / rechazados_contratacion) * 100, 2)
            if rechazados_contratacion > 0
            else 0
        )
        motivos_rechazo_contratacion.append({
            "motivo": motivo_row.get("motivo"),
            "cantidad": cantidad,
            "porcentaje": porcentaje,
        })

    porcentaje_contratados_sobre_avanzan = (
        round(
            (contratados_cohorte / avanzan_contratacion) * 100,
            2,
        )
        if avanzan_contratacion > 0
        else 0
    )

    suma_resultados_cohorte = (
        contratados_cohorte
        + rechazados_contratacion
        + pendientes_contratacion
    )
    cohorte_consistente = (
        suma_resultados_cohorte == avanzan_contratacion
    )
    contratados_periodos_anteriores = max(
        contrataciones_finalizadas_periodo - contratados_cohorte,
        0,
    )

    return {
        "modo_consulta": "general",
        "filtros": {"anio": anio, "mes": mes},
        "fecha_inicio_aplicativo": "2026-03-01",
        "criterio_fecha": {
            "registrados_seleccion": "RegistroPersonal.FechaCreacion",
            "avanzan_contratacion": (
                "Fecha del estado 24; para los seis contratados reales "
                "sin historial se usa ContratacionBasica.FechaIngreso"
            ),
            "contratados": (
                "Personas de la cohorte con evidencia final de contratación "
                "mediante estado 25 o ContratacionBasica.FechaIngreso"
            ),
            "rechazados_contratacion": (
                "Personas de la cohorte con rechazo posterior al avance "
                "que nunca alcanzaron contratación"
            ),
            "pendientes_contratacion": (
                "Personas de la cohorte que aún no presentan contratación "
                "ni rechazo después del avance"
            ),
            "tiempo_contratacion": (
                "Solo casos con fechas reales de estados 24 y 25"
            ),
        },
        "registrados_seleccion": registrados_seleccion,
        "comparativo_registrados_contratados": {
            "registrados_seleccion": registrados_seleccion,
            "contratados": contrataciones_finalizadas_periodo,
            "porcentaje_contratados": None,
            "aplica_porcentaje": False,
            "nota": (
                "Comparativo informativo. No se calcula porcentaje porque "
                "los registrados en Selección y las contrataciones finalizadas "
                "pueden corresponder a cohortes diferentes."
            ),
        },
        "comparativo_avanzan_contratados": {
            "avanzan_contratacion": avanzan_contratacion,
            "contratados": contratados_cohorte,
            "porcentaje_contratados": porcentaje_contratados_sobre_avanzan,
            "fuente": (
                "HistorialEstadoContratacion y ContratacionBasica "
                "para seis excepciones verificadas"
            ),
            "nota": (
                "Este comparativo usa una sola cohorte: personas que "
                "avanzaron dentro del periodo. El resultado refleja el "
                "desenlace final conocido de cada integrante."
            ),
        },
        "contrataciones_finalizadas_periodo": {
            "total": contratados_cohorte,
            "de_cohorte_periodo": contratados_cohorte,
            "total_operativo_periodo": contrataciones_finalizadas_periodo,
            "provenientes_periodos_anteriores": (
                contratados_periodos_anteriores
            ),
            "nota": (
                "El campo total corresponde a los contratados de la misma "
                "cohorte que avanzó en el periodo, para mantener coherencia "
                "entre Avanzan, Contratados, Rechazados y Pendientes. "
                "total_operativo_periodo conserva el número de contrataciones "
                "finalizadas físicamente dentro del periodo."
            ),
        },
        "auditoria_cohorte": {
            "avanzan": avanzan_contratacion,
            "contratados": contratados_cohorte,
            "rechazados": rechazados_contratacion,
            "pendientes": pendientes_contratacion,
            "suma_resultados": suma_resultados_cohorte,
            "es_consistente": cohorte_consistente,
        },
        "pendientes_contratacion": {
            "total": pendientes_contratacion,
            "fuente": (
                "Cohorte de personas que avanzaron en el periodo y aún "
                "no presentan evidencia de contratación ni rechazo"
            ),
            "criterio": (
                "Trabajadores de la cohorte sin desenlace final conocido"
            ),
            "estado_actual": 24,
            "modifica_base_datos": False,
        },
        "rechazados": {
            "total": rechazados_contratacion,
            "fuente": (
                "HistorialEstadoContratacion: transición real 24 a 28"
            ),
            "motivos": motivos_rechazo_contratacion,
            "pendiente_fuente_contratacion": False,
            "filtro_fecha": (
                "Avance al estado 24 dentro del periodo y rechazo posterior "
                "sin evidencia previa o posterior de contratación"
            ),
            "origenes_incluidos": [
                "ESTADO_ANTERIOR_24",
                "ESTADO_NUEVO_28",
            ],
        },
        "tiempo_contratacion": {
            "promedio_minutos": promedio_minutos,
            "promedio_segundos": promedio_segundos,
            "promedio_formateado": promedio_formateado,
            "casos_medidos": casos_tiempo_medidos,
            "disponible": promedio_segundos is not None,
            "fuente_inicio": "Estado 24",
            "fuente_fin": "Estado 25",
            "criterio_periodo": (
                "Fecha de contratación dentro del periodo consultado"
            ),
            "mensaje": (
                "Tiempo promedio real entre el avance a contratación "
                "y la contratación, calculado solo con trazabilidad completa."
                if promedio_segundos is not None
                else
                "No existen casos con trazabilidad completa 24 a 25 "
                "en el periodo seleccionado."
            ),
        },
        "auditoria_contratados_sin_historial": {
            "cantidad_periodo": contratados_sin_historial,
            "documentos_configurados": len(
                documentos_contratados_sin_historial
            ),
            "fuente": "ContratacionBasica.FechaIngreso",
            "modifica_base_datos": False,
        },
        "exclusiones": {
            "fecha_anterior_inicio_aplicativo": True,
            "fecha_inicio_aplicativo": "2026-03-01",
            "activo_migrado_historial_laboral": True,
            "usuarios_migracion": True,
            "ajuste_no_activos_maestro": True,
            "legacy_achill_no_reconstruido": True,
            "registros_prueba_no_reconstruidos": True,
        },
    }


@router.get("/reporte-excel")
def generar_reporte_excel_seleccion(db: Annotated[Session, Depends(get_db)]):
    rows = db.execute(text("""
        SELECT
            rp."Nombres",
            rp."Apellidos",
            rp."NumeroIdentificacion" as cedula,
            COALESCE(rp."Celular", '') as telefono,
            COALESCE(rp."Email", '') as correo,
            cg."NombreCargo" as cargo,
            rp."FechaCreacion" as fecha_registro,
            mcp."MotivoCierre" as motivo_rechazo,
            CASE rp."IdEstadoProceso"
                WHEN 18 THEN 'Nuevo'
                WHEN 19 THEN 'Entrevista'
                WHEN 20 THEN 'Entrevista jefe inmediato'
                WHEN 21 THEN 'Exámenes'
                WHEN 22 THEN 'Seguridad'
                WHEN 24 THEN 'Avanza a contratación'
                WHEN 25 THEN 'Contratado'
                WHEN 26 THEN 'Referenciación'
                WHEN 27 THEN 'Desiste del proceso'
                WHEN 28 THEN 'Rechazado'
                WHEN 30 THEN 'Abierto'
                WHEN 34 THEN 'Pendiente de Contratación'
                ELSE CONCAT('Estado ', rp."IdEstadoProceso")
            END as estado
        FROM public."RegistroPersonal" rp
        LEFT JOIN public."AsignacionCargoCliente" acc
            ON acc."IdRegistroPersonal" = rp."IdRegistroPersonal"
        LEFT JOIN public."Cargo" cg
            ON cg."IdCargo" = acc."IdCargo"
        LEFT JOIN (
            SELECT DISTINCT ON ("IdRegistroPersonal")
                "IdRegistroPersonal",
                "MotivoCierre",
                "FechaCreacion"
            FROM public."MotivoCierreProceso"
            ORDER BY "IdRegistroPersonal", "FechaCreacion" DESC
        ) mcp
            ON mcp."IdRegistroPersonal" = rp."IdRegistroPersonal"
    """)).mappings().all()

    filas = [dict(r) for r in rows]

    if not filas:
        raise HTTPException(status_code=404, detail="No hay datos")

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    headers = list(filas[0].keys())
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill

    for fila in filas:
        fila_limpia = {}
        for key, value in fila.items():
            if isinstance(value, datetime):
                value = value.replace(tzinfo=None)
            fila_limpia[key] = value

        ws.append(list(fila_limpia.values()))

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = max_length + 3

    total = len(filas)
    rechazados = len([f for f in filas if str(f.get("estado", "")).strip().lower() == "rechazado"])
    contratados = len([f for f in filas if str(f.get("estado", "")).strip().lower() == "contratado"])
    avanza = len([f for f in filas if str(f.get("estado", "")).strip().lower() == "avanza a contratación"])

    porcentaje_contratacion = round((contratados / total) * 100) if total else 0
    porcentaje_rechazo = round((rechazados / total) * 100) if total else 0

    motivos_rechazo = {}

    for fila in filas:
        estado = str(fila.get("estado", "")).strip().lower()
        motivo = fila.get("motivo_rechazo")

        if estado == "rechazado" and motivo and str(motivo).strip().upper() != "SIN_MOTIVO":
            motivo = str(motivo).strip()
            motivos_rechazo[motivo] = motivos_rechazo.get(motivo, 0) + 1

    tendencia_mensual = {}

    meses = {
        1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
        5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
        9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
    }

    for fila in filas:
        fecha = fila.get("fecha_registro")

        if isinstance(fecha, datetime):
            clave_orden = fecha.strftime("%Y-%m")
            etiqueta_mes = f"{meses[fecha.month]}-{str(fecha.year)[-2:]}"
        else:
            clave_orden = "9999-99"
            etiqueta_mes = "Sin fecha"

        if clave_orden not in tendencia_mensual:
            tendencia_mensual[clave_orden] = {
                "etiqueta": etiqueta_mes,
                "cantidad": 0
            }

        tendencia_mensual[clave_orden]["cantidad"] += 1

    tendencia_mensual = dict(sorted(tendencia_mensual.items()))

    ws_dash = wb.create_sheet("Dashboard")

    ws_dash["A1"] = "Dashboard Selección"
    ws_dash["A1"].font = Font(bold=True, size=18)

    dash_fill = PatternFill("solid", fgColor="70AD47")
    dash_font = Font(bold=True, color="FFFFFF")
    kpi_fill = PatternFill("solid", fgColor="E2F0D9")
    kpi_title_font = Font(bold=True, color="375623", size=11)
    kpi_value_font = Font(bold=True, color="000000", size=16)

    ws_dash["A2"] = "Total candidatos"
    ws_dash["B2"] = total
    ws_dash["D2"] = "Contratación"
    ws_dash["E2"] = f"{porcentaje_contratacion}%"
    ws_dash["G2"] = "Rechazo"
    ws_dash["H2"] = f"{porcentaje_rechazo}%"

    for cell_ref in ["A2", "D2", "G2"]:
        ws_dash[cell_ref].fill = kpi_fill
        ws_dash[cell_ref].font = kpi_title_font

    for cell_ref in ["B2", "E2", "H2"]:
        ws_dash[cell_ref].fill = kpi_fill
        ws_dash[cell_ref].font = kpi_value_font

    ws_dash["A4"] = "Resumen general"
    ws_dash["A4"].font = Font(bold=True, size=12)

    ws_dash["A5"] = "Métrica"
    ws_dash["B5"] = "Valor"

    for col in ["A", "B"]:
        cell = ws_dash[f"{col}5"]
        cell.fill = dash_fill
        cell.font = dash_font

    ws_dash["A6"] = "Total registros"
    ws_dash["B6"] = total

    ws_dash["A7"] = "Rechazados"
    ws_dash["B7"] = rechazados

    ws_dash["A8"] = "Avanza a contratación"
    ws_dash["B8"] = avanza

    ws_dash["A9"] = "Contratados"
    ws_dash["B9"] = contratados

    ws_dash["J2"] = "Registros por mes"
    ws_dash["J2"].font = Font(bold=True, size=12)

    ws_dash["J3"] = "Mes"
    ws_dash["K3"] = "Registros"

    for col in ["J", "K"]:
        cell = ws_dash[f"{col}3"]
        cell.fill = dash_fill
        cell.font = dash_font

    fila_tendencia = 4
    for item in tendencia_mensual.values():
        ws_dash[f"J{fila_tendencia}"] = item["etiqueta"]
        ws_dash[f"K{fila_tendencia}"] = item["cantidad"]
        fila_tendencia += 1

    if fila_tendencia > 5:
        line = LineChart()
        line.style = 2
        line.title = "Evolución mensual de candidatos registrados"
        line.y_axis.title = "Cantidad de registros"
        line.x_axis.title = "Mes"

        data_line = Reference(ws_dash, min_col=11, min_row=4, max_row=fila_tendencia - 1)
        labels_line = Reference(ws_dash, min_col=10, min_row=4, max_row=fila_tendencia - 1)

        line.add_data(data_line, titles_from_data=False, from_rows=False)
        line.set_categories(labels_line)

        line.series = line.series[:1]
        line.series[0].graphicalProperties.line.solidFill = "70AD47"
        line.series[0].graphicalProperties.line.width = 30000
        line.series[0].marker.symbol = "circle"
        line.series[0].marker.size = 8
        line.series[0].marker.graphicalProperties.solidFill = "70AD47"

        line.dLbls = DataLabelList()
        line.dLbls.showVal = False
        line.dLbls.showSerName = False
        line.dLbls.showCatName = False

        line.legend = None
        line.smooth = False
        line.height = 8
        line.width = 22

        ws_dash.add_chart(line, "J10")

    pie = PieChart()
    pie.title = "Distribución de candidatos en estados finales"

    data = Reference(ws_dash, min_col=2, min_row=7, max_row=9)
    labels = Reference(ws_dash, min_col=1, min_row=7, max_row=9)

    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)

    pie.dataLabels = DataLabelList()
    pie.dataLabels.showVal = False
    pie.dataLabels.showPercent = True
    pie.dataLabels.showCatName = False
    pie.dataLabels.showSerName = False
    pie.dataLabels.showLeaderLines = True

    ws_dash.add_chart(pie, "A12")

    ws_dash["A33"] = "Distribución de candidatos en estados finales"
    ws_dash["A33"].font = Font(bold=True, size=14)

    estados_barra = [
        {"estado": "Rechazados", "cantidad": rechazados, "color": "4F81BD"},
        {"estado": "Avanza a contratación", "cantidad": avanza, "color": "C0504D"},
        {"estado": "Contratados", "cantidad": contratados, "color": "9BBB59"},
    ]

    estados_barra = sorted(estados_barra, key=lambda x: x["cantidad"], reverse=True)

    ws_dash["A34"] = "Estado"
    ws_dash["B34"] = "Cantidad"
    ws_dash["C34"] = "%"

    for col in ["A", "B", "C"]:
        cell = ws_dash[f"{col}34"]
        cell.fill = dash_fill
        cell.font = dash_font

    fila_barra = 35
    for item in estados_barra:
        porcentaje = round((item["cantidad"] / total) * 100) if total else 0
        ws_dash[f"A{fila_barra}"] = item["estado"]
        ws_dash[f"B{fila_barra}"] = item["cantidad"]
        ws_dash[f"C{fila_barra}"] = f"{porcentaje}%"
        fila_barra += 1

    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = None
    bar.y_axis.title = None
    bar.x_axis.title = None

    data_bar = Reference(ws_dash, min_col=2, min_row=34, max_row=37)
    labels_bar = Reference(ws_dash, min_col=1, min_row=35, max_row=37)

    bar.add_data(data_bar, titles_from_data=True)
    bar.set_categories(labels_bar)

    bar.legend = None

    bar.dLbls = DataLabelList()
    bar.dLbls.showVal = True
    bar.dLbls.showSerName = False
    bar.dLbls.showCatName = False
    bar.dLbls.showLegendKey = False

    bar.height = 8
    bar.width = 16

    bar.y_axis.majorGridlines = None
    bar.x_axis.majorGridlines = None

    if bar.series:
        for idx, item in enumerate(estados_barra):
            try:
                bar.series[0].data_points[idx].graphicalProperties.solidFill = item["color"]
            except (AttributeError, IndexError) as error:
                print(f"No fue posible aplicar el color de la gráfica: {error}")

    ws_dash.add_chart(bar, "D33")

    ws_dash["A50"] = "Motivos de rechazo"
    ws_dash["A50"].font = Font(bold=True, size=14)

    ws_dash["A51"] = "Motivo"
    ws_dash["B51"] = "Cantidad"
    ws_dash["C51"] = "%"
    ws_dash["D51"] = "Ref."

    for col in ["A", "B", "C", "D"]:
        cell = ws_dash[f"{col}51"]
        cell.fill = dash_fill
        cell.font = dash_font

    motivos_base = [
        "Desiste del Proceso",
        "No Cumple Perfil",
        "No asiste a Examenes Medicos",
        "Examenes No Aptos",
        "Documentacion Incompleta",
        "No asiste a Contratacion",
    ]

    equivalencias_motivos = {
        "DESISTE DEL PROCESO": "Desiste del Proceso",
        "NO CUMPLE PERFIL": "No Cumple Perfil",
        "NO ASISTE A EXAMENES MEDICOS": "No asiste a Examenes Medicos",
        "NO ASISTE A EXÁMENES MEDICOS": "No asiste a Examenes Medicos",
        "NO ASISTE A EXÁMENES MÉDICOS": "No asiste a Examenes Medicos",
        "EXAMENES NO APTOS": "Examenes No Aptos",
        "EXÁMENES NO APTOS": "Examenes No Aptos",
        "DOCUMENTACION INCOMPLETA": "Documentacion Incompleta",
        "DOCUMENTACIÓN INCOMPLETA": "Documentacion Incompleta",
        "NO ASISTE A CONTRATACION": "No asiste a Contratacion",
        "NO ASISTE A CONTRATACIÓN": "No asiste a Contratacion",
    }

    colores_motivos = {
        "Desiste del Proceso": "C0504D",
        "No Cumple Perfil": "4F81BD",
        "No asiste a Examenes Medicos": "F79646",
        "Examenes No Aptos": "8064A2",
        "Documentacion Incompleta": "9BBB59",
        "No asiste a Contratacion": "4BACC6",
    }

    motivos_normalizados = {motivo: 0 for motivo in motivos_base}

    for motivo, cantidad in motivos_rechazo.items():
        clave = str(motivo).strip().upper()
        motivo_final = equivalencias_motivos.get(clave, str(motivo).strip())

        if motivo_final not in motivos_normalizados:
            motivos_normalizados[motivo_final] = 0

        motivos_normalizados[motivo_final] += cantidad

    motivos_tabla = []
    for motivo, cantidad in motivos_normalizados.items():
        motivos_tabla.append({
            "motivo": motivo,
            "cantidad": cantidad,
            "color": colores_motivos.get(motivo, "70AD47")
        })

    motivos_tabla = sorted(
        motivos_tabla,
        key=lambda x: (x["cantidad"] == 0, -x["cantidad"], x["motivo"])
    )

    total_motivos = sum(item["cantidad"] for item in motivos_tabla)

    fila_motivo = 52
    for item in motivos_tabla:
        porcentaje = round((item["cantidad"] / total_motivos) * 100) if total_motivos else 0

        ws_dash[f"A{fila_motivo}"] = item["motivo"]
        ws_dash[f"B{fila_motivo}"] = item["cantidad"]
        ws_dash[f"C{fila_motivo}"] = f"{porcentaje}%"
        ws_dash[f"D{fila_motivo}"] = ""

        ws_dash[f"D{fila_motivo}"].fill = PatternFill("solid", fgColor=item["color"])

        fila_motivo += 1

    fila_grafica_inicio = fila_motivo + 2
    ws_dash[f"A{fila_grafica_inicio}"] = "Motivo grafica"
    ws_dash[f"B{fila_grafica_inicio}"] = "Cantidad"

    fila_grafica = fila_grafica_inicio + 1
    motivos_grafica = [item for item in motivos_tabla if item["cantidad"] > 0]

    for item in motivos_grafica:
        ws_dash[f"A{fila_grafica}"] = item["motivo"]
        ws_dash[f"B{fila_grafica}"] = item["cantidad"]
        fila_grafica += 1

    if motivos_grafica:
        bar_motivos = BarChart()
        bar_motivos.type = "bar"
        bar_motivos.style = 10
        bar_motivos.title = None
        bar_motivos.y_axis.title = None
        bar_motivos.x_axis.title = None

        data_motivos = Reference(
            ws_dash,
            min_col=2,
            min_row=fila_grafica_inicio,
            max_row=fila_grafica - 1
        )

        labels_motivos = Reference(
            ws_dash,
            min_col=1,
            min_row=fila_grafica_inicio + 1,
            max_row=fila_grafica - 1
        )

        bar_motivos.add_data(data_motivos, titles_from_data=True)
        bar_motivos.set_categories(labels_motivos)

        bar_motivos.legend = None

        bar_motivos.dLbls = DataLabelList()
        bar_motivos.dLbls.showVal = True
        bar_motivos.dLbls.showSerName = False
        bar_motivos.dLbls.showCatName = False
        bar_motivos.dLbls.showLegendKey = False

        bar_motivos.height = 8
        bar_motivos.width = 16

        bar_motivos.y_axis.majorGridlines = None
        bar_motivos.x_axis.majorGridlines = None

        if bar_motivos.series:
            for idx, item in enumerate(motivos_grafica):
                try:
                    bar_motivos.series[0].data_points[idx].graphicalProperties.solidFill = item["color"]
                except (AttributeError, IndexError) as error:
                    print(f"No fue posible aplicar el color de la gráfica: {error}")

        ws_dash.add_chart(bar_motivos, "E50")

    ws_dash.column_dimensions["A"].width = 34
    ws_dash.column_dimensions["B"].width = 14
    ws_dash.column_dimensions["C"].width = 10
    ws_dash.column_dimensions["D"].width = 8
    ws_dash.column_dimensions["E"].width = 14
    ws_dash.column_dimensions["G"].width = 14
    ws_dash.column_dimensions["H"].width = 14
    ws_dash.column_dimensions["J"].width = 18
    ws_dash.column_dimensions["K"].width = 14

    ruta = Path("reporte_seleccion_backend.xlsx")
    wb.save(ruta)

    return FileResponse(
        path=ruta,
        filename="reporte_seleccion.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@router.get("/{id_registro_personal}", response_model=DatosSeleccionResponse)
def obtener_datos_seleccion(
    id_registro_personal: int,
    db: Annotated[Session, Depends(get_db)],
):
    data = service.obtener_por_registro_personal(db, id_registro_personal)
    if not data:
        raise HTTPException(status_code=404, detail="No existen datos")

    if getattr(data, "FechaActualizacion", None) is None:
        data.FechaActualizacion = datetime.now(timezone.utc)

    return data


@router.post("/upsert", response_model=DatosSeleccionResponse)
def upsert_datos_seleccion(
    body: DatosSeleccionUpsertRequest,
    db: Annotated[Session, Depends(get_db)],
):
    payload = body.model_dump(exclude_none=True)

    if "HaTrabajadoAntesEnLaEmpresa" in payload:
        parsed = _parse_bool(payload.get("HaTrabajadoAntesEnLaEmpresa"))
        if parsed is None:
            payload.pop("HaTrabajadoAntesEnLaEmpresa", None)
        else:
            payload["HaTrabajadoAntesEnLaEmpresa"] = parsed

    payload.pop("HaTrabajadoAntes", None)

    data = service.upsert(db, payload)

    if getattr(data, "FechaActualizacion", None) is None:
        data.FechaActualizacion = datetime.now(timezone.utc)

    return data