from collections import Counter
from datetime import date, datetime
from io import BytesIO
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from infrastructure.db.deps import get_db


router = APIRouter(
    prefix="/api/rrll-excel",
    tags=["RRLL Excel"]
)


def _quote_identifier(identifier: str) -> str:
    """
    Protege nombres de tablas y columnas obtenidos desde information_schema.
    No recibe valores escritos por el usuario.
    """
    return '"' + str(identifier).replace('"', '""') + '"'


def _obtener_columnas_tabla(db: Session, nombre_tabla: str) -> dict:
    """
    Retorna las columnas reales de una tabla pública, conservando mayúsculas
    y minúsculas. La llave del diccionario queda normalizada en minúsculas.
    """
    filas = db.execute(
        text("""
            SELECT column_name
            FROM information_schema.columns
            WHERE table_schema = 'public'
              AND table_name = :nombre_tabla
        """),
        {"nombre_tabla": nombre_tabla},
    ).mappings().all()

    return {
        str(fila["column_name"]).lower(): str(fila["column_name"])
        for fila in filas
    }


def _buscar_columna(columnas: dict, candidatos: list[str]):
    for candidato in candidatos:
        columna_real = columnas.get(str(candidato).lower())
        if columna_real:
            return columna_real
    return None


def _normalizar_numero_identificacion(valor) -> str:
    return str(valor or "").strip().replace(".", "").replace(" ", "")


def _convertir_a_fecha(valor):
    """
    Convierte fechas provenientes de PostgreSQL o cadenas comunes a date.
    Si el valor no es válido, retorna None.
    """
    if valor is None or valor == "":
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    texto_fecha = str(valor).strip()
    if not texto_fecha:
        return None

    try:
        return date.fromisoformat(texto_fecha[:10])
    except ValueError:
        pass

    try:
        dia, mes, anio = texto_fecha[:10].split("/")
        return date(int(anio), int(mes), int(dia))
    except (TypeError, ValueError):
        return None


def _completar_total_tiempo_trabajo(resultados):
    """
    Conserva el total que ya venga calculado por la función SQL.

    Solo cuando está vacío y existen fecha de ingreso y fecha de retiro,
    calcula los días transcurridos como:

        fecha_retiro - fecha_ingreso

    Ejemplo:
        2026-07-10 a 2026-07-11 = 1 día.

    No actualiza la base de datos.
    """
    resultados_mutables = [dict(row) for row in resultados]

    for row in resultados_mutables:
        total_actual = row.get("total_tiempo_de_trabajo")

        if total_actual not in (None, ""):
            continue

        fecha_ingreso = _convertir_a_fecha(row.get("fecha_ingreso"))
        fecha_retiro = _convertir_a_fecha(row.get("fecha_retiro"))

        if not fecha_ingreso or not fecha_retiro:
            continue

        dias_trabajados = (fecha_retiro - fecha_ingreso).days

        if dias_trabajados >= 0:
            row["total_tiempo_de_trabajo"] = dias_trabajados

    return resultados_mutables


def _completar_fechas_ingreso_migrados(db: Session, resultados):
    """
    Completa únicamente las fechas de ingreso que la función principal dejó
    vacías. Prioridad:

    1. ContratacionBasica.
    2. HistorialLaboral, incluyendo registros ACTIVO MIGRADO.
    3. Tabla histórica migracionactivossynergy, si existe y contiene columnas
       identificables de documento y fecha de ingreso.

    No modifica información en base de datos.
    """
    resultados_mutables = [dict(row) for row in resultados]

    numeros_faltantes = {
        _normalizar_numero_identificacion(row.get("numero_identificacion"))
        for row in resultados_mutables
        if not row.get("fecha_ingreso")
        and _normalizar_numero_identificacion(row.get("numero_identificacion"))
    }

    if not numeros_faltantes:
        return resultados_mutables

    fechas_por_documento = {}

    columnas_rp = _obtener_columnas_tabla(db, "RegistroPersonal")
    rp_id = _buscar_columna(columnas_rp, ["IdRegistroPersonal"])
    rp_documento = _buscar_columna(
        columnas_rp,
        ["NumeroIdentificacion", "NumeroDocumento", "Documento"],
    )

    if rp_id and rp_documento:
        fuentes_relacionadas = [
            (
                "ContratacionBasica",
                ["IdRegistroPersonal"],
                ["FechaIngreso", "FechaInicio", "FechaIngresoEmpresa"],
                None,
            ),
            (
                "HistorialLaboral",
                ["IdRegistroPersonal"],
                ["FechaIngreso", "FechaInicio", "FechaVinculacion"],
                ["FechaActualizacion", "FechaCreacion", "IdHistorialLaboral"],
            ),
        ]

        for nombre_tabla, candidatos_fk, candidatos_fecha, candidatos_orden in fuentes_relacionadas:
            columnas = _obtener_columnas_tabla(db, nombre_tabla)
            if not columnas:
                continue

            fk = _buscar_columna(columnas, candidatos_fk)
            fecha = _buscar_columna(columnas, candidatos_fecha)

            if not fk or not fecha:
                continue

            columna_orden = _buscar_columna(columnas, candidatos_orden or [])
            expresion_documento = (
                "REPLACE(REPLACE(TRIM(CAST("
                f"rp.{_quote_identifier(rp_documento)} AS TEXT"
                ")), '.', ''), ' ', '')"
            )

            columna_orden_real = columna_orden or fecha

            orden_sql = (
                "ORDER BY "
                f"{expresion_documento}, "
                f"fuente.{_quote_identifier(columna_orden_real)} DESC NULLS LAST"
            )

            consulta = text(f"""
                SELECT DISTINCT ON (
                    {expresion_documento}
                )
                    {expresion_documento} AS numero_identificacion,
                    fuente.{_quote_identifier(fecha)} AS fecha_ingreso
                FROM public.{_quote_identifier("RegistroPersonal")} rp
                INNER JOIN public.{_quote_identifier(nombre_tabla)} fuente
                    ON fuente.{_quote_identifier(fk)}
                     = rp.{_quote_identifier(rp_id)}
                WHERE REPLACE(REPLACE(TRIM(CAST(
                        rp.{_quote_identifier(rp_documento)} AS TEXT
                    )), '.', ''), ' ', '') = ANY(:numeros)
                  AND fuente.{_quote_identifier(fecha)} IS NOT NULL
                {orden_sql}
            """)

            filas = db.execute(
                consulta,
                {"numeros": list(numeros_faltantes)},
            ).mappings().all()

            for fila in filas:
                numero = _normalizar_numero_identificacion(
                    fila.get("numero_identificacion")
                )
                if numero and numero not in fechas_por_documento:
                    fechas_por_documento[numero] = fila.get("fecha_ingreso")

    # Último respaldo: fuente histórica usada en la migración.
    columnas_migracion = _obtener_columnas_tabla(db, "migracionactivossynergy")

    if columnas_migracion:
        columna_documento = _buscar_columna(
            columnas_migracion,
            [
                "NumeroIdentificacion",
                "NumeroDocumento",
                "Documento",
                "Cedula",
                "Identificacion",
                "CC",
            ],
        )
        columna_fecha = _buscar_columna(
            columnas_migracion,
            [
                "FechaIngreso",
                "FechaInicio",
                "FechaVinculacion",
                "FechaIngresoEmpresa",
                "FechaDeIngreso",
            ],
        )

        if columna_documento and columna_fecha:
            consulta_migracion = text(f"""
                SELECT DISTINCT ON (
                    REPLACE(REPLACE(TRIM(CAST(
                        {_quote_identifier(columna_documento)} AS TEXT
                    )), '.', ''), ' ', '')
                )
                    REPLACE(REPLACE(TRIM(CAST(
                        {_quote_identifier(columna_documento)} AS TEXT
                    )), '.', ''), ' ', '') AS numero_identificacion,
                    {_quote_identifier(columna_fecha)} AS fecha_ingreso
                FROM public.{_quote_identifier("migracionactivossynergy")}
                WHERE REPLACE(REPLACE(TRIM(CAST(
                        {_quote_identifier(columna_documento)} AS TEXT
                    )), '.', ''), ' ', '') = ANY(:numeros)
                  AND {_quote_identifier(columna_fecha)} IS NOT NULL
                ORDER BY
                    REPLACE(REPLACE(TRIM(CAST(
                        {_quote_identifier(columna_documento)} AS TEXT
                    )), '.', ''), ' ', ''),
                    {_quote_identifier(columna_fecha)} DESC NULLS LAST
            """)

            filas_migracion = db.execute(
                consulta_migracion,
                {"numeros": list(numeros_faltantes)},
            ).mappings().all()

            for fila in filas_migracion:
                numero = _normalizar_numero_identificacion(
                    fila.get("numero_identificacion")
                )
                if numero and numero not in fechas_por_documento:
                    fechas_por_documento[numero] = fila.get("fecha_ingreso")

    for row in resultados_mutables:
        if row.get("fecha_ingreso"):
            continue

        numero = _normalizar_numero_identificacion(
            row.get("numero_identificacion")
        )
        fecha_encontrada = fechas_por_documento.get(numero)

        if fecha_encontrada:
            row["fecha_ingreso"] = fecha_encontrada

    return resultados_mutables


def _agregar_motivo_oficial_validado_rrll(db: Session, resultados):
    """
    Agrega la validación oficial registrada por RRLL en una columna separada.

    La descripción original del trabajador se conserva en:
        descripcion_motivo_especifico_del_retiro

    La validación oficial de RRLL se agrega en:
        motivo_oficial_validado_rrll

    No modifica información en la base de datos.
    """
    resultados_mutables = [dict(row) for row in resultados]

    numeros = {
        _normalizar_numero_identificacion(row.get("numero_identificacion"))
        for row in resultados_mutables
        if _normalizar_numero_identificacion(row.get("numero_identificacion"))
    }

    if not numeros:
        return resultados_mutables

    filas_validacion = db.execute(
        text("""
            SELECT
                REPLACE(
                    REPLACE(
                        TRIM(CAST(rp."NumeroIdentificacion" AS TEXT)),
                        '.',
                        ''
                    ),
                    ' ',
                    ''
                ) AS numero_identificacion,
                rl."FechaRetiro" AS fecha_retiro,
                rl."DescripcionRetiroRRLL" AS descripcion_retiro_rrll,
                rl."IdRetiroLaboral" AS id_retiro_laboral
            FROM public."RetiroLaboral" rl
            INNER JOIN public."RegistroPersonal" rp
                ON rp."IdRegistroPersonal" = rl."IdRegistroPersonal"
            WHERE REPLACE(
                    REPLACE(
                        TRIM(CAST(rp."NumeroIdentificacion" AS TEXT)),
                        '.',
                        ''
                    ),
                    ' ',
                    ''
                ) = ANY(:numeros)
              AND NULLIF(
                    BTRIM(COALESCE(rl."DescripcionRetiroRRLL", '')),
                    ''
                  ) IS NOT NULL
            ORDER BY
                numero_identificacion,
                rl."FechaRetiro" DESC NULLS LAST,
                rl."IdRetiroLaboral" DESC
        """),
        {"numeros": list(numeros)},
    ).mappings().all()

    validacion_por_documento_y_fecha = {}
    validacion_mas_reciente_por_documento = {}

    for fila in filas_validacion:
        numero = _normalizar_numero_identificacion(
            fila.get("numero_identificacion")
        )

        descripcion = str(
            fila.get("descripcion_retiro_rrll") or ""
        ).strip()

        if not numero or not descripcion:
            continue

        fecha_retiro = _convertir_a_fecha(fila.get("fecha_retiro"))

        if fecha_retiro:
            clave = (numero, fecha_retiro)

            if clave not in validacion_por_documento_y_fecha:
                validacion_por_documento_y_fecha[clave] = descripcion

        if numero not in validacion_mas_reciente_por_documento:
            validacion_mas_reciente_por_documento[numero] = descripcion

    for row in resultados_mutables:
        numero = _normalizar_numero_identificacion(
            row.get("numero_identificacion")
        )

        fecha_retiro = _convertir_a_fecha(row.get("fecha_retiro"))

        descripcion_rrll = None

        if numero and fecha_retiro:
            descripcion_rrll = validacion_por_documento_y_fecha.get(
                (numero, fecha_retiro)
            )

        if not descripcion_rrll and numero:
            descripcion_rrll = validacion_mas_reciente_por_documento.get(
                numero
            )

        row["motivo_oficial_validado_rrll"] = descripcion_rrll or ""

    return resultados_mutables


@router.get("/exportar-retiros")
def exportar_excel_retiros(
    fecha_inicio: Annotated[
        str,
        Query(description="Fecha inicio en formato YYYY-MM-DD"),
    ],
    fecha_fin: Annotated[
        str,
        Query(description="Fecha fin en formato YYYY-MM-DD"),
    ],
    db: Annotated[Session, Depends(get_db)],
):
    try:
        fecha_inicio_dt = date.fromisoformat(fecha_inicio)
        fecha_fin_dt = date.fromisoformat(fecha_fin)
    except ValueError as error:
        raise HTTPException(
            status_code=400,
            detail="Las fechas deben estar en formato YYYY-MM-DD.",
        ) from error

    if fecha_inicio_dt > fecha_fin_dt:
        raise HTTPException(
            status_code=400,
            detail="La fecha de inicio no puede ser mayor que la fecha final."
        )

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Retiros RRLL"

        # =========================
        # ESTILOS BASE
        # =========================
        fill_header = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3"
        )

        fill_title = PatternFill(
            fill_type="solid",
            start_color="B6D7A8",
            end_color="B6D7A8"
        )

        fill_metric = PatternFill(
            fill_type="solid",
            start_color="EAF4E2",
            end_color="EAF4E2"
        )

        thin_side = Side(style="thin", color="B7B7B7")
        border_tabla = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )

        def style_header(cell):
            cell.font = Font(bold=True)
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_tabla

        def style_metric_label(cell):
            cell.font = Font(bold=True)
            cell.fill = fill_metric
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border_tabla

        def style_metric_value(cell):
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_tabla

        # =========================
        # TITULO Y FILTROS HOJA 1
        # =========================
        ws.merge_cells("A1:N1")
        ws["A1"] = "REPORTE RRLL - RETIROS"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A1"].fill = fill_title
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws["A2"] = "Fecha inicio"
        ws["B2"] = fecha_inicio
        ws["A3"] = "Fecha fin"
        ws["B3"] = fecha_fin

        ws["A2"].font = Font(bold=True)
        ws["A3"].font = Font(bold=True)

        # =========================
        # ENCABEZADOS HOJA 1
        # =========================
        headers = [
            "FECHA LEGALIZADOR",
            "NUMERO DE IDENTIFICACION",
            "NOMBRE",
            "APELLIDO",
            "CARGO",
            "SEDE",
            "FECHA DE INGRESO",
            "FECHA DE RETIRO",
            "TOTAL TIEMPO DE TRABAJO",
            "RETIRO LEGALIZADO",
            "DESCRIPCIÓN MOTIVO ESPECIFICO DEL RETIRO",
            "MOTIVO OFICIAL VALIDADO POR RRLL",
            "TIPIFICACION DE RETIRO",
            "OBSERVACION ¿QUÉ DEBE MEJORAR LA COMPAÑÍA?"
        ]

        header_row = 5

        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            style_header(cell)

        ws.row_dimensions[header_row].height = 35

        # =========================
        # DATOS DESDE SQL
        # =========================
        q = text("""
            SELECT *
            FROM public.fn_reporte_retiros_excel(:fecha_inicio, :fecha_fin)
        """)

        resultados = db.execute(
            q,
            {
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin
            }
        ).mappings().all()

        # Completa solo las fechas de ingreso faltantes, especialmente para
        # trabajadores históricos provenientes de la migración.
        resultados = _completar_fechas_ingreso_migrados(db, resultados)

        # Si el total de tiempo viene vacío, lo calcula usando las fechas ya
        # consolidadas. Los valores existentes se conservan sin cambios.
        resultados = _completar_total_tiempo_trabajo(resultados)

        # Conserva la descripción original registrada por el trabajador.
        # Agrega en una columna independiente la validación oficial realizada por RRLL.
        resultados = _agregar_motivo_oficial_validado_rrll(db, resultados)

        # =========================
        # TIPIFICACIONES PARA LISTA
        # =========================
        q_tipificaciones = text("""
            SELECT "Nombre"
            FROM public."TipificacionRetiro"
            ORDER BY "IdTipificacionRetiro"
        """)

        tipificaciones = [
            row["Nombre"]
            for row in db.execute(q_tipificaciones).mappings().all()
        ]

        # =========================
        # ARMAR DATOS HOJA 1
        # =========================
        datos = [
            [
                row.get("fecha_legalizador"),
                row.get("numero_identificacion"),
                row.get("nombre"),
                row.get("apellido"),
                row.get("cargo"),
                row.get("sede"),
                row.get("fecha_ingreso"),
                row.get("fecha_retiro"),
                row.get("total_tiempo_de_trabajo"),
                (
                    "PRESENCIAL"
                    if str(row.get("retiro_legalizado") or "").strip().upper() == "SI"
                    else "VIRTUAL"
                    if str(row.get("retiro_legalizado") or "").strip().upper() == "NO"
                    else ""
                ),
                row.get("descripcion_motivo_especifico_del_retiro"),
                row.get("motivo_oficial_validado_rrll"),
                row.get("tipificacion_de_retiro"),
                row.get("observacion_que_debe_mejorar_la_compania"),
            ]
            for row in resultados
        ]

        fila_datos_inicio = header_row + 1

        for fila in datos:
            ws.append(fila)

        fila_datos_fin = header_row + len(datos)

        # =========================
        # FILTROS Y CONGELAR PANELES
        # =========================
        ws.auto_filter.ref = f"A{header_row}:N{max(fila_datos_fin, header_row)}"
        ws.freeze_panes = "A6"

        # =========================
        # BORDES TABLA
        # =========================
        for row in ws.iter_rows(
            min_row=header_row,
            max_row=max(fila_datos_fin, header_row),
            min_col=1,
            max_col=len(headers)
        ):
            for cell in row:
                cell.border = border_tabla

        # =========================
        # ALTURAS DE FILA
        # =========================
        for row_num in range(fila_datos_inicio, fila_datos_fin + 1):
            ws.row_dimensions[row_num].height = 30

            texto_k = ws[f"K{row_num}"].value or ""
            texto_l = ws[f"L{row_num}"].value or ""
            texto_n = ws[f"N{row_num}"].value or ""

            if (
                len(str(texto_k)) > 40
                or len(str(texto_l)) > 40
                or len(str(texto_n)) > 40
            ):
                ws.row_dimensions[row_num].height = 45

        # =========================
        # ALINEACIONES
        # =========================
        columnas_centradas = ["A", "B", "G", "H", "I", "J", "M"]
        for row_num in range(fila_datos_inicio, fila_datos_fin + 1):
            for col in columnas_centradas:
                ws[f"{col}{row_num}"].alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        columnas_texto_largo = ["K", "L", "N"]
        for row_num in range(fila_datos_inicio, fila_datos_fin + 1):
            for col in columnas_texto_largo:
                ws[f"{col}{row_num}"].alignment = Alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True
                )

        # =========================
        # HOJA OCULTA PARA LISTAS
        # =========================
        ws_listas = wb.create_sheet(title="Listas")
        for idx, tip in enumerate(tipificaciones, start=1):
            ws_listas.cell(row=idx, column=1, value=tip)
        ws_listas.sheet_state = "hidden"

        # =========================
        # LISTA DESPLEGABLE TIPIFICACION
        # =========================
        if tipificaciones:
            dv_tipificacion = DataValidation(
                type="list",
                formula1=f"=Listas!$A$1:$A${len(tipificaciones)}",
                allow_blank=True
            )
            dv_tipificacion.prompt = "Seleccione una tipificación de retiro"
            dv_tipificacion.promptTitle = "Tipificación de Retiro"
            dv_tipificacion.error = "Seleccione una tipificación válida de la lista"
            dv_tipificacion.errorTitle = "Valor no válido"

            ws.add_data_validation(dv_tipificacion)
            dv_tipificacion.add(f"M6:M{max(fila_datos_fin, 1000)}")

        # =========================
        # ANCHOS HOJA 1
        # =========================
        anchos = {
            "A": 20,
            "B": 24,
            "C": 22,
            "D": 22,
            "E": 28,
            "F": 32,
            "G": 20,
            "H": 20,
            "I": 22,
            "J": 18,
            "K": 40,
            "L": 55,
            "M": 30,
            "N": 45,
        }

        for col, ancho in anchos.items():
            ws.column_dimensions[col].width = ancho

        # =========================
        # HOJA 2 - DASHBOARD
        # =========================
        ws_dashboard = wb.create_sheet(title="Dashboard")

        # -------------------------
        # TÍTULO PRINCIPAL
        # -------------------------
        ws_dashboard.merge_cells("A1:J1")
        ws_dashboard["A1"] = "DASHBOARD EJECUTIVO - RETIROS RRLL"
        ws_dashboard["A1"].font = Font(bold=True, size=15)
        ws_dashboard["A1"].fill = fill_title
        ws_dashboard["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws_dashboard.merge_cells("A2:J2")
        ws_dashboard["A2"] = f"Periodo analizado: {fecha_inicio} a {fecha_fin}"
        ws_dashboard["A2"].font = Font(italic=True, size=10)
        ws_dashboard["A2"].alignment = Alignment(horizontal="center", vertical="center")

        # -------------------------
        # NORMALIZADORES
        # -------------------------
        def normalizar_texto(valor, default):
            texto = str(valor or "").strip()
            return texto if texto else default

        def truncar_texto(texto, max_len=28):
            texto = normalizar_texto(texto, "")
            return texto if len(texto) <= max_len else texto[:max_len - 3] + "..."

        # -------------------------
        # CONTADORES
        # -------------------------
        legalizados_counter = Counter()
        tipificacion_counter = Counter()
        motivo_counter = Counter()

        for row in resultados:
            retiro_legalizado_raw = normalizar_texto(
                row.get("retiro_legalizado"),
                ""
            ).upper()

            if retiro_legalizado_raw == "SI":
                retiro_legalizado = "PRESENCIAL"
            elif retiro_legalizado_raw == "NO":
                retiro_legalizado = "VIRTUAL"
            else:
                retiro_legalizado = ""

            tipificacion = normalizar_texto(
                row.get("tipificacion_de_retiro"),
                "SIN TIPIFICACION"
            )

            motivo = (
                row.get("motivo_de_retiro")
                or row.get("motivo_retiro")
                or row.get("nombre_motivo_retiro")
                or row.get("descripcion_motivo_especifico_del_retiro")
            )
            motivo = normalizar_texto(motivo, "SIN MOTIVO REGISTRADO")

            if retiro_legalizado:
                legalizados_counter[retiro_legalizado] += 1

            tipificacion_counter[tipificacion] += 1
            motivo_counter[motivo] += 1

        total_retiros = len(resultados)
        total_legalizados_presencial = legalizados_counter.get("PRESENCIAL", 0)
        total_legalizados_virtual = legalizados_counter.get("VIRTUAL", 0)

        tipificacion_top = (
            max(tipificacion_counter.items(), key=lambda x: x[1])[0]
            if tipificacion_counter else "SIN DATOS"
        )
        motivo_top = (
            max(motivo_counter.items(), key=lambda x: x[1])[0]
            if motivo_counter else "SIN DATOS"
        )

        tipificaciones_ordenadas = sorted(
            tipificacion_counter.items(),
            key=lambda x: x[1],
            reverse=True
        )

        motivos_ordenados = sorted(
            motivo_counter.items(),
            key=lambda x: x[1],
            reverse=True
        )

        top_tipificaciones = tipificaciones_ordenadas[:5]
        top_motivos = motivos_ordenados[:5]

        # -------------------------
        # BLOQUE DE MÉTRICAS
        # -------------------------
        ws_dashboard["A4"] = "INDICADOR"
        ws_dashboard["B4"] = "VALOR"
        style_header(ws_dashboard["A4"])
        style_header(ws_dashboard["B4"])

        metricas = [
            ("Total de retiros analizados", total_retiros),
            ("Retiros presenciales", total_legalizados_presencial),
            ("Retiros virtuales", total_legalizados_virtual),
            ("Tipificación más frecuente", tipificacion_top),
            ("Motivo de retiro más frecuente", motivo_top),
        ]

        fila_metrica = 5
        for etiqueta, valor in metricas:
            ws_dashboard[f"A{fila_metrica}"] = etiqueta
            ws_dashboard[f"B{fila_metrica}"] = valor
            style_metric_label(ws_dashboard[f"A{fila_metrica}"])
            style_metric_value(ws_dashboard[f"B{fila_metrica}"])
            fila_metrica += 1

        # ==================================================
        # BLOQUE 1 - RETIRO LEGALIZADO (TABLA + GRÁFICA)
        # ==================================================
        ws_dashboard.merge_cells("A12:B12")
        ws_dashboard["A12"] = "RESUMEN DE RETIRO LEGALIZADO"
        ws_dashboard["A12"].font = Font(bold=True, size=11)
        ws_dashboard["A12"].fill = fill_title
        ws_dashboard["A12"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dashboard["A12"].border = border_tabla
        ws_dashboard["B12"].border = border_tabla

        ws_dashboard["A13"] = "ESTADO"
        ws_dashboard["B13"] = "CANTIDAD"
        style_header(ws_dashboard["A13"])
        style_header(ws_dashboard["B13"])

        estados_legalizados = ["PRESENCIAL", "VIRTUAL"]
        fila_legalizados_inicio = 14

        for i, estado in enumerate(estados_legalizados, start=fila_legalizados_inicio):
            ws_dashboard[f"A{i}"] = estado
            ws_dashboard[f"B{i}"] = legalizados_counter.get(estado, 0)
            ws_dashboard[f"A{i}"].border = border_tabla
            ws_dashboard[f"B{i}"].border = border_tabla
            ws_dashboard[f"A{i}"].alignment = Alignment(horizontal="left", vertical="center")
            ws_dashboard[f"B{i}"].alignment = Alignment(horizontal="center", vertical="center")

        pie_legalizados = PieChart()
        pie_legalizados.title = "Distribución de retiros legalizados"
        pie_legalizados.height = 6.2
        pie_legalizados.width = 7.8
        pie_legalizados.varyColors = True
        pie_legalizados.legend = None

        labels_legalizados = Reference(ws_dashboard, min_col=1, min_row=14, max_row=15)
        data_legalizados = Reference(ws_dashboard, min_col=2, min_row=14, max_row=15)

        pie_legalizados.add_data(data_legalizados, titles_from_data=False)
        pie_legalizados.set_categories(labels_legalizados)

        pie_legalizados.dLbls = DataLabelList()
        pie_legalizados.dLbls.showCatName = True
        pie_legalizados.dLbls.showPercent = True
        pie_legalizados.dLbls.showVal = False
        pie_legalizados.dLbls.showLegendKey = False
        pie_legalizados.dLbls.showSerName = False

        ws_dashboard.add_chart(pie_legalizados, "E12")

        # ==================================================
        # BLOQUE 2 - TIPIFICACIONES (SOLO TABLA)
        # ==================================================
        ws_dashboard.merge_cells("A28:B28")
        ws_dashboard["A28"] = "RESUMEN DE TIPIFICACIONES"
        ws_dashboard["A28"].font = Font(bold=True, size=11)
        ws_dashboard["A28"].fill = fill_title
        ws_dashboard["A28"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dashboard["A28"].border = border_tabla
        ws_dashboard["B28"].border = border_tabla

        ws_dashboard["A29"] = "TIPIFICACIÓN"
        ws_dashboard["B29"] = "CANTIDAD"
        style_header(ws_dashboard["A29"])
        style_header(ws_dashboard["B29"])

        fila_tip_inicio = 30
        for idx, (tip, cantidad) in enumerate(top_tipificaciones, start=fila_tip_inicio):
            ws_dashboard[f"A{idx}"] = tip
            ws_dashboard[f"B{idx}"] = cantidad

            ws_dashboard[f"A{idx}"].border = border_tabla
            ws_dashboard[f"B{idx}"].border = border_tabla

            ws_dashboard[f"A{idx}"].alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )
            ws_dashboard[f"B{idx}"].alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            ws_dashboard[f"A{idx}"].font = Font(size=11)
            ws_dashboard[f"B{idx}"].font = Font(size=11, bold=True)

            ws_dashboard.row_dimensions[idx].height = 34

        # ==================================================
        # BLOQUE 3 - MOTIVOS DE RETIRO (TABLA + GRÁFICA)
        # ==================================================
        ws_dashboard.merge_cells("A44:B44")
        ws_dashboard["A44"] = "RESUMEN DE MOTIVOS DE RETIRO"
        ws_dashboard["A44"].font = Font(bold=True, size=11)
        ws_dashboard["A44"].fill = fill_title
        ws_dashboard["A44"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dashboard["A44"].border = border_tabla
        ws_dashboard["B44"].border = border_tabla

        ws_dashboard["A45"] = "MOTIVO"
        ws_dashboard["B45"] = "CANTIDAD"
        style_header(ws_dashboard["A45"])
        style_header(ws_dashboard["B45"])

        fila_motivo_inicio = 46
        for idx, (motivo, cantidad) in enumerate(top_motivos, start=fila_motivo_inicio):
            ws_dashboard[f"A{idx}"] = motivo
            ws_dashboard[f"B{idx}"] = cantidad

            ws_dashboard[f"A{idx}"].border = border_tabla
            ws_dashboard[f"B{idx}"].border = border_tabla

            ws_dashboard[f"A{idx}"].alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True
            )
            ws_dashboard[f"B{idx}"].alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

            ws_dashboard[f"A{idx}"].font = Font(size=11)
            ws_dashboard[f"B{idx}"].font = Font(size=11, bold=True)

            ws_dashboard.row_dimensions[idx].height = 34

        fila_motivo_fin = max(fila_motivo_inicio, fila_motivo_inicio + len(top_motivos) - 1)

        ws_dashboard["C45"] = "ETIQUETA_GRAFICA_MOT"
        for idx, (motivo, _) in enumerate(top_motivos, start=fila_motivo_inicio):
            ws_dashboard[f"C{idx}"] = truncar_texto(motivo, 24)

        bar_motivos = BarChart()
        bar_motivos.type = "bar"
        bar_motivos.style = 11
        bar_motivos.title = "Motivos de retiro"
        bar_motivos.height = 7.4
        bar_motivos.width = 8.8
        bar_motivos.legend = None
        bar_motivos.gapWidth = 110
        bar_motivos.overlap = 0

        bar_motivos.x_axis.title = None
        bar_motivos.y_axis.title = None

        data_mot = Reference(ws_dashboard, min_col=2, min_row=46, max_row=fila_motivo_fin)
        categories_mot = Reference(ws_dashboard, min_col=1, min_row=46, max_row=fila_motivo_fin)

        bar_motivos.add_data(data_mot, titles_from_data=False)
        bar_motivos.set_categories(categories_mot)

        bar_motivos.dLbls = DataLabelList()
        bar_motivos.dLbls.showVal = True
        bar_motivos.dLbls.showCatName = False
        bar_motivos.dLbls.showSerName = False
        bar_motivos.dLbls.showLegendKey = False
        bar_motivos.dLbls.position = "outEnd"

        bar_motivos.x_axis.delete = True
        bar_motivos.y_axis.delete = True
        bar_motivos.x_axis.majorGridlines = None

        ws_dashboard.add_chart(bar_motivos, "D45")

        # -------------------------
        # AJUSTES VISUALES DASHBOARD
        # -------------------------
        ws_dashboard.column_dimensions["A"].width = 52
        ws_dashboard.column_dimensions["B"].width = 16
        ws_dashboard.column_dimensions["C"].width = 18
        ws_dashboard.column_dimensions["D"].width = 4
        ws_dashboard.column_dimensions["E"].width = 4
        ws_dashboard.column_dimensions["F"].width = 4
        ws_dashboard.column_dimensions["G"].width = 4
        ws_dashboard.column_dimensions["H"].width = 4
        ws_dashboard.column_dimensions["I"].width = 4
        ws_dashboard.column_dimensions["J"].width = 4

        for fila in range(1, 65):
            ws_dashboard.row_dimensions[fila].height = 22

        for fila in [1, 2, 12, 28, 44]:
            ws_dashboard.row_dimensions[fila].height = 26

        # =========================
        # GENERAR ARCHIVO
        # =========================
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        nombre_archivo = f"reporte_retiros_rrll_{fecha_inicio}_a_{fecha_fin}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{nombre_archivo}"'
            }
        )

    except (
        SQLAlchemyError,
        AttributeError,
        KeyError,
        OSError,
        TypeError,
        ValueError,
    ) as error:
        raise HTTPException(
            status_code=500,
            detail=f"Error al generar el Excel de retiros: {error!s}",
        ) from error
# ============================================================
# EXCEL DE PROCESOS DISCIPLINARIOS
# Una fila = un expediente disciplinario
# ============================================================

def _valor_fila(fila, candidatos, default=None):
    """
    Busca un valor en un mapping sin depender de mayúsculas/minúsculas.
    Permite que el reporte siga funcionando aunque algunos nombres de columna
    cambien entre ambientes.
    """
    if not fila:
        return default

    mapa = {
        str(clave).lower(): valor
        for clave, valor in dict(fila).items()
    }

    for candidato in candidatos:
        clave = str(candidato).lower()
        if clave in mapa:
            valor = mapa[clave]
            if valor not in (None, ""):
                return valor

    return default


def _primer_valor(filas, candidatos, default=None):
    for fila in filas:
        valor = _valor_fila(fila, candidatos, None)
        if valor not in (None, ""):
            return valor
    return default


def _formatear_novedad_disciplinaria(valor) -> str:
    """
    Convierte el código interno del motivo de citación a un texto legible
    para el Excel de procesos disciplinarios.
    """
    codigo = str(valor or "").strip().upper()

    if not codigo:
        return ""

    etiquetas = {
        "AUSENCIA_INJUSTIFICADA": "Ausencia injustificada",
        "RETARDOS_INJUSTIFICADOS": "Retardos injustificados",
        "INCUMPLIMIENTO_FUNCIONES": "Incumplimiento de funciones",
        "INCUMPLIMIENTO_NORMAS": "Incumplimiento de normas",
        "CLIMA_LABORAL": "Clima laboral",
        "DANOS_BIEN_AJENO_AFECTACION_CLIENTE": (
            "Daños en bien ajeno - afectación al cliente"
        ),
        "PERIODO_PRUEBA": "Período de prueba",
        "ATENCION_LINEA_VERDE": "Atención línea verde",
    }

    return etiquetas.get(
        codigo,
        str(valor or "").replace("_", " ").strip().capitalize(),
    )


def _dias_entre_fechas(fecha_inicio, fecha_fin):
    """
    Calcula días calendario inclusivos cuando existen ambas fechas.
    Si no hay fechas suficientes, retorna vacío.
    """
    inicio = _convertir_a_fecha(fecha_inicio)
    fin = _convertir_a_fecha(fecha_fin)

    if not inicio or not fin or fin < inicio:
        return ""

    return (fin - inicio).days + 1


def _obtener_ultimos_registros_por_proceso(
    db: Session,
    nombre_tabla: str,
    ids_proceso: list[int],
    candidatos_fk: list[str],
    candidatos_orden: list[str],
):
    """
    Obtiene el registro más reciente de una tabla por proceso disciplinario.
    Usa las columnas reales de information_schema para evitar acoplar el
    reporte a una sola variante de nombres.
    """
    if not ids_proceso:
        return {}

    columnas = _obtener_columnas_tabla(db, nombre_tabla)
    if not columnas:
        return {}

    columna_fk = _buscar_columna(columnas, candidatos_fk)
    if not columna_fk:
        return {}

    columna_orden = _buscar_columna(columnas, candidatos_orden)
    if not columna_orden:
        columna_orden = columna_fk

    consulta = text(f"""
        SELECT *
        FROM public.{_quote_identifier(nombre_tabla)}
        WHERE {_quote_identifier(columna_fk)} = ANY(:ids_proceso)
        ORDER BY
            {_quote_identifier(columna_fk)},
            {_quote_identifier(columna_orden)} DESC NULLS LAST
    """)

    filas = db.execute(
        consulta,
        {"ids_proceso": ids_proceso},
    ).mappings().all()

    resultado = {}

    for fila in filas:
        id_proceso = _valor_fila(
            fila,
            [columna_fk],
        )

        if id_proceso is None:
            continue

        id_proceso = int(id_proceso)

        if id_proceso not in resultado:
            resultado[id_proceso] = dict(fila)

    return resultado


@router.get("/exportar-procesos-disciplinarios")
def exportar_excel_procesos_disciplinarios(
    fecha_inicio: Annotated[
        str,
        Query(description="Fecha inicio en formato YYYY-MM-DD"),
    ],
    fecha_fin: Annotated[
        str,
        Query(description="Fecha fin en formato YYYY-MM-DD"),
    ],
    db: Annotated[Session, Depends(get_db)],
):
    """
    Genera el Excel de Procesos Disciplinarios para RRLL.

    Regla principal:
        1 fila = 1 expediente disciplinario.

    El filtro se realiza por la fecha de creación del expediente.

    Los campos que todavía no existan en el flujo se entregan vacíos;
    el reporte no inventa información ni modifica la base de datos.
    """
    try:
        fecha_inicio_dt = date.fromisoformat(fecha_inicio)
        fecha_fin_dt = date.fromisoformat(fecha_fin)
    except ValueError as error:
        raise HTTPException(
            status_code=400,
            detail="Las fechas deben estar en formato YYYY-MM-DD.",
        ) from error

    if fecha_inicio_dt > fecha_fin_dt:
        raise HTTPException(
            status_code=400,
            detail=(
                "La fecha de inicio no puede ser mayor "
                "que la fecha final."
            ),
        )

    try:
        # =====================================================
        # 1. UNIVERSO DE EXPEDIENTES
        # =====================================================

        consulta_procesos = text("""
            SELECT
                p."IdProcesoDisciplinario" AS id_proceso_disciplinario,
                p."IdRegistroPersonal" AS id_registro_personal,
                p."EstadoProceso" AS estado_proceso,
                p."FechaCreacion" AS fecha_creacion_proceso,
                rp."NumeroIdentificacion" AS cedula,
                rp."Nombres" AS nombres,
                rp."Apellidos" AS apellidos
            FROM public."ProcesoDisciplinario" p
            INNER JOIN public."RegistroPersonal" rp
                ON rp."IdRegistroPersonal" = p."IdRegistroPersonal"
            WHERE p."FechaCreacion"::date
                  BETWEEN CAST(:fecha_inicio AS date)
                      AND CAST(:fecha_fin AS date)
            ORDER BY
                p."FechaCreacion" ASC,
                p."IdProcesoDisciplinario" ASC
        """)

        procesos = db.execute(
            consulta_procesos,
            {
                "fecha_inicio": fecha_inicio,
                "fecha_fin": fecha_fin,
            },
        ).mappings().all()

        ids_proceso = [
            int(row["id_proceso_disciplinario"])
            for row in procesos
        ]

        # =====================================================
        # 2. ÚLTIMA INFORMACIÓN DE CADA ETAPA
        # =====================================================

        citaciones = _obtener_ultimos_registros_por_proceso(
            db=db,
            nombre_tabla="CitacionProcesoDisciplinario",
            ids_proceso=ids_proceso,
            candidatos_fk=[
                "IdProcesoDisciplinario",
                "IdProceso",
            ],
            candidatos_orden=[
                "FechaActualizacion",
                "FechaCreacion",
                "IdCitacionProcesoDisciplinario",
            ],
        )

        descargos = _obtener_ultimos_registros_por_proceso(
            db=db,
            nombre_tabla="DescargoProcesoDisciplinario",
            ids_proceso=ids_proceso,
            candidatos_fk=[
                "IdProcesoDisciplinario",
                "IdProceso",
            ],
            candidatos_orden=[
                "FechaActualizacion",
                "FechaCreacion",
                "IdDescargoProcesoDisciplinario",
                "IdDescargo",
            ],
        )

        cierres = _obtener_ultimos_registros_por_proceso(
            db=db,
            nombre_tabla="CierreProcesoDisciplinario",
            ids_proceso=ids_proceso,
            candidatos_fk=[
                "IdProcesoDisciplinario",
                "IdProceso",
            ],
            candidatos_orden=[
                "FechaActualizacion",
                "FechaCreacion",
                "IdCierreProcesoDisciplinario",
                "IdCierre",
            ],
        )

        agendas = _obtener_ultimos_registros_por_proceso(
            db=db,
            nombre_tabla="AgendaProcesoDisciplinario",
            ids_proceso=ids_proceso,
            candidatos_fk=[
                "IdProcesoDisciplinario",
                "IdProceso",
            ],
            candidatos_orden=[
                "FechaActualizacion",
                "FechaCreacion",
                "IdAgendaProcesoDisciplinario",
                "IdAgenda",
            ],
        )

        # =====================================================
        # 3. ARMAR FILAS DEL REPORTE
        # =====================================================

        filas_excel = []

        for proceso in procesos:
            id_proceso = int(
                proceso["id_proceso_disciplinario"]
            )

            citacion = citaciones.get(id_proceso, {})
            descargo = descargos.get(id_proceso, {})
            cierre = cierres.get(id_proceso, {})
            agenda = agendas.get(id_proceso, {})

            fuentes = [
                cierre,
                descargo,
                citacion,
                agenda,
                proceso,
            ]

            cedula = str(
                proceso.get("cedula") or ""
            ).strip()

            colaborador = " ".join(
                [
                    str(proceso.get("nombres") or "").strip(),
                    str(proceso.get("apellidos") or "").strip(),
                ]
            ).strip()

            novedad = _formatear_novedad_disciplinaria(
                _primer_valor(
                    [citacion, descargo],
                    [
                        "MotivoCitacion",
                        "Motivo",
                        "TipoFalta",
                        "Novedad",
                    ],
                    "",
                )
            )

            inicio_ausencia = _primer_valor(
                fuentes,
                [
                    "InicioAusencia",
                    "FechaInicioAusencia",
                    "FechaDesdeAusencia",
                ],
                "",
            )

            fin_ausencia = _primer_valor(
                fuentes,
                [
                    "FinAusencia",
                    "FechaFinAusencia",
                    "FechaHastaAusencia",
                ],
                "",
            )

            dias_ausencia = _primer_valor(
                fuentes,
                [
                    "DiasAusencia",
                    "DiasDeAusencia",
                ],
                None,
            )

            if dias_ausencia is None:
                dias_ausencia = _dias_entre_fechas(
                    inicio_ausencia,
                    fin_ausencia,
                )

            medida = _primer_valor(
                [cierre],
                [
                    "MedidaDisciplinaria",
                    "Medida",
                    "Decision",
                    "TipoMedida",
                ],
                "",
            )

            inicio_suspension = _primer_valor(
                [cierre, descargo],
                [
                    "DiaInicioSuspension",
                    "FechaInicioSuspension",
                    "InicioSuspension",
                ],
                "",
            )

            fin_suspension = _primer_valor(
                [cierre, descargo],
                [
                    "DiaFinSuspension",
                    "FechaFinSuspension",
                    "FinSuspension",
                ],
                "",
            )

            dias_suspension = _primer_valor(
                [cierre, descargo],
                [
                    "DiasSuspension",
                    "DiasDeSuspension",
                ],
                None,
            )

            if dias_suspension is None:
                dias_suspension = _dias_entre_fechas(
                    inicio_suspension,
                    fin_suspension,
                )

            ultimo_dia_laborado = _primer_valor(
                fuentes,
                [
                    "UltimoDiaLaborado",
                    "FechaUltimoDiaLaborado",
                ],
                "",
            )

            estado_agenda = _primer_valor(
                [agenda],
                [
                    "EstadoAgenda",
                    "Estado",
                ],
                "",
            )

            presentacion = _primer_valor(
                [agenda, descargo],
                [
                    "Presentacion",
                    "SePresento",
                    "Asistencia",
                    "EstadoAsistencia",
                ],
                "",
            )

            if not presentacion:
                estado_agenda_normalizado = str(
                    estado_agenda or ""
                ).strip().upper()

                if estado_agenda_normalizado == "ATENDIDO":
                    presentacion = "SÍ"
                elif estado_agenda_normalizado == "CANCELADO":
                    presentacion = "NO"

            sede = _primer_valor(
                [citacion],
                [
                    "Cliente",
                    "Sede",
                ],
                "",
            )

            fecha_citacion = _primer_valor(
                [citacion],
                [
                    "FechaCitacion",
                    "Fecha",
                ],
                "",
            )

            fecha_entrega_carta = _primer_valor(
                [cierre, descargo],
                [
                    "FechaEntregaCarta",
                    "FechaCarta",
                    "FechaNotificacionCarta",
                ],
                "",
            )

            fecha_agenda = _primer_valor(
                [agenda],
                [
                    "FechaEvento",
                    "FechaAgenda",
                    "Fecha",
                ],
                "",
            )

            informacion = _primer_valor(
                [citacion, descargo],
                [
                    "RelatoHechos",
                    "Informacion",
                    "ManifestacionSupervisor",
                ],
                "",
            )

            observacion = _primer_valor(
                [citacion, descargo, cierre, agenda],
                [
                    "ObservacionOperaciones",
                    "Observacion",
                    "Observaciones",
                    "Conclusion",
                    "ConclusionCierre",
                ],
                "",
            )

            modalidad = _primer_valor(
                [citacion, agenda],
                [
                    "Modalidad",
                ],
                "",
            )

            verificacion = _primer_valor(
                [citacion, cierre],
                [
                    "DesempenoContinua",
                    "Verificacion",
                    "NivelDesempeno",
                ],
                "",
            )

            filas_excel.append(
                [
                    cedula,
                    colaborador,
                    novedad,
                    dias_ausencia,
                    inicio_ausencia,
                    fin_ausencia,
                    medida,
                    dias_suspension,
                    inicio_suspension,
                    fin_suspension,
                    ultimo_dia_laborado,
                    presentacion,
                    sede,
                    fecha_citacion,
                    fecha_entrega_carta,
                    fecha_agenda,
                    informacion,
                    observacion,
                    estado_agenda,
                    modalidad,
                    verificacion,
                ]
            )

        # =====================================================
        # 4. CREAR EXCEL
        # =====================================================

        wb = Workbook()
        ws = wb.active
        ws.title = "Procesos Disciplinarios"

        fill_header = PatternFill(
            fill_type="solid",
            start_color="D9EAD3",
            end_color="D9EAD3",
        )

        fill_title = PatternFill(
            fill_type="solid",
            start_color="B6D7A8",
            end_color="B6D7A8",
        )

        fill_metric = PatternFill(
            fill_type="solid",
            start_color="EAF4E2",
            end_color="EAF4E2",
        )

        thin_side = Side(
            style="thin",
            color="B7B7B7",
        )

        border_tabla = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        def style_header_disciplinario(cell):
            cell.font = Font(
                bold=True,
                color="1F1F1F",
            )
            cell.fill = fill_header
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = border_tabla

        # =====================================================
        # 5. TÍTULO Y PERIODO
        # =====================================================

        ws.merge_cells("A1:U1")
        ws["A1"] = (
            "REPORTE RRLL - PROCESOS DISCIPLINARIOS"
        )
        ws["A1"].font = Font(
            bold=True,
            size=14,
        )
        ws["A1"].fill = fill_title
        ws["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        ws["A2"] = "Fecha inicio"
        ws["B2"] = fecha_inicio
        ws["A3"] = "Fecha fin"
        ws["B3"] = fecha_fin

        ws["A2"].font = Font(bold=True)
        ws["A3"].font = Font(bold=True)

        # =====================================================
        # 6. ENCABEZADOS SOLICITADOS
        # =====================================================

        headers = [
            "CÉDULA",
            "COLABORADOR",
            "NOVEDAD",
            "DÍAS DE AUSENCIA",
            "INICIO AUSENCIA",
            "FIN AUSENCIA",
            "MEDIDA",
            "DÍAS DE SUSPENSIÓN",
            "DÍA INICIO SUSPENSIÓN",
            "DÍA FIN SUSPENSIÓN",
            "ÚLTIMO DÍA LABORADO",
            "PRESENTACIÓN",
            "SEDE",
            "FECHA CITACIÓN",
            "FECHA ENTREGA CARTA",
            "FECHA AGENDA",
            "INFORMACIÓN",
            "OBSERVACIÓN",
            "AGENDADOS",
            "MODALIDAD",
            "VERIFICACIÓN",
        ]

        header_row = 5

        for col_num, header in enumerate(
            headers,
            start=1,
        ):
            cell = ws.cell(
                row=header_row,
                column=col_num,
                value=header,
            )
            style_header_disciplinario(cell)

        ws.row_dimensions[header_row].height = 42

        # =====================================================
        # 7. DATOS
        # =====================================================

        for fila in filas_excel:
            ws.append(fila)

        fila_datos_inicio = header_row + 1
        fila_datos_fin = (
            header_row + len(filas_excel)
        )

        for row_num in range(
            fila_datos_inicio,
            fila_datos_fin + 1,
        ):
            for col_num in range(
                1,
                len(headers) + 1,
            ):
                cell = ws.cell(
                    row=row_num,
                    column=col_num,
                )
                cell.border = border_tabla
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

            # Fechas con formato colombiano
            for col_num in [
                5,   # Inicio ausencia
                6,   # Fin ausencia
                9,   # Inicio suspensión
                10,  # Fin suspensión
                11,  # Último día laborado
                14,  # Fecha citación
                15,  # Fecha entrega carta
                16,  # Fecha agenda
            ]:
                cell = ws.cell(
                    row=row_num,
                    column=col_num,
                )
                fecha = _convertir_a_fecha(
                    cell.value
                )
                if fecha:
                    cell.value = fecha
                    cell.number_format = "DD/MM/YYYY"

            ws.row_dimensions[row_num].height = 34

        # =====================================================
        # 8. FILTROS Y PANELES
        # =====================================================

        ultima_fila_filtro = max(
            fila_datos_fin,
            header_row,
        )

        ws.auto_filter.ref = (
            f"A{header_row}:U{ultima_fila_filtro}"
        )

        ws.freeze_panes = "A6"

        # =====================================================
        # 9. ANCHOS
        # =====================================================

        anchos = {
            "A": 18,
            "B": 34,
            "C": 32,
            "D": 18,
            "E": 18,
            "F": 18,
            "G": 28,
            "H": 20,
            "I": 22,
            "J": 22,
            "K": 22,
            "L": 18,
            "M": 34,
            "N": 19,
            "O": 22,
            "P": 19,
            "Q": 45,
            "R": 45,
            "S": 20,
            "T": 18,
            "U": 22,
        }

        for col, ancho in anchos.items():
            ws.column_dimensions[col].width = ancho

        # =====================================================
        # 10. HOJA RESUMEN
        # =====================================================

        ws_resumen = wb.create_sheet(
            title="Resumen"
        )

        ws_resumen.merge_cells("A1:F1")
        ws_resumen["A1"] = (
            "RESUMEN - PROCESOS DISCIPLINARIOS"
        )
        ws_resumen["A1"].font = Font(
            bold=True,
            size=14,
        )
        ws_resumen["A1"].fill = fill_title
        ws_resumen["A1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        ws_resumen["A2"] = (
            f"Periodo analizado: "
            f"{fecha_inicio} a {fecha_fin}"
        )
        ws_resumen["A2"].font = Font(
            italic=True,
            size=10,
        )

        modalidades_counter = Counter(
            str(fila[19] or "SIN MODALIDAD").strip()
            for fila in filas_excel
        )

        medidas_counter = Counter(
            str(fila[6] or "SIN MEDIDA").strip()
            for fila in filas_excel
        )

        estados_agenda_counter = Counter(
            str(fila[18] or "SIN ESTADO").strip()
            for fila in filas_excel
        )

        metricas = [
            (
                "Total expedientes",
                len(filas_excel),
            ),
            (
                "Presenciales",
                modalidades_counter.get(
                    "PRESENCIAL",
                    0,
                ),
            ),
            (
                "Virtuales",
                modalidades_counter.get(
                    "VIRTUAL",
                    0,
                ),
            ),
            (
                "Atendidos",
                estados_agenda_counter.get(
                    "ATENDIDO",
                    0,
                ),
            ),
            (
                "Con medida registrada",
                sum(
                    cantidad
                    for medida, cantidad
                    in medidas_counter.items()
                    if medida != "SIN MEDIDA"
                ),
            ),
        ]

        ws_resumen["A4"] = "INDICADOR"
        ws_resumen["B4"] = "VALOR"

        for cell in [
            ws_resumen["A4"],
            ws_resumen["B4"],
        ]:
            style_header_disciplinario(cell)

        fila_metrica = 5

        for etiqueta, valor in metricas:
            ws_resumen[
                f"A{fila_metrica}"
            ] = etiqueta
            ws_resumen[
                f"B{fila_metrica}"
            ] = valor

            ws_resumen[
                f"A{fila_metrica}"
            ].fill = fill_metric

            ws_resumen[
                f"A{fila_metrica}"
            ].font = Font(bold=True)

            ws_resumen[
                f"A{fila_metrica}"
            ].border = border_tabla

            ws_resumen[
                f"B{fila_metrica}"
            ].border = border_tabla

            fila_metrica += 1

        ws_resumen.column_dimensions["A"].width = 36
        ws_resumen.column_dimensions["B"].width = 18

        # =====================================================
        # 11. GENERAR ARCHIVO
        # =====================================================

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        nombre_archivo = (
            "reporte_procesos_disciplinarios_"
            f"{fecha_inicio}_a_{fecha_fin}.xlsx"
        )

        return StreamingResponse(
            output,
            media_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            headers={
                "Content-Disposition": (
                    f'attachment; filename="{nombre_archivo}"'
                )
            },
        )

    except HTTPException:
        raise

    except (
        SQLAlchemyError,
        AttributeError,
        KeyError,
        OSError,
        TypeError,
        ValueError,
    ) as error:
        raise HTTPException(
            status_code=500,
            detail=(
                "Error al generar el Excel de procesos "
                f"disciplinarios: {error!s}"
            ),
        ) from error