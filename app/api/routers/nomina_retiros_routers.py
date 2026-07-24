import os
from io import BytesIO
from datetime import datetime, date

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from sqlalchemy import text
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from infrastructure.db.deps import get_db

router = APIRouter(prefix="/api/nomina-retiros", tags=["Nómina Retiros"])


class ObservacionNominaRequest(BaseModel):
    observacion_nomina: str | None = None


class FinalizarRetiroNominaRequest(BaseModel):
    fecha_pago_liquidacion: date


FILTRO_NOMINA_SQL = """
(
    rp."IdEstadoProceso" IN (30, 32, 35)
    OR UPPER(COALESCE(rl."EstadoCasoRRLL", '')) IN ('ABIERTO', 'ENVIADO_NOMINA', 'CERRADO')
)
"""


def _consultar_retiros_nomina(db: Session):
    query = text(f"""
        SELECT
            rl."IdRetiroLaboral",
            rl."IdRegistroPersonal",
            rp."NumeroIdentificacion",
            rp."Nombres",
            rp."Apellidos",
            COALESCE(c."Nombre", 'SIN CLIENTE') AS "NombreCliente",
            rl."FechaProceso",
            rl."FechaRetiro",
            rl."FechaCierre",
            rl."FechaEnvioNomina",
            rl."FechaPagoLiquidacion",
            rl."EstadoCasoRRLL",
            rp."IdEstadoProceso",
            ep."Nombre" AS "EstadoProceso",
            mr."Nombre" AS "MotivoRetiro",
            tr."Nombre" AS "TipificacionRetiro",
            rl."ObservacionGeneral",
            rl."ObservacionRetiro",
            rl."ObservacionNomina",
            rl."UsuarioObservacionNomina",
            rl."FechaObservacionNomina",
            CASE
                WHEN rp."IdEstadoProceso" = 32
                  OR UPPER(COALESCE(rl."EstadoCasoRRLL", '')) = 'ENVIADO_NOMINA'
                THEN true
                ELSE false
            END AS "PuedeGestionarNomina"
        FROM public."RetiroLaboral" rl
        INNER JOIN public."RegistroPersonal" rp
            ON rp."IdRegistroPersonal" = rl."IdRegistroPersonal"
        LEFT JOIN public."Cliente" c
            ON c."IdCliente" = rl."IdCliente"
        LEFT JOIN public."EstadoProceso" ep
            ON ep."IdEstadoProceso" = rp."IdEstadoProceso"
        LEFT JOIN public."MotivoRetiro" mr
            ON mr."IdMotivoRetiro" = rl."IdMotivoRetiro"
        LEFT JOIN public."TipificacionRetiro" tr
            ON tr."IdTipificacionRetiro" = rl."IdTipificacionRetiro"
        WHERE {FILTRO_NOMINA_SQL}
        ORDER BY
            CASE
                WHEN rp."IdEstadoProceso" = 32
                  OR UPPER(COALESCE(rl."EstadoCasoRRLL", '')) = 'ENVIADO_NOMINA'
                THEN 0
                WHEN rp."IdEstadoProceso" = 35 THEN 1
                ELSE 2
            END,
            rl."FechaCreacion" DESC;
    """)

    return [dict(row) for row in db.execute(query).mappings().all()]


def _valor_fecha_excel(fecha):
    if not fecha:
        return ""

    if isinstance(fecha, datetime):
        return fecha.date()

    return fecha


def _nombre_completo(row):
    nombres = row.get("Nombres") or ""
    apellidos = row.get("Apellidos") or ""
    return " ".join(f"{nombres} {apellidos}".split()).upper()


def _configurar_hoja_detalle(ws, titulo, registros):
    verde = "008060"
    verde_claro = "E8F7F0"
    gris_texto = "374151"
    blanco = "FFFFFF"
    gris_fila = "F9FAFB"
    borde_color = "D9E2EC"

    header_fill = PatternFill("solid", fgColor=verde)
    title_fill = PatternFill("solid", fgColor=verde_claro)
    thin = Side(style="thin", color=borde_color)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:F1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=16, color=verde)
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Generado: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}"
    ws["A2"].font = Font(italic=True, size=10, color="6B7280")
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    encabezados = [
        "Identificación",
        "Nombre completo",
        "Fecha de retiro",
        "Fecha pago liquidación",
        "Cliente",
        "Estado",
    ]

    for col, header in enumerate(encabezados, start=1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True, color=blanco)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    ws.row_dimensions[4].height = 24

    for row_idx, registro in enumerate(registros, start=5):
        estado_caso = str(registro.get("EstadoCasoRRLL") or "").upper()
        id_estado = int(registro.get("IdEstadoProceso") or 0)

        if id_estado == 35:
            estado_excel = "Retirado"
        elif id_estado == 32 or estado_caso == "ENVIADO_NOMINA":
            estado_excel = "Cerrado"
        elif estado_caso == "CERRADO":
            estado_excel = "Cerrado"
        else:
           estado_excel = "Abierto"

        valores = [
            registro.get("NumeroIdentificacion") or "",
            _nombre_completo(registro),
            _valor_fecha_excel(registro.get("FechaRetiro")),
            _valor_fecha_excel(registro.get("FechaPagoLiquidacion")),
            " ".join(
                str(registro.get("NombreCliente") or "SIN CLIENTE")
                .replace("\r", " ")
                .replace("\n", " ")
                .split()
            ).upper(),
            estado_excel.upper(),
        ]

        for col_idx, valor in enumerate(valores, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.font = Font(color=gris_texto)

            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=gris_fila)

            if col_idx in (3, 4) and valor:
                cell.number_format = "DD/MM/YYYY"

            if col_idx == 6:
                estado = str(valor).upper()
                if "RETIRADO" in estado:
                    cell.fill = PatternFill("solid", fgColor="DCFCE7")
                    cell.font = Font(color="166534", bold=True)
                elif "CERRADO" in estado or "NÓMINA" in estado or "NOMINA" in estado:
                    cell.fill = PatternFill("solid", fgColor="DBEAFE")
                    cell.font = Font(color="1D4ED8", bold=True)
                else:
                    cell.fill = PatternFill("solid", fgColor="FEF3C7")
                    cell.font = Font(color="92400E", bold=True)

        ws.row_dimensions[row_idx].height = 30

    widths = {
        "A": 20,
        "B": 42,
        "C": 18,
        "D": 24,
        "E": 60,
        "F": 24,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:F{max(4, len(registros) + 4)}"


@router.get("")
def listar_retiros_nomina(db: Session = Depends(get_db)):
    try:
        query = text(f"""
            SELECT
                rl."IdRetiroLaboral",
                rl."IdRegistroPersonal",
                rp."NumeroIdentificacion",
                rp."Nombres",
                rp."Apellidos",
                COALESCE(c."Nombre", 'SIN CLIENTE') AS "NombreCliente",
                rl."FechaProceso",
                rl."FechaRetiro",
                pso."FechaCreacion" AS "FechaPazYSalvo",
                rl."FechaCierre",
                rl."FechaEnvioNomina",
                rl."FechaPagoLiquidacion",
                rl."EstadoCasoRRLL",

                -- Último día laborado hasta creación del Paz y Salvo.
                -- Esta medición se conserva en días calendario.
                CASE
                    WHEN rl."FechaRetiro" IS NOT NULL
                     AND pso."FechaCreacion" IS NOT NULL
                     AND DATE(pso."FechaCreacion") >= rl."FechaRetiro"
                    THEN DATE(pso."FechaCreacion") - rl."FechaRetiro"
                    ELSE NULL
                END AS "DiasRetiroPazYSalvo",

                -- Paz y Salvo hasta cierre de RRLL.
                -- Se entrega en segundos para mostrar minutos, horas o días
                -- sin perder la hora real del proceso.
                CASE
                    WHEN pso."FechaCreacion" IS NOT NULL
                     AND rl."FechaCierre" IS NOT NULL
                     AND rl."FechaCierre" >= pso."FechaCreacion"
                    THEN ROUND(
                        EXTRACT(
                            EPOCH FROM (
                                rl."FechaCierre" - pso."FechaCreacion"
                            )
                        )
                    )::bigint
                    ELSE NULL
                END AS "SegundosPazYSalvoCierreRRLL",

                -- Cierre de RRLL hasta gestión/finalización de Nómina.
                -- También se entrega en segundos para conservar la precisión.
                CASE
                    WHEN rl."FechaCierre" IS NOT NULL
                     AND rl."FechaEnvioNomina" IS NOT NULL
                     AND rl."FechaEnvioNomina" >= rl."FechaCierre"
                    THEN ROUND(
                        EXTRACT(
                            EPOCH FROM (
                                rl."FechaEnvioNomina" - rl."FechaCierre"
                            )
                        )
                    )::bigint
                    ELSE NULL
                END AS "SegundosCierreRRLLNomina",

                rp."IdEstadoProceso",
                ep."Nombre" AS "EstadoProceso",
                mr."Nombre" AS "MotivoRetiro",
                tr."Nombre" AS "TipificacionRetiro",
                rl."ObservacionGeneral",
                rl."ObservacionRetiro",
                rl."ObservacionNomina",
                rl."UsuarioObservacionNomina",
                rl."FechaObservacionNomina",

                CASE
                    WHEN rp."IdEstadoProceso" = 32
                      OR UPPER(
                            TRIM(
                                COALESCE(rl."EstadoCasoRRLL", '')
                            )
                         ) = 'ENVIADO_NOMINA'
                    THEN true
                    ELSE false
                END AS "PuedeGestionarNomina"

            FROM public."RetiroLaboral" rl

            INNER JOIN public."RegistroPersonal" rp
                ON rp."IdRegistroPersonal" = rl."IdRegistroPersonal"

            -- Se toma un solo Paz y Salvo por retiro.
            -- Esto evita que el mismo retiro aparezca repetido cuando
            -- existen varios registros asociados.
            LEFT JOIN LATERAL (
                SELECT
                    pso_ultimo."FechaCreacion"
                FROM public."PazYSalvoOperaciones" pso_ultimo
                WHERE pso_ultimo."IdRetiroLaboral" = rl."IdRetiroLaboral"
                ORDER BY
                    pso_ultimo."FechaCreacion" DESC NULLS LAST
                LIMIT 1
            ) pso ON true

            LEFT JOIN public."Cliente" c
                ON c."IdCliente" = rl."IdCliente"

            LEFT JOIN public."EstadoProceso" ep
                ON ep."IdEstadoProceso" = rp."IdEstadoProceso"

            LEFT JOIN public."MotivoRetiro" mr
                ON mr."IdMotivoRetiro" = rl."IdMotivoRetiro"

            LEFT JOIN public."TipificacionRetiro" tr
                ON tr."IdTipificacionRetiro" = rl."IdTipificacionRetiro"

            WHERE {FILTRO_NOMINA_SQL}

            ORDER BY
                CASE
                    WHEN rp."IdEstadoProceso" = 32
                      OR UPPER(
                            TRIM(
                                COALESCE(rl."EstadoCasoRRLL", '')
                            )
                         ) = 'ENVIADO_NOMINA'
                    THEN 0

                    WHEN rp."IdEstadoProceso" = 35
                    THEN 1

                    ELSE 2
                END,
                rl."FechaCreacion" DESC;
        """)

        rows = db.execute(query).mappings().all()

        return {
            "success": True,
            "message": "Retiros de nómina consultados correctamente.",
            "data": [dict(row) for row in rows],
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al consultar retiros de nómina: {str(e)}"
        )

@router.put("/{id_retiro_laboral}/observacion-nomina")
def guardar_observacion_nomina(
    id_retiro_laboral: int,
    payload: ObservacionNominaRequest,
    db: Session = Depends(get_db)
):
    try:
        query_retiro = text("""
            SELECT
                rl."IdRetiroLaboral",
                rl."EstadoCasoRRLL",
                rp."IdEstadoProceso"
            FROM public."RetiroLaboral" rl
            INNER JOIN public."RegistroPersonal" rp
                ON rp."IdRegistroPersonal" = rl."IdRegistroPersonal"
            WHERE rl."IdRetiroLaboral" = :id_retiro_laboral;
        """)

        retiro = db.execute(
            query_retiro,
            {"id_retiro_laboral": id_retiro_laboral}
        ).mappings().first()

        if not retiro:
            raise HTTPException(status_code=404, detail="Retiro laboral no encontrado.")

        estado_caso = str(retiro["EstadoCasoRRLL"] or "").upper()

        if int(retiro["IdEstadoProceso"] or 0) not in (32, 35) and estado_caso not in ("ENVIADO_NOMINA", "CERRADO"):
            raise HTTPException(
                status_code=400,
                detail="Solo se puede registrar observación de nómina en retiros enviados a nómina o retirados."
            )

        query_update = text("""
            UPDATE public."RetiroLaboral"
            SET
                "ObservacionNomina" = :observacion_nomina,
                "UsuarioObservacionNomina" = 'nomina',
                "FechaObservacionNomina" = NOW(),
                "FechaActualizacion" = NOW(),
                "UsuarioActualizacion" = 'nomina'
            WHERE "IdRetiroLaboral" = :id_retiro_laboral
            RETURNING
                "IdRetiroLaboral",
                "ObservacionNomina",
                "UsuarioObservacionNomina",
                "FechaObservacionNomina";
        """)

        actualizado = db.execute(query_update, {
            "id_retiro_laboral": id_retiro_laboral,
            "observacion_nomina": payload.observacion_nomina,
        }).mappings().first()

        db.commit()

        return {
            "success": True,
            "message": "Observación de nómina guardada correctamente.",
            "data": dict(actualizado),
        }

    except HTTPException:
        db.rollback()
        raise

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error guardando observación de nómina: {str(e)}"
        )


@router.get("/indicadores")
def obtener_indicadores_nomina_retiros(db: Session = Depends(get_db)):
    try:
        # ============================================================
        # 1. TOTALES POR ESTADO
        # ============================================================
        query_totales = text(f"""
            WITH clasificados AS (
                SELECT
                    rl."IdRetiroLaboral",

                    CASE
                        WHEN rp."IdEstadoProceso" = 35
                        THEN 'RETIRADO'

                        WHEN rp."IdEstadoProceso" = 32
                          OR UPPER(
                                TRIM(
                                    COALESCE(rl."EstadoCasoRRLL", '')
                                )
                             ) IN (
                                'ENVIADO_NOMINA',
                                'CERRADO'
                             )
                        THEN 'CERRADO'

                        WHEN UPPER(
                                TRIM(
                                    COALESCE(rl."EstadoCasoRRLL", '')
                                )
                             ) = 'ABIERTO'
                        THEN 'ABIERTO'

                        ELSE 'OTRO'
                    END AS grupo

                FROM public."RetiroLaboral" rl

                INNER JOIN public."RegistroPersonal" rp
                    ON rp."IdRegistroPersonal" = rl."IdRegistroPersonal"

                WHERE {FILTRO_NOMINA_SQL}
            )

            SELECT
                COUNT(*) FILTER (
                    WHERE grupo = 'CERRADO'
                )::int AS cerrados,

                COUNT(*) FILTER (
                    WHERE grupo = 'ABIERTO'
                )::int AS abiertos,

                COUNT(*) FILTER (
                    WHERE grupo = 'RETIRADO'
                )::int AS retirados,

                COUNT(*)::int AS total

            FROM clasificados;
        """)

        totales_row = db.execute(query_totales).mappings().first()

        abiertos = int(totales_row["abiertos"] or 0)
        cerrados = int(totales_row["cerrados"] or 0)
        retirados = int(totales_row["retirados"] or 0)
        total = int(totales_row["total"] or 0)

        # ============================================================
        # 2. RETIROS POR MES
        # ============================================================
        query_retiros_mes = text(f"""
            SELECT
                EXTRACT(YEAR FROM datos.fecha_base)::int AS anio,
                EXTRACT(MONTH FROM datos.fecha_base)::int AS mes_numero,
                COUNT(*)::int AS cantidad

            FROM (
                SELECT
                    rl."IdRetiroLaboral",

                    COALESCE(
                        rl."FechaRetiro",
                        pso."FechaCreacion",
                        rl."FechaCierre",
                        rl."FechaEnvioNomina",
                        rl."FechaPagoLiquidacion",
                        rl."FechaProceso",
                        rl."FechaCreacion"
                    ) AS fecha_base

                FROM public."RetiroLaboral" rl

                INNER JOIN public."RegistroPersonal" rp
                    ON rp."IdRegistroPersonal" = rl."IdRegistroPersonal"

                -- Se selecciona un único Paz y Salvo por retiro
                -- para impedir que el gráfico mensual duplique registros.
                LEFT JOIN LATERAL (
                    SELECT
                        pso_ultimo."FechaCreacion"
                    FROM public."PazYSalvoOperaciones" pso_ultimo
                    WHERE
                        pso_ultimo."IdRetiroLaboral" =
                        rl."IdRetiroLaboral"
                    ORDER BY
                        pso_ultimo."FechaCreacion" DESC NULLS LAST
                    LIMIT 1
                ) pso ON true

                WHERE {FILTRO_NOMINA_SQL}
            ) datos

            WHERE datos.fecha_base IS NOT NULL

            GROUP BY
                EXTRACT(YEAR FROM datos.fecha_base),
                EXTRACT(MONTH FROM datos.fecha_base)

            ORDER BY
                anio,
                mes_numero;
        """)

        meses_nombre = {
            1: "enero",
            2: "febrero",
            3: "marzo",
            4: "abril",
            5: "mayo",
            6: "junio",
            7: "julio",
            8: "agosto",
            9: "septiembre",
            10: "octubre",
            11: "noviembre",
            12: "diciembre",
        }

        retiros_por_mes_rows = (
            db.execute(query_retiros_mes)
            .mappings()
            .all()
        )

        retiros_por_mes = [
            {
                "mes": (
                    f"{meses_nombre.get(int(row['mes_numero']), 'sin-mes')}"
                    f"-{int(row['anio'])}"
                ),
                "cantidad": int(row["cantidad"] or 0),
            }
            for row in retiros_por_mes_rows
        ]

        # ============================================================
        # 3. PROMEDIOS DE TIEMPO
        # ============================================================
        query_promedios = text(f"""
            SELECT
                -- Último día laborado hasta Paz y Salvo:
                -- se conserva en días calendario.
                ROUND(
                    AVG(
                        CASE
                            WHEN rl."FechaRetiro" IS NOT NULL
                             AND pso."FechaCreacion" IS NOT NULL
                             AND DATE(pso."FechaCreacion") >=
                                 rl."FechaRetiro"
                            THEN
                                DATE(pso."FechaCreacion")
                                - rl."FechaRetiro"
                            ELSE NULL
                        END
                    ),
                    2
                ) AS promedio_paz_y_salvo_dias,

                -- Paz y Salvo hasta cierre de RRLL:
                -- resultado exacto en segundos.
                ROUND(
                    AVG(
                        CASE
                            WHEN pso."FechaCreacion" IS NOT NULL
                             AND rl."FechaCierre" IS NOT NULL
                             AND rl."FechaCierre" >=
                                 pso."FechaCreacion"
                            THEN EXTRACT(
                                EPOCH FROM (
                                    rl."FechaCierre"
                                    - pso."FechaCreacion"
                                )
                            )
                            ELSE NULL
                        END
                    ),
                    0
                ) AS promedio_rrll_segundos,

                -- Cierre de RRLL hasta finalización de Nómina:
                -- resultado exacto en segundos.
                ROUND(
                    AVG(
                        CASE
                            WHEN rl."FechaCierre" IS NOT NULL
                             AND rl."FechaEnvioNomina" IS NOT NULL
                             AND rl."FechaEnvioNomina" >=
                                 rl."FechaCierre"
                            THEN EXTRACT(
                                EPOCH FROM (
                                    rl."FechaEnvioNomina"
                                    - rl."FechaCierre"
                                )
                            )
                            ELSE NULL
                        END
                    ),
                    0
                ) AS promedio_nomina_segundos

            FROM public."RetiroLaboral" rl

            INNER JOIN public."RegistroPersonal" rp
                ON rp."IdRegistroPersonal" = rl."IdRegistroPersonal"

            -- Se usa un único Paz y Salvo por retiro para que
            -- cada proceso participe una sola vez en el promedio.
            LEFT JOIN LATERAL (
                SELECT
                    pso_ultimo."FechaCreacion"
                FROM public."PazYSalvoOperaciones" pso_ultimo
                WHERE
                    pso_ultimo."IdRetiroLaboral" =
                    rl."IdRetiroLaboral"
                ORDER BY
                    pso_ultimo."FechaCreacion" DESC NULLS LAST
                LIMIT 1
            ) pso ON true

            WHERE {FILTRO_NOMINA_SQL};
        """)

        promedios_row = (
            db.execute(query_promedios)
            .mappings()
            .first()
        )

        promedio_paz_y_salvo_dias = float(
            promedios_row["promedio_paz_y_salvo_dias"] or 0
        )

        promedio_rrll_segundos = int(
            promedios_row["promedio_rrll_segundos"] or 0
        )

        promedio_nomina_segundos = int(
            promedios_row["promedio_nomina_segundos"] or 0
        )

        # Estos valores en días se conservan temporalmente por
        # compatibilidad con el frontend actual.
        promedio_rrll_dias = round(
            promedio_rrll_segundos / 86400,
            2
        )

        promedio_nomina_dias = round(
            promedio_nomina_segundos / 86400,
            2
        )

        return {
            "success": True,
            "message": (
                "Indicadores de nómina retiros consultados "
                "correctamente."
            ),
            "data": {
                "totales": {
                    "abiertos": abiertos,
                    "cerrados": cerrados,
                    "retirados": retirados,
                    "total": total,
                },

                "promedios": {
                    # Valores anteriores conservados para no romper
                    # inmediatamente el frontend actual.
                    "pazYSalvo": promedio_paz_y_salvo_dias,
                    "rrll": promedio_rrll_dias,
                    "nomina": promedio_nomina_dias,
                    "operaciones": promedio_paz_y_salvo_dias,

                    # Nuevos valores exactos que utilizará el frontend.
                    "pazYSalvoDias": promedio_paz_y_salvo_dias,
                    "rrllSegundos": promedio_rrll_segundos,
                    "nominaSegundos": promedio_nomina_segundos,
                },

                "distribucionEstados": [
                    {
                        "estado": "Abierto",
                        "cantidad": abiertos,
                    },
                    {
                        "estado": "Cerrado",
                        "cantidad": cerrados,
                    },
                    {
                        "estado": "Retirado",
                        "cantidad": retirados,
                    },
                ],

                "retirosPorMes": retiros_por_mes,
            },
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=(
                "Error al consultar indicadores de nómina retiros: "
                f"{str(e)}"
            )
        )

@router.get("/reporte-excel")
def descargar_reporte_excel_nomina_retiros(db: Session = Depends(get_db)):
    try:
        rows = _consultar_retiros_nomina(db)

        def prioridad_excel(row):
            estado_caso = str(row.get("EstadoCasoRRLL") or "").upper()
            id_estado = int(row.get("IdEstadoProceso") or 0)

            if id_estado == 32 or estado_caso == "ENVIADO_NOMINA":
                return 1  # CERRADO

            if id_estado == 35:
                return 3  # RETIRADO

            return 2  # ABIERTO


        rows = sorted(
            rows,
            key=lambda row: (
                prioridad_excel(row),
                str(row.get("FechaRetiro") or ""),
                _nombre_completo(row),
            )
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Nomina Retiros"

        _configurar_hoja_detalle(ws, "Reporte Nómina Retiros", rows)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        nombre_archivo = f"Reporte_Nomina_Retiros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        headers = {
            "Content-Disposition": f'attachment; filename="{nombre_archivo}"'
        }

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error al generar reporte Excel de nómina retiros: {str(e)}"
        )


@router.put("/{id_retiro_laboral}/finalizar")
def finalizar_retiro_nomina(
    id_retiro_laboral: int,
    payload: FinalizarRetiroNominaRequest,
    db: Session = Depends(get_db)
):
    try:
        query_estado_retirado = text("""
            SELECT "IdEstadoProceso"
            FROM public."EstadoProceso"
            WHERE "Nombre" ILIKE 'Retirado'
              AND "Estado" = B'1'
            LIMIT 1;
        """)

        estado_retirado = db.execute(query_estado_retirado).mappings().first()

        if not estado_retirado:
            raise HTTPException(
                status_code=400,
                detail="No existe el estado Retirado activo en EstadoProceso."
            )

        id_estado_retirado = estado_retirado["IdEstadoProceso"]

        query_retiro = text("""
            SELECT
                rl."IdRetiroLaboral",
                rl."IdRegistroPersonal",
                rl."EstadoCasoRRLL",
                rp."IdEstadoProceso"
            FROM public."RetiroLaboral" rl
            INNER JOIN public."RegistroPersonal" rp
                ON rp."IdRegistroPersonal" = rl."IdRegistroPersonal"
            WHERE rl."IdRetiroLaboral" = :id_retiro_laboral;
        """)

        retiro = db.execute(
            query_retiro,
            {"id_retiro_laboral": id_retiro_laboral}
        ).mappings().first()

        if not retiro:
            raise HTTPException(status_code=404, detail="Retiro laboral no encontrado.")

        estado_caso = str(retiro["EstadoCasoRRLL"] or "").upper()

        if int(retiro["IdEstadoProceso"] or 0) != 32 and estado_caso != "ENVIADO_NOMINA":
            raise HTTPException(
                status_code=400,
                detail="Solo se pueden finalizar retiros enviados a nómina."
            )

        query_documentos_obligatorios = text("""
            SELECT
                tdr."IdTipoDocumentoRetiro",
                tdr."Nombre" AS "NombreDocumento",
                CASE
                    WHEN EXISTS (
                        SELECT 1
                        FROM public."RetiroLaboralAdjunto" rla
                        WHERE rla."IdRetiroLaboral" = :id_retiro_laboral
                          AND rla."IdTipoDocumentoRetiro" = tdr."IdTipoDocumentoRetiro"
                          AND COALESCE(rla."Activo", true) = true
                          AND COALESCE(rla."Eliminado", false) = false
                          AND rla."RutaArchivo" IS NOT NULL
                    ) THEN true
                    ELSE false
                END AS "Adjuntado"
            FROM public."TipoDocumentoRetiro" tdr
            -- ÚNICOS DOCUMENTOS OBLIGATORIOS PARA FINALIZAR:
            -- 15 = Retiro ARL
            -- 16 = Liquidación de contrato
            WHERE tdr."IdTipoDocumentoRetiro" IN (15, 16)
            ORDER BY tdr."IdTipoDocumentoRetiro";
        """)

        documentos_estado = db.execute(
            query_documentos_obligatorios,
            {"id_retiro_laboral": id_retiro_laboral}
        ).mappings().all()

        documentos_faltantes = [
            doc["NombreDocumento"]
            for doc in documentos_estado
            if not bool(doc["Adjuntado"])
        ]

        if documentos_faltantes:
            raise HTTPException(
                status_code=400,
                detail={
                    "message": "No se puede finalizar el retiro. Faltan documentos obligatorios de nómina.",
                    "documentos_faltantes": documentos_faltantes,
                }
            )

        query_update_retiro = text("""
            UPDATE public."RetiroLaboral"
            SET
                "EstadoCasoRRLL" = 'CERRADO',
                "Activo" = false,
                "FechaEnvioNomina" = NOW(),
                "FechaPagoLiquidacion" = :fecha_pago_liquidacion,
                "FechaActualizacion" = NOW(),
                "UsuarioActualizacion" = 'nomina'
            WHERE "IdRetiroLaboral" = :id_retiro_laboral;
        """)

        query_update_registro = text("""
            UPDATE public."RegistroPersonal"
            SET
                "IdEstadoProceso" = :id_estado_retirado,
                "FechaActualizacion" = NOW(),
                "UsuarioActualizacion" = 'nomina_contratacion'
            WHERE "IdRegistroPersonal" = :id_registro_personal;
        """)

        db.execute(
            query_update_retiro,
            {
                "id_retiro_laboral": id_retiro_laboral,
                "fecha_pago_liquidacion": payload.fecha_pago_liquidacion,
            }
        )

        db.execute(
            query_update_registro,
            {
                "id_estado_retirado": id_estado_retirado,
                "id_registro_personal": retiro["IdRegistroPersonal"],
            }
        )

        db.commit()

        return {
            "success": True,
            "message": "Retiro finalizado correctamente por nómina.",
            "data": {
                "IdRetiroLaboral": id_retiro_laboral,
                "IdRegistroPersonal": retiro["IdRegistroPersonal"],
                "IdEstadoProceso": id_estado_retirado,
                "EstadoProceso": "Retirado",
                "FechaPagoLiquidacion": payload.fecha_pago_liquidacion.isoformat(),
            },
        }

    except HTTPException:
        db.rollback()
        raise

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error al finalizar retiro desde nómina: {str(e)}"
        )

@router.put("/{id_retiro_laboral}/devolver")
def devolver_retiro_rrll(
    id_retiro_laboral: int,
    db: Session = Depends(get_db)
):
    try:
        query_retiro = text("""
            SELECT
                rl."IdRetiroLaboral",
                rl."IdRegistroPersonal",
                rl."EstadoCasoRRLL",
                rp."IdEstadoProceso"
            FROM public."RetiroLaboral" rl
            INNER JOIN public."RegistroPersonal" rp
                ON rp."IdRegistroPersonal" = rl."IdRegistroPersonal"
            WHERE rl."IdRetiroLaboral" = :id_retiro_laboral;
        """)

        retiro = db.execute(
            query_retiro,
            {"id_retiro_laboral": id_retiro_laboral}
        ).mappings().first()

        if not retiro:
            raise HTTPException(status_code=404, detail="Retiro laboral no encontrado.")

        estado_caso = str(retiro["EstadoCasoRRLL"] or "").upper()

        if int(retiro["IdEstadoProceso"] or 0) != 32 and estado_caso != "ENVIADO_NOMINA":
            raise HTTPException(
                status_code=400,
                detail="Solo se pueden devolver retiros enviados a nómina."
            )

        query_update_retiro = text("""
            UPDATE public."RetiroLaboral"
            SET
                "EstadoCasoRRLL" = 'ABIERTO',
                "Activo" = true,
                "FechaEnvioNomina" = NULL,
                "FechaCierre" = NULL,
                "FechaActualizacion" = NOW(),
                "UsuarioActualizacion" = 'nomina'
            WHERE "IdRetiroLaboral" = :id_retiro_laboral;
        """)

        query_update_registro = text("""
            UPDATE public."RegistroPersonal"
            SET
                "IdEstadoProceso" = 30
            WHERE "IdRegistroPersonal" = :id_registro_personal;
        """)

        db.execute(
            query_update_retiro,
            {"id_retiro_laboral": id_retiro_laboral}
        )

        db.execute(
            query_update_registro,
            {"id_registro_personal": retiro["IdRegistroPersonal"]}
        )

        db.commit()

        return {
            "success": True,
            "message": "Retiro devuelto correctamente a Relaciones Laborales.",
            "data": {
                "IdRetiroLaboral": id_retiro_laboral,
                "IdRegistroPersonal": retiro["IdRegistroPersonal"],
                "IdEstadoProceso": 30,
                "EstadoProceso": "Abierto",
            },
        }

    except HTTPException:
        db.rollback()
        raise

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error al devolver retiro a RRLL: {str(e)}"
        )
    
@router.get("/{id_retiro_laboral}/adjuntos")
def listar_adjuntos_nomina_retiro(
    id_retiro_laboral: int,
    db: Session = Depends(get_db)
):
    try:
        query = text("""
            WITH ultimo_adjunto AS (
                SELECT DISTINCT ON (rla."IdTipoDocumentoRetiro")
                    rla."IdRetiroLaboralAdjunto",
                    rla."IdRetiroLaboral",
                    rla."IdTipoDocumentoRetiro",
                    rla."NombreArchivo",
                    rla."NombreArchivoOriginal",
                    rla."RutaArchivo",
                    rla."ExtensionArchivo",
                    rla."PesoArchivo",
                    rla."OrigenArchivo",
                    rla."FechaCreacion",
                    rla."CreadoPor"
                FROM public."RetiroLaboralAdjunto" rla
                WHERE rla."IdRetiroLaboral" = :id_retiro_laboral
                  AND COALESCE(rla."Activo", true) = true
                  AND COALESCE(rla."Eliminado", false) = false
                ORDER BY
                    rla."IdTipoDocumentoRetiro",
                    rla."FechaCreacion" DESC,
                    rla."IdRetiroLaboralAdjunto" DESC
            )
            SELECT
                tdr."IdTipoDocumentoRetiro",
                tdr."Nombre" AS "NombreDocumento",
                ua."IdRetiroLaboralAdjunto",
                ua."IdRetiroLaboral",
                ua."NombreArchivo",
                ua."NombreArchivoOriginal",
                ua."RutaArchivo",
                ua."ExtensionArchivo",
                ua."PesoArchivo",
                ua."OrigenArchivo",
                ua."FechaCreacion",
                ua."CreadoPor",
                CASE
                    WHEN ua."IdRetiroLaboralAdjunto" IS NOT NULL THEN true
                    ELSE false
                END AS "Adjuntado"
            FROM public."TipoDocumentoRetiro" tdr
            LEFT JOIN ultimo_adjunto ua
                ON ua."IdTipoDocumentoRetiro" = tdr."IdTipoDocumentoRetiro"
            WHERE
                tdr."IdTipoDocumentoRetiro" IN (1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16)
                OR UPPER(TRIM(tdr."Nombre")) IN (
                    'SOPORTE NÓMINA',
                    'SOPORTE NOMINA',
                    'AUTORIZACIÓN DE DESCUENTO',
                    'AUTORIZACION DE DESCUENTO'
                )
            ORDER BY
                CASE
                    WHEN tdr."IdTipoDocumentoRetiro" BETWEEN 1 AND 16
                        THEN tdr."IdTipoDocumentoRetiro"
                    WHEN UPPER(TRIM(tdr."Nombre")) IN ('SOPORTE NÓMINA', 'SOPORTE NOMINA')
                        THEN 17
                    WHEN UPPER(TRIM(tdr."Nombre")) IN (
                        'AUTORIZACIÓN DE DESCUENTO',
                        'AUTORIZACION DE DESCUENTO'
                    )
                        THEN 18
                    ELSE 99
                END,
                tdr."IdTipoDocumentoRetiro";
        """)

        rows = db.execute(
            query,
            {"id_retiro_laboral": id_retiro_laboral}
        ).mappings().all()

        return {
            "success": True,
            "message": "Adjuntos del retiro consultados correctamente.",
            "data": [dict(row) for row in rows],
        }

    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error consultando adjuntos del retiro: {str(e)}"
        )

@router.post("/{id_retiro_laboral}/adjuntos")
async def subir_adjunto_nomina_retiro(
    id_retiro_laboral: int,
    IdTipoDocumentoRetiro: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    try:
        query_tipo_documento = text("""
            SELECT
                "IdTipoDocumentoRetiro",
                "Nombre"
            FROM public."TipoDocumentoRetiro"
            WHERE "IdTipoDocumentoRetiro" = :id_tipo_documento_retiro
            LIMIT 1;
        """)

        tipo_documento = db.execute(
            query_tipo_documento,
            {"id_tipo_documento_retiro": IdTipoDocumentoRetiro}
        ).mappings().first()

        if not tipo_documento:
            raise HTTPException(
                status_code=404,
                detail="El tipo de documento de retiro no existe."
            )

        nombre_tipo_documento = str(tipo_documento["Nombre"] or "").strip()
        nombre_normalizado = (
            nombre_tipo_documento
            .upper()
            .replace("Á", "A")
            .replace("É", "E")
            .replace("Í", "I")
            .replace("Ó", "O")
            .replace("Ú", "U")
        )

        documentos_permitidos = {
            "RETIRO ARL",
            "LIQUIDACION DE CONTRATO",
            "SOPORTE NOMINA",
            "AUTORIZACION DE DESCUENTO",
        }

        if nombre_normalizado not in documentos_permitidos:
            raise HTTPException(
                status_code=400,
                detail=(
                    "Nómina solo puede adjuntar Retiro ARL, Liquidación de contrato, "
                    "Soporte Nómina o Autorización de descuento."
                )
            )

        query_retiro = text("""
            SELECT
                rl."IdRetiroLaboral",
                rl."IdRegistroPersonal",
                rl."EstadoCasoRRLL",
                rp."IdEstadoProceso"
            FROM public."RetiroLaboral" rl
            INNER JOIN public."RegistroPersonal" rp
                ON rp."IdRegistroPersonal" = rl."IdRegistroPersonal"
            WHERE rl."IdRetiroLaboral" = :id_retiro_laboral;
        """)

        retiro = db.execute(
            query_retiro,
            {"id_retiro_laboral": id_retiro_laboral}
        ).mappings().first()

        if not retiro:
            raise HTTPException(status_code=404, detail="Retiro laboral no encontrado.")

        estado_caso = str(retiro["EstadoCasoRRLL"] or "").upper()

        if int(retiro["IdEstadoProceso"] or 0) not in (32, 35) and estado_caso not in ("ENVIADO_NOMINA", "CERRADO"):
            raise HTTPException(
                status_code=400,
                detail="Solo se pueden adjuntar documentos de nómina a retiros enviados a nómina o retirados."
            )

        contenido = await file.read()

        if not contenido:
            raise HTTPException(status_code=400, detail="El archivo está vacío.")

        extension = os.path.splitext(file.filename or "")[1].lower() or ".pdf"
        fecha_nombre = datetime.now().strftime("%Y%m%d_%H%M%S")

        nombre_archivo = (
            f"retiro_{id_retiro_laboral}_tipo_{IdTipoDocumentoRetiro}_{fecha_nombre}{extension}"
        )

        carpeta_destino = os.path.join(
            "storage",
            "rrll",
            "retiros",
            str(id_retiro_laboral)
        )

        os.makedirs(carpeta_destino, exist_ok=True)

        ruta_archivo = os.path.join(carpeta_destino, nombre_archivo)

        with open(ruta_archivo, "wb") as f:
            f.write(contenido)

        ruta_bd = ruta_archivo.replace("\\", "/")

        query_insert = text("""
            INSERT INTO public."RetiroLaboralAdjunto"
            (
                "IdRetiroLaboral",
                "IdTipoDocumentoRetiro",
                "NombreArchivo",
                "RutaArchivo",
                "ExtensionArchivo",
                "PesoArchivo",
                "Observacion",
                "Activo",
                "FechaCreacion",
                "FechaActualizacion",
                "CreadoPor",
                "UsuarioActualizacion",
                "OrigenArchivo",
                "MimeType",
                "NombreArchivoOriginal",
                "Eliminado"
            )
            VALUES
            (
                :id_retiro_laboral,
                :id_tipo_documento_retiro,
                :nombre_archivo,
                :ruta_archivo,
                :extension_archivo,
                :peso_archivo,
                :observacion,
                true,
                NOW(),
                NOW(),
                'nomina',
                'nomina',
                'NOMINA',
                :mime_type,
                :nombre_archivo_original,
                false
            )
            RETURNING "IdRetiroLaboralAdjunto";
        """)

        nuevo = db.execute(query_insert, {
            "id_retiro_laboral": id_retiro_laboral,
            "id_tipo_documento_retiro": IdTipoDocumentoRetiro,
            "nombre_archivo": nombre_archivo,
            "ruta_archivo": ruta_bd,
            "extension_archivo": extension,
            "peso_archivo": len(contenido),
            "observacion": "Carga desde módulo Nómina Retiros",
            "mime_type": file.content_type or "application/pdf",
            "nombre_archivo_original": file.filename or nombre_archivo,
        }).mappings().first()

        db.commit()

        return {
            "success": True,
            "message": "Documento de nómina adjuntado correctamente.",
            "data": {
                "IdRetiroLaboralAdjunto": nuevo["IdRetiroLaboralAdjunto"],
                "IdRetiroLaboral": id_retiro_laboral,
                "IdTipoDocumentoRetiro": IdTipoDocumentoRetiro,
                "NombreDocumento": nombre_tipo_documento,
                "NombreArchivo": nombre_archivo,
            },
        }

    except HTTPException:
        db.rollback()
        raise

    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Error al adjuntar documento de nómina: {str(e)}"
        )